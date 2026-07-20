import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
import sqlite3
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ============================================================
# BME-ERP - TESORERÍA INTEGRADA v1.0
# ============================================================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
RUTA_DB = os.path.join(BASE_DIR, "erp_cafe.db")

C_FONDO = "#EEF3F8"
C_AZUL = "#0F5C8E"
C_OSCURO = "#153B5B"
C_VERDE = "#15803D"
C_NARANJA = "#C56A00"
C_ROJO = "#B42318"
C_MORADO = "#7C3AED"
C_TEXTO = "#1F2937"
C_SUAVE = "#64748B"
C_BLANCO = "#FFFFFF"
C_BORDE = "#D7E0E8"


def conectar():
    con = sqlite3.connect(RUTA_DB, timeout=20)
    con.execute("PRAGMA foreign_keys = ON")
    return con


def columnas_tabla(con, tabla):
    return {
        fila[1]
        for fila in con.execute(
            f"PRAGMA table_info({tabla})"
        ).fetchall()
    }


def tabla_existe(con, tabla):
    return con.execute("""
        SELECT COUNT(*)
        FROM sqlite_master
        WHERE type='table' AND name=?
    """, (tabla,)).fetchone()[0] > 0


def agregar_columna(con, tabla, columna, definicion):
    if columna not in columnas_tabla(con, tabla):
        con.execute(
            f"ALTER TABLE {tabla} "
            f"ADD COLUMN {columna} {definicion}"
        )


def inicializar_bd():
    with conectar() as con:
        con.executescript("""
            CREATE TABLE IF NOT EXISTS bancos(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                banco TEXT,
                numero_cuenta TEXT,
                tipo_cuenta TEXT,
                titular TEXT,
                saldo REAL DEFAULT 0,
                estado TEXT DEFAULT 'ACTIVA'
            );

            CREATE TABLE IF NOT EXISTS movimientos_bancos(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                fecha TEXT,
                banco_id INTEGER,
                tipo TEXT,
                concepto TEXT,
                valor REAL,
                saldo_anterior REAL,
                saldo_nuevo REAL,
                autorizado_por TEXT
            );

            CREATE TABLE IF NOT EXISTS transferencias_bancarias(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                fecha TEXT,
                banco_origen TEXT,
                banco_destino TEXT,
                valor REAL,
                autorizado_por TEXT
            );

            CREATE TABLE IF NOT EXISTS conciliaciones_bancarias(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                fecha TEXT NOT NULL,
                banco_id INTEGER NOT NULL,
                saldo_erp REAL DEFAULT 0,
                saldo_extracto REAL DEFAULT 0,
                diferencia REAL DEFAULT 0,
                observaciones TEXT DEFAULT '',
                usuario TEXT DEFAULT '',
                estado TEXT DEFAULT 'CONCILIADA',
                creado_en TEXT DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS movimientos_tesoreria(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                fecha TEXT NOT NULL,
                tipo TEXT NOT NULL,
                categoria TEXT DEFAULT '',
                concepto TEXT NOT NULL,
                valor REAL NOT NULL,
                banco_id INTEGER,
                referencia TEXT DEFAULT '',
                origen TEXT DEFAULT 'TESORERÍA',
                estado TEXT DEFAULT 'ACTIVO',
                usuario TEXT DEFAULT '',
                creado_en TEXT DEFAULT CURRENT_TIMESTAMP,
                reversado_en TEXT DEFAULT '',
                motivo_reversion TEXT DEFAULT ''
            );

            CREATE TABLE IF NOT EXISTS auditoria_erp(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                fecha_hora TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                usuario TEXT,
                rol TEXT,
                accion TEXT NOT NULL,
                detalle TEXT,
                modulo TEXT
            );

            CREATE INDEX IF NOT EXISTS idx_mov_tes_fecha
            ON movimientos_tesoreria(fecha);

            CREATE INDEX IF NOT EXISTS idx_mov_tes_banco
            ON movimientos_tesoreria(banco_id);

            CREATE INDEX IF NOT EXISTS idx_conciliacion_banco
            ON conciliaciones_bancarias(banco_id);
        """)

        if tabla_existe(con, "movimientos_bancos"):
            for columna, definicion in [
                ("saldo_anterior", "REAL DEFAULT 0"),
                ("saldo_nuevo", "REAL DEFAULT 0"),
                ("autorizado_por", "TEXT DEFAULT ''")
            ]:
                agregar_columna(
                    con,
                    "movimientos_bancos",
                    columna,
                    definicion
                )

        con.commit()


def ahora():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def hoy():
    return datetime.now().strftime("%Y-%m-%d")


def periodo_actual():
    return datetime.now().strftime("%Y-%m")


def moneda(valor):
    return f"${float(valor or 0):,.0f}"


def numero(valor):
    return f"{float(valor or 0):,.2f}"


def a_numero(valor, nombre):
    try:
        n = float(
            str(valor).replace(",", "").strip()
        )
    except ValueError:
        raise ValueError(
            f"{nombre} debe ser numérico."
        )

    if n <= 0:
        raise ValueError(
            f"{nombre} debe ser mayor que cero."
        )

    return n


def extraer_id_combo(valor):
    try:
        return int(
            str(valor).split("|")[0].strip()
        )
    except Exception:
        return None


def auditoria(con, accion, detalle):
    con.execute("""
        INSERT INTO auditoria_erp(
            usuario, rol, accion, detalle, modulo
        )
        VALUES (?, ?, ?, ?, 'Tesorería')
    """, (
        os.environ.get(
            "ERP_USUARIO",
            "usuario_local"
        ),
        os.environ.get(
            "ERP_ROL",
            "OPERADOR"
        ),
        accion,
        detalle
    ))


def registrar_movimiento_banco(
    con,
    banco_id,
    tipo,
    concepto,
    valor,
    saldo_anterior,
    saldo_nuevo
):
    con.execute("""
        INSERT INTO movimientos_bancos(
            fecha,
            banco_id,
            tipo,
            concepto,
            valor,
            saldo_anterior,
            saldo_nuevo,
            autorizado_por
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        ahora(),
        banco_id,
        tipo,
        concepto,
        valor,
        saldo_anterior,
        saldo_nuevo,
        os.environ.get(
            "ERP_USUARIO",
            "usuario_local"
        )
    ))


def registrar_movimiento_tesoreria(
    con,
    tipo,
    categoria,
    concepto,
    valor,
    banco_id,
    referencia="",
    origen="TESORERÍA"
):
    con.execute("""
        INSERT INTO movimientos_tesoreria(
            fecha,
            tipo,
            categoria,
            concepto,
            valor,
            banco_id,
            referencia,
            origen,
            estado,
            usuario
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'ACTIVO', ?)
    """, (
        hoy(),
        tipo,
        categoria,
        concepto,
        valor,
        banco_id,
        referencia,
        origen,
        os.environ.get(
            "ERP_USUARIO",
            "usuario_local"
        )
    ))


# ============================================================
# CATÁLOGOS Y KPIs
# ============================================================

def cargar_bancos():
    with conectar() as con:
        filas = con.execute("""
            SELECT
                id,
                banco,
                COALESCE(numero_cuenta, ''),
                COALESCE(tipo_cuenta, ''),
                COALESCE(titular, ''),
                COALESCE(saldo, 0),
                COALESCE(estado, 'ACTIVA')
            FROM bancos
            ORDER BY banco, numero_cuenta
        """).fetchall()

    tabla_bancos.delete(
        *tabla_bancos.get_children()
    )

    activos = []

    for fila in filas:
        tabla_bancos.insert(
            "",
            "end",
            iid=str(fila[0]),
            values=(
                fila[1],
                fila[2],
                fila[3],
                fila[4],
                moneda(fila[5]),
                fila[6]
            ),
            tags=(
                "activo"
                if fila[6].upper() == "ACTIVA"
                else "inactivo"
            )
        )

        if fila[6].upper() == "ACTIVA":
            activos.append(
                (
                    f"{fila[0]} | {fila[1]} | "
                    f"{fila[2]} | {moneda(fila[5])}"
                )
            )

    for combo in (
        combo_mov_banco,
        combo_origen,
        combo_destino,
        combo_conc_banco
    ):
        combo["values"] = activos


def actualizar_kpis():
    periodo = entry_periodo.get().strip()

    with conectar() as con:
        saldo_bancos = float(
            con.execute("""
                SELECT IFNULL(SUM(saldo), 0)
                FROM bancos
                WHERE UPPER(estado)='ACTIVA'
            """).fetchone()[0] or 0
        )

        ingresos_mes = float(
            con.execute("""
                SELECT IFNULL(SUM(valor), 0)
                FROM movimientos_tesoreria
                WHERE tipo='INGRESO'
                  AND estado='ACTIVO'
                  AND substr(fecha,1,7)=?
            """, (periodo,)).fetchone()[0] or 0
        )

        egresos_mes = float(
            con.execute("""
                SELECT IFNULL(SUM(valor), 0)
                FROM movimientos_tesoreria
                WHERE tipo='EGRESO'
                  AND estado='ACTIVO'
                  AND substr(fecha,1,7)=?
            """, (periodo,)).fetchone()[0] or 0
        )

        cartera = float(
            con.execute("""
                SELECT IFNULL(SUM(saldo), 0)
                FROM cuentas_cobrar
                WHERE saldo>0
                  AND UPPER(estado) NOT IN (
                      'PAGADA',
                      'ANULADA'
                  )
            """).fetchone()[0] or 0
        ) if tabla_existe(
            con,
            "cuentas_cobrar"
        ) else 0

        cxp = float(
            con.execute("""
                SELECT IFNULL(SUM(saldo), 0)
                FROM cuentas_pagar
                WHERE saldo>0
                  AND UPPER(estado) <> 'PAGADA'
            """).fetchone()[0] or 0
        ) if tabla_existe(
            con,
            "cuentas_pagar"
        ) else 0

    flujo = ingresos_mes - egresos_mes
    liquidez_proyectada = (
        saldo_bancos
        + cartera
        - cxp
    )

    lbl_kpi_bancos.config(
        text=moneda(saldo_bancos)
    )
    lbl_kpi_ingresos.config(
        text=moneda(ingresos_mes)
    )
    lbl_kpi_egresos.config(
        text=moneda(egresos_mes)
    )
    lbl_kpi_flujo.config(
        text=moneda(flujo)
    )
    lbl_kpi_liquidez.config(
        text=moneda(liquidez_proyectada)
    )

    if liquidez_proyectada > 0 and flujo >= 0:
        estado = "VERDE"
        color = C_VERDE
    elif liquidez_proyectada > 0:
        estado = "AMARILLO"
        color = C_NARANJA
    else:
        estado = "ROJO"
        color = C_ROJO

    lbl_semaforo.config(
        text=f"SEMÁFORO FINANCIERO: {estado}",
        fg=color
    )


# ============================================================
# BANCOS
# ============================================================

def limpiar_banco():
    entry_banco.delete(0, "end")
    entry_cuenta.delete(0, "end")
    combo_tipo_cuenta.set("Ahorros")
    entry_titular.delete(0, "end")
    entry_saldo_inicial.delete(0, "end")


def guardar_banco():
    banco = entry_banco.get().strip()
    cuenta = entry_cuenta.get().strip()
    tipo = combo_tipo_cuenta.get().strip()
    titular = entry_titular.get().strip()

    if not banco:
        messagebox.showerror(
            "Banco",
            "Ingrese el nombre del banco."
        )
        return

    try:
        saldo = float(
            entry_saldo_inicial.get().strip() or 0
        )
    except ValueError:
        messagebox.showerror(
            "Banco",
            "El saldo inicial debe ser numérico."
        )
        return

    with conectar() as con:
        con.execute("""
            INSERT INTO bancos(
                banco,
                numero_cuenta,
                tipo_cuenta,
                titular,
                saldo,
                estado
            )
            VALUES (?, ?, ?, ?, ?, 'ACTIVA')
        """, (
            banco,
            cuenta,
            tipo,
            titular,
            saldo
        ))

        auditoria(
            con,
            "CREAR CUENTA BANCARIA",
            f"{banco}; cuenta {cuenta}; saldo {saldo:.2f}"
        )
        con.commit()

    limpiar_banco()
    cargar_bancos()
    actualizar_kpis()

    messagebox.showinfo(
        "Banco",
        "Cuenta bancaria creada correctamente."
    )


def cambiar_estado_banco():
    seleccion = tabla_bancos.selection()

    if not seleccion:
        messagebox.showwarning(
            "Banco",
            "Seleccione una cuenta bancaria."
        )
        return

    banco_id = int(seleccion[0])

    with conectar() as con:
        fila = con.execute("""
            SELECT banco, estado
            FROM bancos
            WHERE id=?
        """, (banco_id,)).fetchone()

        nuevo = (
            "INACTIVA"
            if fila[1].upper() == "ACTIVA"
            else "ACTIVA"
        )

        con.execute("""
            UPDATE bancos
            SET estado=?
            WHERE id=?
        """, (
            nuevo,
            banco_id
        ))

        auditoria(
            con,
            "CAMBIAR ESTADO BANCO",
            f"{fila[0]} a {nuevo}"
        )
        con.commit()

    cargar_bancos()
    actualizar_kpis()


# ============================================================
# MOVIMIENTOS MANUALES
# ============================================================

def registrar_movimiento_manual():
    banco_id = extraer_id_combo(
        combo_mov_banco.get()
    )
    tipo = combo_mov_tipo.get().strip()
    categoria = combo_mov_categoria.get().strip()
    concepto = entry_mov_concepto.get().strip()
    referencia = entry_mov_referencia.get().strip()

    if not banco_id:
        messagebox.showerror(
            "Movimiento",
            "Seleccione la cuenta bancaria."
        )
        return

    if tipo not in ("INGRESO", "EGRESO"):
        messagebox.showerror(
            "Movimiento",
            "Seleccione el tipo de movimiento."
        )
        return

    if not concepto:
        messagebox.showerror(
            "Movimiento",
            "Ingrese el concepto."
        )
        return

    try:
        valor = a_numero(
            entry_mov_valor.get(),
            "Valor"
        )
    except ValueError as error:
        messagebox.showerror(
            "Movimiento",
            str(error)
        )
        return

    con = conectar()

    try:
        con.execute("BEGIN IMMEDIATE")

        banco = con.execute("""
            SELECT banco, saldo
            FROM bancos
            WHERE id=?
              AND UPPER(estado)='ACTIVA'
        """, (banco_id,)).fetchone()

        if not banco:
            raise ValueError(
                "La cuenta bancaria no está activa."
            )

        saldo_anterior = float(
            banco[1] or 0
        )

        if tipo == "INGRESO":
            saldo_nuevo = saldo_anterior + valor
            tipo_banco = "CONSIGNACION"
        else:
            if valor > saldo_anterior:
                raise ValueError(
                    "Fondos insuficientes."
                )
            saldo_nuevo = saldo_anterior - valor
            tipo_banco = "RETIRO"

        con.execute("""
            UPDATE bancos
            SET saldo=?
            WHERE id=?
        """, (
            saldo_nuevo,
            banco_id
        ))

        registrar_movimiento_banco(
            con,
            banco_id,
            tipo_banco,
            concepto,
            valor,
            saldo_anterior,
            saldo_nuevo
        )

        registrar_movimiento_tesoreria(
            con,
            tipo,
            categoria,
            concepto,
            valor,
            banco_id,
            referencia,
            "TESORERÍA MANUAL"
        )

        auditoria(
            con,
            "REGISTRAR MOVIMIENTO TESORERÍA",
            (
                f"{tipo}; {concepto}; "
                f"valor {valor:.2f}"
            )
        )

        con.commit()

        messagebox.showinfo(
            "Movimiento",
            (
                "Movimiento registrado correctamente.\n\n"
                f"Nuevo saldo bancario: "
                f"{moneda(saldo_nuevo)}"
            )
        )

        entry_mov_concepto.delete(
            0,
            "end"
        )
        entry_mov_valor.delete(
            0,
            "end"
        )
        entry_mov_referencia.delete(
            0,
            "end"
        )

        cargar_bancos()
        cargar_movimientos()
        actualizar_kpis()

    except Exception as error:
        con.rollback()
        messagebox.showerror(
            "Movimiento",
            (
                "La operación fue revertida.\n\n"
                f"{error}"
            )
        )
    finally:
        con.close()


def cargar_movimientos():
    periodo = entry_periodo.get().strip()
    criterio = entry_buscar_mov.get().strip()

    sql = """
        SELECT
            m.id,
            m.fecha,
            b.banco,
            b.numero_cuenta,
            m.tipo,
            m.categoria,
            m.concepto,
            m.valor,
            m.referencia,
            m.origen,
            m.estado
        FROM movimientos_tesoreria m
        LEFT JOIN bancos b
          ON b.id=m.banco_id
        WHERE substr(m.fecha,1,7)=?
    """
    params = [periodo]

    if criterio:
        patron = f"%{criterio}%"
        sql += """
            AND (
                m.concepto LIKE ?
                OR m.referencia LIKE ?
                OR m.categoria LIKE ?
                OR b.banco LIKE ?
            )
        """
        params.extend([
            patron,
            patron,
            patron,
            patron
        ])

    sql += " ORDER BY m.id DESC"

    with conectar() as con:
        filas = con.execute(
            sql,
            params
        ).fetchall()

    tabla_movimientos.delete(
        *tabla_movimientos.get_children()
    )

    for fila in filas:
        tabla_movimientos.insert(
            "",
            "end",
            iid=str(fila[0]),
            values=(
                fila[1],
                f"{fila[2] or ''} / {fila[3] or ''}",
                fila[4],
                fila[5],
                fila[6],
                moneda(fila[7]),
                fila[8],
                fila[9],
                fila[10]
            ),
            tags=(
                "ingreso"
                if fila[4] == "INGRESO"
                else "egreso"
            )
        )


def reversar_movimiento():
    seleccion = tabla_movimientos.selection()

    if not seleccion:
        messagebox.showwarning(
            "Reversión",
            "Seleccione un movimiento."
        )
        return

    movimiento_id = int(seleccion[0])

    with conectar() as con:
        fila = con.execute("""
            SELECT
                tipo,
                concepto,
                valor,
                banco_id,
                estado
            FROM movimientos_tesoreria
            WHERE id=?
        """, (movimiento_id,)).fetchone()

    if not fila:
        return

    if fila[4] == "REVERSADO":
        messagebox.showinfo(
            "Reversión",
            "El movimiento ya fue reversado."
        )
        return

    motivo = simpledialog.askstring(
        "Motivo de reversión",
        "Indique el motivo:",
        parent=ventana
    )

    if not motivo:
        return

    con = conectar()

    try:
        con.execute("BEGIN IMMEDIATE")

        banco = con.execute("""
            SELECT saldo
            FROM bancos
            WHERE id=?
        """, (fila[3],)).fetchone()

        saldo_anterior = float(
            banco[0] or 0
        )

        if fila[0] == "INGRESO":
            if fila[2] > saldo_anterior:
                raise ValueError(
                    "El banco no tiene saldo suficiente "
                    "para reversar el ingreso."
                )
            saldo_nuevo = (
                saldo_anterior
                - float(fila[2])
            )
            tipo_banco = "REVERSIÓN INGRESO"
        else:
            saldo_nuevo = (
                saldo_anterior
                + float(fila[2])
            )
            tipo_banco = "REVERSIÓN EGRESO"

        con.execute("""
            UPDATE bancos
            SET saldo=?
            WHERE id=?
        """, (
            saldo_nuevo,
            fila[3]
        ))

        con.execute("""
            UPDATE movimientos_tesoreria
            SET estado='REVERSADO',
                reversado_en=?,
                motivo_reversion=?
            WHERE id=?
        """, (
            ahora(),
            motivo,
            movimiento_id
        ))

        registrar_movimiento_banco(
            con,
            fila[3],
            tipo_banco,
            f"Reversión: {fila[1]}",
            fila[2],
            saldo_anterior,
            saldo_nuevo
        )

        auditoria(
            con,
            "REVERSAR MOVIMIENTO TESORERÍA",
            (
                f"Movimiento {movimiento_id}; "
                f"motivo {motivo}"
            )
        )

        con.commit()

        messagebox.showinfo(
            "Reversión",
            "Movimiento reversado correctamente."
        )

        cargar_bancos()
        cargar_movimientos()
        actualizar_kpis()

    except Exception as error:
        con.rollback()
        messagebox.showerror(
            "Reversión",
            (
                "La operación fue revertida.\n\n"
                f"{error}"
            )
        )
    finally:
        con.close()


# ============================================================
# TRANSFERENCIAS
# ============================================================

def realizar_transferencia():
    origen_id = extraer_id_combo(
        combo_origen.get()
    )
    destino_id = extraer_id_combo(
        combo_destino.get()
    )

    if not origen_id or not destino_id:
        messagebox.showerror(
            "Transferencia",
            "Seleccione banco origen y destino."
        )
        return

    if origen_id == destino_id:
        messagebox.showerror(
            "Transferencia",
            "Seleccione cuentas diferentes."
        )
        return

    try:
        valor = a_numero(
            entry_transferencia_valor.get(),
            "Valor"
        )
    except ValueError as error:
        messagebox.showerror(
            "Transferencia",
            str(error)
        )
        return

    concepto = (
        entry_transferencia_concepto.get().strip()
        or "Transferencia interna"
    )

    con = conectar()

    try:
        con.execute("BEGIN IMMEDIATE")

        origen = con.execute("""
            SELECT banco, numero_cuenta, saldo
            FROM bancos
            WHERE id=? AND UPPER(estado)='ACTIVA'
        """, (origen_id,)).fetchone()

        destino = con.execute("""
            SELECT banco, numero_cuenta, saldo
            FROM bancos
            WHERE id=? AND UPPER(estado)='ACTIVA'
        """, (destino_id,)).fetchone()

        if not origen or not destino:
            raise ValueError(
                "Una de las cuentas no está activa."
            )

        saldo_origen = float(
            origen[2] or 0
        )
        saldo_destino = float(
            destino[2] or 0
        )

        if valor > saldo_origen:
            raise ValueError(
                "Saldo insuficiente en la cuenta origen."
            )

        nuevo_origen = saldo_origen - valor
        nuevo_destino = saldo_destino + valor

        con.execute("""
            UPDATE bancos
            SET saldo=?
            WHERE id=?
        """, (
            nuevo_origen,
            origen_id
        ))

        con.execute("""
            UPDATE bancos
            SET saldo=?
            WHERE id=?
        """, (
            nuevo_destino,
            destino_id
        ))

        con.execute("""
            INSERT INTO transferencias_bancarias(
                fecha,
                banco_origen,
                banco_destino,
                valor,
                autorizado_por
            )
            VALUES (?, ?, ?, ?, ?)
        """, (
            ahora(),
            f"{origen[0]} / {origen[1]}",
            f"{destino[0]} / {destino[1]}",
            valor,
            os.environ.get(
                "ERP_USUARIO",
                "usuario_local"
            )
        ))

        registrar_movimiento_banco(
            con,
            origen_id,
            "TRANSFERENCIA SALIENTE",
            (
                f"{concepto} a "
                f"{destino[0]} / {destino[1]}"
            ),
            valor,
            saldo_origen,
            nuevo_origen
        )

        registrar_movimiento_banco(
            con,
            destino_id,
            "TRANSFERENCIA ENTRANTE",
            (
                f"{concepto} desde "
                f"{origen[0]} / {origen[1]}"
            ),
            valor,
            saldo_destino,
            nuevo_destino
        )

        auditoria(
            con,
            "TRANSFERENCIA BANCARIA",
            (
                f"De {origen[0]} a {destino[0]}; "
                f"valor {valor:.2f}"
            )
        )

        con.commit()

        messagebox.showinfo(
            "Transferencia",
            "Transferencia realizada correctamente."
        )

        entry_transferencia_valor.delete(
            0,
            "end"
        )
        entry_transferencia_concepto.delete(
            0,
            "end"
        )

        cargar_bancos()
        cargar_transferencias()
        actualizar_kpis()

    except Exception as error:
        con.rollback()
        messagebox.showerror(
            "Transferencia",
            (
                "La operación fue revertida.\n\n"
                f"{error}"
            )
        )
    finally:
        con.close()


def cargar_transferencias():
    with conectar() as con:
        filas = con.execute("""
            SELECT
                id,
                fecha,
                banco_origen,
                banco_destino,
                valor,
                autorizado_por
            FROM transferencias_bancarias
            ORDER BY id DESC
        """).fetchall()

    tabla_transferencias.delete(
        *tabla_transferencias.get_children()
    )

    for fila in filas:
        tabla_transferencias.insert(
            "",
            "end",
            values=(
                fila[1],
                fila[2],
                fila[3],
                moneda(fila[4]),
                fila[5]
            )
        )


# ============================================================
# CONCILIACIÓN
# ============================================================

def cargar_saldo_conciliacion(evento=None):
    banco_id = extraer_id_combo(
        combo_conc_banco.get()
    )

    if not banco_id:
        lbl_saldo_erp.config(
            text="$0"
        )
        return

    with conectar() as con:
        saldo = con.execute("""
            SELECT saldo
            FROM bancos
            WHERE id=?
        """, (banco_id,)).fetchone()

    lbl_saldo_erp.config(
        text=moneda(saldo[0] if saldo else 0)
    )
    calcular_diferencia()


def calcular_diferencia(evento=None):
    banco_id = extraer_id_combo(
        combo_conc_banco.get()
    )

    if not banco_id:
        return

    with conectar() as con:
        fila = con.execute("""
            SELECT saldo
            FROM bancos
            WHERE id=?
        """, (banco_id,)).fetchone()

    saldo_erp = float(
        fila[0] or 0
    ) if fila else 0

    try:
        saldo_extracto = float(
            entry_saldo_extracto.get().replace(
                ",",
                ""
            ).strip() or 0
        )
    except ValueError:
        saldo_extracto = 0

    diferencia = saldo_extracto - saldo_erp

    lbl_diferencia.config(
        text=moneda(diferencia),
        fg=(
            C_VERDE
            if abs(diferencia) < 0.01
            else C_ROJO
        )
    )


def guardar_conciliacion():
    banco_id = extraer_id_combo(
        combo_conc_banco.get()
    )

    if not banco_id:
        messagebox.showerror(
            "Conciliación",
            "Seleccione la cuenta bancaria."
        )
        return

    try:
        saldo_extracto = float(
            entry_saldo_extracto.get().replace(
                ",",
                ""
            ).strip()
        )
    except ValueError:
        messagebox.showerror(
            "Conciliación",
            "Ingrese un saldo de extracto válido."
        )
        return

    with conectar() as con:
        fila = con.execute("""
            SELECT saldo
            FROM bancos
            WHERE id=?
        """, (banco_id,)).fetchone()

        saldo_erp = float(
            fila[0] or 0
        )
        diferencia = saldo_extracto - saldo_erp

        con.execute("""
            INSERT INTO conciliaciones_bancarias(
                fecha,
                banco_id,
                saldo_erp,
                saldo_extracto,
                diferencia,
                observaciones,
                usuario,
                estado
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            hoy(),
            banco_id,
            saldo_erp,
            saldo_extracto,
            diferencia,
            txt_conc_obs.get(
                "1.0",
                "end"
            ).strip(),
            os.environ.get(
                "ERP_USUARIO",
                "usuario_local"
            ),
            (
                "CONCILIADA"
                if abs(diferencia) < 0.01
                else "CON DIFERENCIA"
            )
        ))

        auditoria(
            con,
            "CONCILIACIÓN BANCARIA",
            (
                f"Banco {banco_id}; "
                f"diferencia {diferencia:.2f}"
            )
        )
        con.commit()

    messagebox.showinfo(
        "Conciliación",
        "Conciliación registrada correctamente."
    )

    entry_saldo_extracto.delete(
        0,
        "end"
    )
    txt_conc_obs.delete(
        "1.0",
        "end"
    )

    cargar_conciliaciones()


def cargar_conciliaciones():
    with conectar() as con:
        filas = con.execute("""
            SELECT
                c.fecha,
                b.banco,
                b.numero_cuenta,
                c.saldo_erp,
                c.saldo_extracto,
                c.diferencia,
                c.estado
            FROM conciliaciones_bancarias c
            JOIN bancos b
              ON b.id=c.banco_id
            ORDER BY c.id DESC
        """).fetchall()

    tabla_conciliaciones.delete(
        *tabla_conciliaciones.get_children()
    )

    for fila in filas:
        tabla_conciliaciones.insert(
            "",
            "end",
            values=(
                fila[0],
                f"{fila[1]} / {fila[2]}",
                moneda(fila[3]),
                moneda(fila[4]),
                moneda(fila[5]),
                fila[6]
            ),
            tags=(
                "ok"
                if fila[6] == "CONCILIADA"
                else "diferencia"
            )
        )


# ============================================================
# FLUJO DE CAJA
# ============================================================

def cargar_flujo_caja():
    periodo = entry_periodo.get().strip()

    with conectar() as con:
        ingresos = con.execute("""
            SELECT
                categoria,
                IFNULL(SUM(valor), 0)
            FROM movimientos_tesoreria
            WHERE tipo='INGRESO'
              AND estado='ACTIVO'
              AND substr(fecha,1,7)=?
            GROUP BY categoria
            ORDER BY categoria
        """, (periodo,)).fetchall()

        egresos = con.execute("""
            SELECT
                categoria,
                IFNULL(SUM(valor), 0)
            FROM movimientos_tesoreria
            WHERE tipo='EGRESO'
              AND estado='ACTIVO'
              AND substr(fecha,1,7)=?
            GROUP BY categoria
            ORDER BY categoria
        """, (periodo,)).fetchall()

        saldo_bancos = float(
            con.execute("""
                SELECT IFNULL(SUM(saldo), 0)
                FROM bancos
                WHERE UPPER(estado)='ACTIVA'
            """).fetchone()[0] or 0
        )

        cxc = float(
            con.execute("""
                SELECT IFNULL(SUM(saldo), 0)
                FROM cuentas_cobrar
                WHERE saldo>0
                  AND UPPER(estado) NOT IN (
                      'PAGADA',
                      'ANULADA'
                  )
            """).fetchone()[0] or 0
        ) if tabla_existe(
            con,
            "cuentas_cobrar"
        ) else 0

        cxp = float(
            con.execute("""
                SELECT IFNULL(SUM(saldo), 0)
                FROM cuentas_pagar
                WHERE saldo>0
                  AND UPPER(estado) <> 'PAGADA'
            """).fetchone()[0] or 0
        ) if tabla_existe(
            con,
            "cuentas_pagar"
        ) else 0

    tabla_flujo.delete(
        *tabla_flujo.get_children()
    )

    total_ingresos = 0
    total_egresos = 0

    for categoria, valor in ingresos:
        total_ingresos += float(
            valor or 0
        )
        tabla_flujo.insert(
            "",
            "end",
            values=(
                "INGRESO",
                categoria,
                moneda(valor)
            ),
            tags=("ingreso",)
        )

    for categoria, valor in egresos:
        total_egresos += float(
            valor or 0
        )
        tabla_flujo.insert(
            "",
            "end",
            values=(
                "EGRESO",
                categoria,
                moneda(valor)
            ),
            tags=("egreso",)
        )

    flujo = total_ingresos - total_egresos
    proyeccion = saldo_bancos + cxc - cxp

    lbl_flujo_ingresos.config(
        text=moneda(total_ingresos)
    )
    lbl_flujo_egresos.config(
        text=moneda(total_egresos)
    )
    lbl_flujo_neto.config(
        text=moneda(flujo)
    )
    lbl_flujo_proyectado.config(
        text=moneda(proyeccion)
    )


# ============================================================
# IMPORTAR MOVIMIENTOS EXISTENTES
# ============================================================

def sincronizar_movimientos_existentes():
    with conectar() as con:
        # Recaudos de cartera
        if tabla_existe(
            con,
            "recaudos_cartera"
        ):
            filas = con.execute("""
                SELECT
                    r.id,
                    r.fecha,
                    r.banco_id,
                    r.valor,
                    r.referencia,
                    c.cliente
                FROM recaudos_cartera r
                JOIN cuentas_cobrar c
                  ON c.id=r.cuenta_id
                WHERE r.estado='ACTIVO'
            """).fetchall()

            for fila in filas:
                referencia = (
                    f"CXC-{fila[0]}"
                )

                existe = con.execute("""
                    SELECT id
                    FROM movimientos_tesoreria
                    WHERE referencia=?
                      AND origen='CUENTAS POR COBRAR'
                """, (referencia,)).fetchone()

                if not existe:
                    registrar_movimiento_tesoreria(
                        con,
                        "INGRESO",
                        "RECAUDOS",
                        f"Recaudo de {fila[5]}",
                        fila[3],
                        fila[2],
                        referencia,
                        "CUENTAS POR COBRAR"
                    )

        # Pagos CxP
        if (
            tabla_existe(
                con,
                "pagos_cxp"
            )
            and tabla_existe(
                con,
                "cuentas_pagar"
            )
        ):
            filas = con.execute("""
                SELECT
                    p.id,
                    p.fecha,
                    p.valor,
                    c.proveedor
                FROM pagos_cxp p
                JOIN cuentas_pagar c
                  ON c.id=p.cuenta_id
            """).fetchall()

            for fila in filas:
                referencia = (
                    f"CXP-{fila[0]}"
                )

                existe = con.execute("""
                    SELECT id
                    FROM movimientos_tesoreria
                    WHERE referencia=?
                      AND origen='CUENTAS POR PAGAR'
                """, (referencia,)).fetchone()

                if not existe:
                    registrar_movimiento_tesoreria(
                        con,
                        "EGRESO",
                        "PAGO PROVEEDORES",
                        f"Pago a {fila[3]}",
                        fila[2],
                        None,
                        referencia,
                        "CUENTAS POR PAGAR"
                    )

        # Ventas de contado
        if tabla_existe(
            con,
            "ventas_integradas"
        ):
            filas = con.execute("""
                SELECT
                    id,
                    fecha,
                    banco_id,
                    total,
                    cliente,
                    numero
                FROM ventas_integradas
                WHERE forma_pago='CONTADO'
                  AND estado='ACTIVA'
            """).fetchall()

            for fila in filas:
                referencia = (
                    f"VTA-{fila[0]}"
                )

                existe = con.execute("""
                    SELECT id
                    FROM movimientos_tesoreria
                    WHERE referencia=?
                      AND origen='VENTAS'
                """, (referencia,)).fetchone()

                if not existe:
                    registrar_movimiento_tesoreria(
                        con,
                        "INGRESO",
                        "VENTAS CONTADO",
                        f"Venta a {fila[4]}",
                        fila[3],
                        fila[2],
                        referencia,
                        "VENTAS"
                    )

        con.commit()


# ============================================================
# REPORTES
# ============================================================

def exportar_excel():
    carpeta = os.path.join(
        BASE_DIR,
        "reportes"
    )
    os.makedirs(
        carpeta,
        exist_ok=True
    )

    ruta = os.path.join(
        carpeta,
        (
            "tesoreria_"
            + datetime.now().strftime(
                "%Y%m%d_%H%M%S"
            )
            + ".xlsx"
        )
    )

    with conectar() as con:
        bancos = con.execute("""
            SELECT
                banco,
                numero_cuenta,
                tipo_cuenta,
                titular,
                saldo,
                estado
            FROM bancos
            ORDER BY banco
        """).fetchall()

        movimientos = con.execute("""
            SELECT
                fecha,
                tipo,
                categoria,
                concepto,
                valor,
                referencia,
                origen,
                estado
            FROM movimientos_tesoreria
            ORDER BY id DESC
        """).fetchall()

        transferencias = con.execute("""
            SELECT
                fecha,
                banco_origen,
                banco_destino,
                valor,
                autorizado_por
            FROM transferencias_bancarias
            ORDER BY id DESC
        """).fetchall()

        conciliaciones = con.execute("""
            SELECT
                c.fecha,
                b.banco,
                b.numero_cuenta,
                c.saldo_erp,
                c.saldo_extracto,
                c.diferencia,
                c.estado
            FROM conciliaciones_bancarias c
            JOIN bancos b
              ON b.id=c.banco_id
            ORDER BY c.id DESC
        """).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Bancos"

    ws.append([
        "Banco",
        "Cuenta",
        "Tipo",
        "Titular",
        "Saldo",
        "Estado"
    ])

    for celda in ws[1]:
        celda.font = Font(
            bold=True,
            color="FFFFFF"
        )
        celda.fill = PatternFill(
            "solid",
            fgColor="153B5B"
        )
        celda.alignment = Alignment(
            horizontal="center"
        )

    for fila in bancos:
        ws.append(list(fila))

    ws2 = wb.create_sheet(
        "Movimientos"
    )
    ws2.append([
        "Fecha",
        "Tipo",
        "Categoría",
        "Concepto",
        "Valor",
        "Referencia",
        "Origen",
        "Estado"
    ])

    for celda in ws2[1]:
        celda.font = Font(
            bold=True,
            color="FFFFFF"
        )
        celda.fill = PatternFill(
            "solid",
            fgColor="0F5C8E"
        )

    for fila in movimientos:
        ws2.append(list(fila))

    ws3 = wb.create_sheet(
        "Transferencias"
    )
    ws3.append([
        "Fecha",
        "Origen",
        "Destino",
        "Valor",
        "Autorizado por"
    ])

    for fila in transferencias:
        ws3.append(list(fila))

    ws4 = wb.create_sheet(
        "Conciliaciones"
    )
    ws4.append([
        "Fecha",
        "Banco",
        "Cuenta",
        "Saldo ERP",
        "Saldo extracto",
        "Diferencia",
        "Estado"
    ])

    for fila in conciliaciones:
        ws4.append(list(fila))

    wb.save(ruta)

    with conectar() as con:
        auditoria(
            con,
            "EXPORTAR TESORERÍA",
            ruta
        )
        con.commit()

    messagebox.showinfo(
        "Exportación",
        (
            "Reporte generado correctamente:\n\n"
            f"{ruta}"
        )
    )


def refrescar_todo():
    sincronizar_movimientos_existentes()
    cargar_bancos()
    cargar_movimientos()
    cargar_transferencias()
    cargar_conciliaciones()
    cargar_flujo_caja()
    actualizar_kpis()


# ============================================================
# INTERFAZ
# ============================================================

inicializar_bd()

ventana = tk.Tk()
ventana.title(
    "BME-ERP - Tesorería Integrada v1.0"
)
ventana.geometry("1500x900")
ventana.minsize(1180, 720)
ventana.configure(bg=C_FONDO)

try:
    ventana.state("zoomed")
except tk.TclError:
    pass

estilo = ttk.Style()
try:
    estilo.theme_use("clam")
except tk.TclError:
    pass

estilo.configure(
    "Treeview",
    rowheight=28,
    font=("Segoe UI", 9)
)
estilo.configure(
    "Treeview.Heading",
    font=("Segoe UI", 9, "bold")
)

header = tk.Frame(
    ventana,
    bg=C_OSCURO,
    height=82
)
header.pack(fill="x")
header.pack_propagate(False)

tk.Label(
    header,
    text="TESORERÍA INTEGRADA",
    font=("Segoe UI", 22, "bold"),
    bg=C_OSCURO,
    fg="white"
).pack(
    side="left",
    padx=24,
    pady=20
)

tk.Label(
    header,
    text=(
        "Bancos, flujo de caja, movimientos, "
        "transferencias y conciliación"
    ),
    font=("Segoe UI", 10),
    bg=C_OSCURO,
    fg="#BFDBFE"
).pack(
    side="right",
    padx=24
)

barra_periodo = tk.Frame(
    ventana,
    bg=C_FONDO
)
barra_periodo.pack(
    fill="x",
    padx=18,
    pady=(12, 0)
)

tk.Label(
    barra_periodo,
    text="Periodo:",
    bg=C_FONDO,
    fg=C_TEXTO,
    font=("Segoe UI", 9, "bold")
).pack(side="left")

entry_periodo = ttk.Entry(
    barra_periodo,
    width=10
)
entry_periodo.pack(
    side="left",
    padx=6
)
entry_periodo.insert(
    0,
    periodo_actual()
)

tk.Button(
    barra_periodo,
    text="Actualizar todo",
    command=refrescar_todo,
    bg=C_AZUL,
    fg="white",
    relief="flat",
    padx=14,
    pady=6
).pack(
    side="left",
    padx=5
)

tk.Button(
    barra_periodo,
    text="Exportar Excel",
    command=exportar_excel,
    bg=C_VERDE,
    fg="white",
    relief="flat",
    padx=14,
    pady=6
).pack(
    side="right"
)

lbl_semaforo = tk.Label(
    barra_periodo,
    text="SEMÁFORO FINANCIERO",
    bg=C_FONDO,
    fg=C_SUAVE,
    font=("Segoe UI", 10, "bold")
)
lbl_semaforo.pack(
    side="right",
    padx=20
)

panel_kpi = tk.Frame(
    ventana,
    bg=C_FONDO
)
panel_kpi.pack(
    fill="x",
    padx=18,
    pady=(8, 5)
)

for columna in range(5):
    panel_kpi.grid_columnconfigure(
        columna,
        weight=1
    )


def crear_kpi(
    columna,
    titulo,
    color
):
    marco = tk.Frame(
        panel_kpi,
        bg=C_BLANCO,
        highlightbackground=C_BORDE,
        highlightthickness=1
    )
    marco.grid(
        row=0,
        column=columna,
        sticky="ew",
        padx=5
    )

    tk.Frame(
        marco,
        bg=color,
        width=5
    ).pack(
        side="left",
        fill="y"
    )

    interior = tk.Frame(
        marco,
        bg=C_BLANCO
    )
    interior.pack(
        fill="both",
        expand=True,
        padx=12,
        pady=9
    )

    tk.Label(
        interior,
        text=titulo,
        bg=C_BLANCO,
        fg=C_SUAVE,
        font=("Segoe UI", 8, "bold")
    ).pack(anchor="w")

    valor = tk.Label(
        interior,
        text="$0",
        bg=C_BLANCO,
        fg=C_TEXTO,
        font=("Segoe UI", 14, "bold")
    )
    valor.pack(
        anchor="w",
        pady=(3, 0)
    )

    return valor


lbl_kpi_bancos = crear_kpi(
    0,
    "SALDO BANCOS",
    C_AZUL
)
lbl_kpi_ingresos = crear_kpi(
    1,
    "INGRESOS DEL MES",
    C_VERDE
)
lbl_kpi_egresos = crear_kpi(
    2,
    "EGRESOS DEL MES",
    C_ROJO
)
lbl_kpi_flujo = crear_kpi(
    3,
    "FLUJO NETO",
    C_MORADO
)
lbl_kpi_liquidez = crear_kpi(
    4,
    "LIQUIDEZ PROYECTADA",
    C_NARANJA
)

notebook = ttk.Notebook(ventana)
notebook.pack(
    fill="both",
    expand=True,
    padx=18,
    pady=10
)

tab_resumen = tk.Frame(
    notebook,
    bg=C_FONDO
)
tab_bancos = tk.Frame(
    notebook,
    bg=C_FONDO
)
tab_movimientos = tk.Frame(
    notebook,
    bg=C_FONDO
)
tab_transferencias = tk.Frame(
    notebook,
    bg=C_FONDO
)
tab_conciliacion = tk.Frame(
    notebook,
    bg=C_FONDO
)

notebook.add(
    tab_resumen,
    text="  Flujo de caja  "
)
notebook.add(
    tab_bancos,
    text="  Bancos  "
)
notebook.add(
    tab_movimientos,
    text="  Movimientos  "
)
notebook.add(
    tab_transferencias,
    text="  Transferencias  "
)
notebook.add(
    tab_conciliacion,
    text="  Conciliación  "
)

# FLUJO
resumen_flujo = tk.Frame(
    tab_resumen,
    bg=C_BLANCO
)
resumen_flujo.pack(
    fill="x",
    padx=10,
    pady=10
)

for columna in range(4):
    resumen_flujo.grid_columnconfigure(
        columna,
        weight=1
    )

for indice, titulo in enumerate([
    "Ingresos",
    "Egresos",
    "Flujo neto",
    "Saldo proyectado"
]):
    tk.Label(
        resumen_flujo,
        text=titulo,
        bg=C_BLANCO,
        fg=C_SUAVE,
        font=("Segoe UI", 8, "bold")
    ).grid(
        row=0,
        column=indice,
        pady=(10, 2)
    )

lbl_flujo_ingresos = tk.Label(
    resumen_flujo,
    text="$0",
    bg=C_BLANCO,
    fg=C_VERDE,
    font=("Segoe UI", 16, "bold")
)
lbl_flujo_ingresos.grid(
    row=1,
    column=0,
    pady=(0, 10)
)

lbl_flujo_egresos = tk.Label(
    resumen_flujo,
    text="$0",
    bg=C_BLANCO,
    fg=C_ROJO,
    font=("Segoe UI", 16, "bold")
)
lbl_flujo_egresos.grid(
    row=1,
    column=1,
    pady=(0, 10)
)

lbl_flujo_neto = tk.Label(
    resumen_flujo,
    text="$0",
    bg=C_BLANCO,
    fg=C_AZUL,
    font=("Segoe UI", 16, "bold")
)
lbl_flujo_neto.grid(
    row=1,
    column=2,
    pady=(0, 10)
)

lbl_flujo_proyectado = tk.Label(
    resumen_flujo,
    text="$0",
    bg=C_BLANCO,
    fg=C_MORADO,
    font=("Segoe UI", 16, "bold")
)
lbl_flujo_proyectado.grid(
    row=1,
    column=3,
    pady=(0, 10)
)

tabla_flujo = ttk.Treeview(
    tab_resumen,
    columns=(
        "Tipo",
        "Categoria",
        "Valor"
    ),
    show="headings"
)

for columna in (
    "Tipo",
    "Categoria",
    "Valor"
):
    tabla_flujo.heading(
        columna,
        text=columna
    )

tabla_flujo.pack(
    fill="both",
    expand=True,
    padx=10,
    pady=(0, 10)
)
tabla_flujo.tag_configure(
    "ingreso",
    foreground=C_VERDE
)
tabla_flujo.tag_configure(
    "egreso",
    foreground=C_ROJO
)

# BANCOS
form_banco = tk.LabelFrame(
    tab_bancos,
    text="NUEVA CUENTA BANCARIA",
    bg=C_BLANCO,
    fg=C_TEXTO,
    font=("Segoe UI", 10, "bold"),
    padx=12,
    pady=10
)
form_banco.pack(
    fill="x",
    padx=10,
    pady=10
)

for columna in range(6):
    form_banco.grid_columnconfigure(
        columna,
        weight=1
    )

for texto, columna in [
    ("Banco", 0),
    ("Número cuenta", 1),
    ("Tipo", 2),
    ("Titular", 3),
    ("Saldo inicial", 4)
]:
    tk.Label(
        form_banco,
        text=texto,
        bg=C_BLANCO,
        fg=C_SUAVE
    ).grid(
        row=0,
        column=columna,
        sticky="w"
    )

entry_banco = ttk.Entry(
    form_banco
)
entry_banco.grid(
    row=1,
    column=0,
    sticky="ew",
    padx=(0, 8)
)

entry_cuenta = ttk.Entry(
    form_banco
)
entry_cuenta.grid(
    row=1,
    column=1,
    sticky="ew",
    padx=(0, 8)
)

combo_tipo_cuenta = ttk.Combobox(
    form_banco,
    values=[
        "Ahorros",
        "Corriente",
        "Caja",
        "Billetera digital"
    ],
    state="readonly"
)
combo_tipo_cuenta.grid(
    row=1,
    column=2,
    sticky="ew",
    padx=(0, 8)
)
combo_tipo_cuenta.set("Ahorros")

entry_titular = ttk.Entry(
    form_banco
)
entry_titular.grid(
    row=1,
    column=3,
    sticky="ew",
    padx=(0, 8)
)

entry_saldo_inicial = ttk.Entry(
    form_banco
)
entry_saldo_inicial.grid(
    row=1,
    column=4,
    sticky="ew",
    padx=(0, 8)
)

tk.Button(
    form_banco,
    text="Guardar cuenta",
    command=guardar_banco,
    bg=C_VERDE,
    fg="white",
    relief="flat",
    padx=14,
    pady=7
).grid(
    row=1,
    column=5,
    sticky="ew"
)

barra_bancos = tk.Frame(
    tab_bancos,
    bg=C_BLANCO
)
barra_bancos.pack(
    fill="x",
    padx=10,
    pady=(0, 5)
)

tk.Button(
    barra_bancos,
    text="Activar / Inactivar",
    command=cambiar_estado_banco,
    bg=C_NARANJA,
    fg="white",
    relief="flat",
    padx=14,
    pady=6
).pack(
    side="right",
    padx=8,
    pady=8
)

tabla_bancos = ttk.Treeview(
    tab_bancos,
    columns=(
        "Banco",
        "Cuenta",
        "Tipo",
        "Titular",
        "Saldo",
        "Estado"
    ),
    show="headings"
)

for columna in (
    "Banco",
    "Cuenta",
    "Tipo",
    "Titular",
    "Saldo",
    "Estado"
):
    tabla_bancos.heading(
        columna,
        text=columna
    )

tabla_bancos.pack(
    fill="both",
    expand=True,
    padx=10,
    pady=(0, 10)
)
tabla_bancos.tag_configure(
    "activo",
    foreground=C_VERDE
)
tabla_bancos.tag_configure(
    "inactivo",
    foreground=C_ROJO
)

# MOVIMIENTOS
form_mov = tk.LabelFrame(
    tab_movimientos,
    text="MOVIMIENTO MANUAL",
    bg=C_BLANCO,
    fg=C_TEXTO,
    font=("Segoe UI", 10, "bold"),
    padx=12,
    pady=10
)
form_mov.pack(
    fill="x",
    padx=10,
    pady=10
)

for columna in range(7):
    form_mov.grid_columnconfigure(
        columna,
        weight=1
    )

titulos_mov = [
    "Banco",
    "Tipo",
    "Categoría",
    "Concepto",
    "Valor",
    "Referencia"
]

for indice, titulo in enumerate(
    titulos_mov
):
    tk.Label(
        form_mov,
        text=titulo,
        bg=C_BLANCO,
        fg=C_SUAVE
    ).grid(
        row=0,
        column=indice,
        sticky="w"
    )

combo_mov_banco = ttk.Combobox(
    form_mov,
    state="readonly"
)
combo_mov_banco.grid(
    row=1,
    column=0,
    sticky="ew",
    padx=(0, 8)
)

combo_mov_tipo = ttk.Combobox(
    form_mov,
    values=[
        "INGRESO",
        "EGRESO"
    ],
    state="readonly"
)
combo_mov_tipo.grid(
    row=1,
    column=1,
    sticky="ew",
    padx=(0, 8)
)

combo_mov_categoria = ttk.Combobox(
    form_mov,
    values=[
        "VENTAS",
        "RECAUDOS",
        "PAGO PROVEEDORES",
        "NÓMINA",
        "SERVICIOS",
        "IMPUESTOS",
        "MANTENIMIENTO",
        "OTROS"
    ]
)
combo_mov_categoria.grid(
    row=1,
    column=2,
    sticky="ew",
    padx=(0, 8)
)

entry_mov_concepto = ttk.Entry(
    form_mov
)
entry_mov_concepto.grid(
    row=1,
    column=3,
    sticky="ew",
    padx=(0, 8)
)

entry_mov_valor = ttk.Entry(
    form_mov
)
entry_mov_valor.grid(
    row=1,
    column=4,
    sticky="ew",
    padx=(0, 8)
)

entry_mov_referencia = ttk.Entry(
    form_mov
)
entry_mov_referencia.grid(
    row=1,
    column=5,
    sticky="ew",
    padx=(0, 8)
)

tk.Button(
    form_mov,
    text="Registrar",
    command=registrar_movimiento_manual,
    bg=C_VERDE,
    fg="white",
    relief="flat",
    padx=14,
    pady=7
).grid(
    row=1,
    column=6,
    sticky="ew"
)

barra_mov = tk.Frame(
    tab_movimientos,
    bg=C_BLANCO
)
barra_mov.pack(
    fill="x",
    padx=10,
    pady=(0, 5)
)

tk.Label(
    barra_mov,
    text="Buscar:",
    bg=C_BLANCO,
    fg=C_TEXTO
).pack(
    side="left",
    padx=(10, 5),
    pady=8
)

entry_buscar_mov = ttk.Entry(
    barra_mov,
    width=30
)
entry_buscar_mov.pack(
    side="left",
    padx=5
)

tk.Button(
    barra_mov,
    text="Actualizar",
    command=cargar_movimientos,
    bg=C_AZUL,
    fg="white",
    relief="flat",
    padx=14,
    pady=6
).pack(
    side="left",
    padx=5
)

tk.Button(
    barra_mov,
    text="Reversar",
    command=reversar_movimiento,
    bg=C_ROJO,
    fg="white",
    relief="flat",
    padx=14,
    pady=6
).pack(
    side="right",
    padx=8
)

tabla_movimientos = ttk.Treeview(
    tab_movimientos,
    columns=(
        "Fecha",
        "Banco",
        "Tipo",
        "Categoria",
        "Concepto",
        "Valor",
        "Referencia",
        "Origen",
        "Estado"
    ),
    show="headings"
)

for columna in (
    "Fecha",
    "Banco",
    "Tipo",
    "Categoria",
    "Concepto",
    "Valor",
    "Referencia",
    "Origen",
    "Estado"
):
    tabla_movimientos.heading(
        columna,
        text=columna
    )

tabla_movimientos.pack(
    fill="both",
    expand=True,
    padx=10,
    pady=(0, 10)
)
tabla_movimientos.tag_configure(
    "ingreso",
    foreground=C_VERDE
)
tabla_movimientos.tag_configure(
    "egreso",
    foreground=C_ROJO
)

# TRANSFERENCIAS
form_trans = tk.LabelFrame(
    tab_transferencias,
    text="TRANSFERENCIA ENTRE CUENTAS",
    bg=C_BLANCO,
    fg=C_TEXTO,
    font=("Segoe UI", 10, "bold"),
    padx=12,
    pady=10
)
form_trans.pack(
    fill="x",
    padx=10,
    pady=10
)

for columna in range(5):
    form_trans.grid_columnconfigure(
        columna,
        weight=1
    )

for texto, columna in [
    ("Cuenta origen", 0),
    ("Cuenta destino", 1),
    ("Valor", 2),
    ("Concepto", 3)
]:
    tk.Label(
        form_trans,
        text=texto,
        bg=C_BLANCO,
        fg=C_SUAVE
    ).grid(
        row=0,
        column=columna,
        sticky="w"
    )

combo_origen = ttk.Combobox(
    form_trans,
    state="readonly"
)
combo_origen.grid(
    row=1,
    column=0,
    sticky="ew",
    padx=(0, 8)
)

combo_destino = ttk.Combobox(
    form_trans,
    state="readonly"
)
combo_destino.grid(
    row=1,
    column=1,
    sticky="ew",
    padx=(0, 8)
)

entry_transferencia_valor = ttk.Entry(
    form_trans
)
entry_transferencia_valor.grid(
    row=1,
    column=2,
    sticky="ew",
    padx=(0, 8)
)

entry_transferencia_concepto = ttk.Entry(
    form_trans
)
entry_transferencia_concepto.grid(
    row=1,
    column=3,
    sticky="ew",
    padx=(0, 8)
)

tk.Button(
    form_trans,
    text="Transferir",
    command=realizar_transferencia,
    bg=C_VERDE,
    fg="white",
    relief="flat",
    padx=14,
    pady=7
).grid(
    row=1,
    column=4,
    sticky="ew"
)

tabla_transferencias = ttk.Treeview(
    tab_transferencias,
    columns=(
        "Fecha",
        "Origen",
        "Destino",
        "Valor",
        "Autorizado"
    ),
    show="headings"
)

for columna in (
    "Fecha",
    "Origen",
    "Destino",
    "Valor",
    "Autorizado"
):
    tabla_transferencias.heading(
        columna,
        text=columna
    )

tabla_transferencias.pack(
    fill="both",
    expand=True,
    padx=10,
    pady=(0, 10)
)

# CONCILIACIÓN
form_conc = tk.LabelFrame(
    tab_conciliacion,
    text="CONCILIACIÓN BANCARIA",
    bg=C_BLANCO,
    fg=C_TEXTO,
    font=("Segoe UI", 10, "bold"),
    padx=12,
    pady=10
)
form_conc.pack(
    fill="x",
    padx=10,
    pady=10
)

for columna in range(5):
    form_conc.grid_columnconfigure(
        columna,
        weight=1
    )

for texto, columna in [
    ("Cuenta bancaria", 0),
    ("Saldo ERP", 1),
    ("Saldo extracto", 2),
    ("Diferencia", 3)
]:
    tk.Label(
        form_conc,
        text=texto,
        bg=C_BLANCO,
        fg=C_SUAVE
    ).grid(
        row=0,
        column=columna,
        sticky="w"
    )

combo_conc_banco = ttk.Combobox(
    form_conc,
    state="readonly"
)
combo_conc_banco.grid(
    row=1,
    column=0,
    sticky="ew",
    padx=(0, 8)
)
combo_conc_banco.bind(
    "<<ComboboxSelected>>",
    cargar_saldo_conciliacion
)

lbl_saldo_erp = tk.Label(
    form_conc,
    text="$0",
    bg=C_BLANCO,
    fg=C_AZUL,
    font=("Segoe UI", 12, "bold")
)
lbl_saldo_erp.grid(
    row=1,
    column=1,
    sticky="w"
)

entry_saldo_extracto = ttk.Entry(
    form_conc
)
entry_saldo_extracto.grid(
    row=1,
    column=2,
    sticky="ew",
    padx=(0, 8)
)
entry_saldo_extracto.bind(
    "<KeyRelease>",
    calcular_diferencia
)

lbl_diferencia = tk.Label(
    form_conc,
    text="$0",
    bg=C_BLANCO,
    fg=C_SUAVE,
    font=("Segoe UI", 12, "bold")
)
lbl_diferencia.grid(
    row=1,
    column=3,
    sticky="w"
)

tk.Button(
    form_conc,
    text="Guardar conciliación",
    command=guardar_conciliacion,
    bg=C_VERDE,
    fg="white",
    relief="flat",
    padx=14,
    pady=7
).grid(
    row=1,
    column=4,
    sticky="ew"
)

tk.Label(
    form_conc,
    text="Observaciones",
    bg=C_BLANCO,
    fg=C_SUAVE
).grid(
    row=2,
    column=0,
    sticky="w",
    pady=(10, 0)
)

txt_conc_obs = tk.Text(
    form_conc,
    height=2,
    relief="solid",
    bd=1
)
txt_conc_obs.grid(
    row=3,
    column=0,
    columnspan=5,
    sticky="ew"
)

tabla_conciliaciones = ttk.Treeview(
    tab_conciliacion,
    columns=(
        "Fecha",
        "Banco",
        "SaldoERP",
        "SaldoExtracto",
        "Diferencia",
        "Estado"
    ),
    show="headings"
)

for columna in (
    "Fecha",
    "Banco",
    "SaldoERP",
    "SaldoExtracto",
    "Diferencia",
    "Estado"
):
    tabla_conciliaciones.heading(
        columna,
        text=columna
    )

tabla_conciliaciones.pack(
    fill="both",
    expand=True,
    padx=10,
    pady=(0, 10)
)
tabla_conciliaciones.tag_configure(
    "ok",
    foreground=C_VERDE
)
tabla_conciliaciones.tag_configure(
    "diferencia",
    foreground=C_ROJO
)

barra_estado = tk.Frame(
    ventana,
    bg=C_BLANCO,
    height=28
)
barra_estado.pack(fill="x")

tk.Label(
    barra_estado,
    text=f"Base de datos: {RUTA_DB}",
    bg=C_BLANCO,
    fg=C_SUAVE,
    font=("Segoe UI", 8)
).pack(
    side="left",
    padx=12
)

tk.Label(
    barra_estado,
    text="BME-ERP Tesorería Integrada v1.0",
    bg=C_BLANCO,
    fg=C_SUAVE,
    font=("Segoe UI", 8)
).pack(
    side="right",
    padx=12
)

refrescar_todo()
ventana.mainloop()
