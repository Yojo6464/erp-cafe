import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
import sqlite3
from datetime import datetime, timedelta
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from proteccion_periodos import validar_periodo_abierto
from motor_contable import (
    contabilizar_evento,
    reversar_comprobante_origen
)

# ============================================================
# BME-ERP
# VENTAS INTEGRADAS v1.0
# ============================================================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
RUTA_DB = os.path.join(BASE_DIR, "erp_cafe.db")

C_FONDO = "#EEF3F8"
C_AZUL = "#0F5C8E"
C_OSCURO = "#153B5B"
C_VERDE = "#15803D"
C_NARANJA = "#C56A00"
C_ROJO = "#B42318"
C_MORADO = "#7C3AED"
C_TEXTO = "#1F2937"
C_SUAVE = "#64748B"
C_BLANCO = "#FFFFFF"
C_BORDE = "#D7E0E8"

detalle_venta = []
venta_seleccionada_id = None


# ============================================================
# BASE DE DATOS
# ============================================================

def conectar():
    conexion = sqlite3.connect(RUTA_DB, timeout=20)
    conexion.execute("PRAGMA foreign_keys = ON")
    return conexion


def columnas_tabla(conexion, tabla):
    return {
        fila[1]
        for fila in conexion.execute(
            f"PRAGMA table_info({tabla})"
        ).fetchall()
    }


def agregar_columna_si_falta(
    conexion,
    tabla,
    columna,
    definicion
):
    if columna not in columnas_tabla(conexion, tabla):
        conexion.execute(
            f"ALTER TABLE {tabla} "
            f"ADD COLUMN {columna} {definicion}"
        )


def inicializar_bd():
    with conectar() as conexion:
        conexion.executescript("""
            CREATE TABLE IF NOT EXISTS ventas_integradas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                numero TEXT NOT NULL UNIQUE,
                fecha TEXT NOT NULL,
                cliente_id INTEGER,
                cliente TEXT NOT NULL,
                documento_cliente TEXT DEFAULT '',
                factura_cliente TEXT DEFAULT '',
                forma_pago TEXT NOT NULL,
                dias_credito INTEGER DEFAULT 0,
                vencimiento TEXT DEFAULT '',
                subtotal REAL DEFAULT 0,
                descuento REAL DEFAULT 0,
                iva REAL DEFAULT 0,
                total REAL DEFAULT 0,
                costo_total REAL DEFAULT 0,
                utilidad REAL DEFAULT 0,
                margen REAL DEFAULT 0,
                banco_id INTEGER,
                estado TEXT DEFAULT 'ACTIVA',
                observaciones TEXT DEFAULT '',
                creada_en TEXT DEFAULT CURRENT_TIMESTAMP,
                anulada_en TEXT DEFAULT '',
                motivo_anulacion TEXT DEFAULT ''
            );

            CREATE TABLE IF NOT EXISTS detalle_ventas_integradas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                venta_id INTEGER NOT NULL,
                producto TEXT NOT NULL,
                presentacion TEXT NOT NULL,
                lote TEXT DEFAULT '',
                cantidad REAL NOT NULL,
                precio_unitario REAL NOT NULL,
                costo_unitario REAL DEFAULT 0,
                subtotal REAL DEFAULT 0,
                costo_total REAL DEFAULT 0,
                utilidad REAL DEFAULT 0,
                FOREIGN KEY(venta_id)
                    REFERENCES ventas_integradas(id)
            );

            CREATE TABLE IF NOT EXISTS devoluciones_ventas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                venta_id INTEGER NOT NULL,
                fecha TEXT NOT NULL,
                producto TEXT NOT NULL,
                presentacion TEXT NOT NULL,
                lote TEXT DEFAULT '',
                cantidad REAL NOT NULL,
                valor REAL DEFAULT 0,
                motivo TEXT DEFAULT '',
                creado_en TEXT DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(venta_id)
                    REFERENCES ventas_integradas(id)
            );

            CREATE TABLE IF NOT EXISTS auditoria_erp(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                fecha_hora TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                usuario TEXT,
                rol TEXT,
                accion TEXT NOT NULL,
                detalle TEXT,
                modulo TEXT
            );

            CREATE INDEX IF NOT EXISTS idx_ventas_fecha
            ON ventas_integradas(fecha);

            CREATE INDEX IF NOT EXISTS idx_ventas_cliente
            ON ventas_integradas(cliente);

            CREATE INDEX IF NOT EXISTS idx_detalle_venta
            ON detalle_ventas_integradas(venta_id);
        """)

        if "clientes" in {
            fila[0]
            for fila in conexion.execute("""
                SELECT name
                FROM sqlite_master
                WHERE type = 'table'
            """).fetchall()
        }:
            agregar_columna_si_falta(
                conexion,
                "clientes",
                "estado",
                "TEXT DEFAULT 'ACTIVO'"
            )

        conexion.commit()


# ============================================================
# UTILIDADES
# ============================================================

def ahora():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def hoy():
    return datetime.now().strftime("%Y-%m-%d")


def moneda(valor):
    return f"${float(valor or 0):,.0f}"


def numero(valor, decimales=2):
    return f"{float(valor or 0):,.{decimales}f}"


def a_numero(valor, nombre, permitir_cero=False):
    texto = str(valor).strip().replace(",", "")

    try:
        dato = float(texto)
    except ValueError:
        raise ValueError(
            f"{nombre} debe ser numérico."
        )

    if permitir_cero:
        if dato < 0:
            raise ValueError(
                f"{nombre} no puede ser negativo."
            )
    elif dato <= 0:
        raise ValueError(
            f"{nombre} debe ser mayor que cero."
        )

    return dato


def registrar_auditoria(
    conexion,
    accion,
    detalle=""
):
    conexion.execute("""
        INSERT INTO auditoria_erp(
            usuario, rol, accion, detalle, modulo
        )
        VALUES (?, ?, ?, ?, 'Ventas')
    """, (
        os.environ.get(
            "ERP_USUARIO",
            "usuario_local"
        ),
        os.environ.get(
            "ERP_ROL",
            "OPERADOR"
        ),
        accion,
        detalle
    ))


def numero_venta():
    with conectar() as conexion:
        siguiente = conexion.execute("""
            SELECT IFNULL(MAX(id), 0) + 1
            FROM ventas_integradas
        """).fetchone()[0]

    return (
        f"VTA-{datetime.now().strftime('%Y%m%d')}-"
        f"{int(siguiente):05d}"
    )


def extraer_id_combo(valor):
    if not valor:
        return None

    try:
        return int(
            str(valor).split("|")[0].strip()
        )
    except (ValueError, IndexError):
        return None


# ============================================================
# CATÁLOGOS
# ============================================================

def cargar_clientes():
    with conectar() as conexion:
        filas = conexion.execute("""
            SELECT
                id,
                COALESCE(nit, ''),
                nombre
            FROM clientes
            WHERE COALESCE(estado, 'ACTIVO') = 'ACTIVO'
            ORDER BY nombre
        """).fetchall()

    combo_cliente["values"] = [
        f"{fila[0]} | {fila[1]} | {fila[2]}"
        for fila in filas
    ]


def cargar_bancos():
    with conectar() as conexion:
        filas = conexion.execute("""
            SELECT
                id,
                banco,
                numero_cuenta,
                saldo
            FROM bancos
            WHERE UPPER(estado) = 'ACTIVA'
            ORDER BY banco, numero_cuenta
        """).fetchall()

    combo_banco["values"] = [
        (
            f"{fila[0]} | {fila[1]} | "
            f"{fila[2]} | {moneda(fila[3])}"
        )
        for fila in filas
    ]


def cargar_productos():
    with conectar() as conexion:
        filas = conexion.execute("""
            SELECT DISTINCT producto
            FROM inventario
            WHERE cantidad > 0
              AND TRIM(COALESCE(producto, '')) <> ''
            ORDER BY producto
        """).fetchall()

    combo_producto["values"] = [
        fila[0]
        for fila in filas
    ]


def cargar_presentaciones(evento=None):
    producto = combo_producto.get().strip()

    with conectar() as conexion:
        filas = conexion.execute("""
            SELECT DISTINCT presentacion
            FROM inventario
            WHERE producto = ?
              AND cantidad > 0
            ORDER BY presentacion
        """, (producto,)).fetchall()

    combo_presentacion["values"] = [
        fila[0]
        for fila in filas
    ]
    combo_presentacion.set("")
    actualizar_existencia()


def actualizar_existencia(evento=None):
    producto = combo_producto.get().strip()
    presentacion = combo_presentacion.get().strip()

    if not producto or not presentacion:
        lbl_existencia.config(
            text="Disponible: 0"
        )
        lbl_costo.config(
            text="Costo promedio: $0"
        )
        return

    with conectar() as conexion:
        fila = conexion.execute("""
            SELECT
                IFNULL(SUM(cantidad), 0),
                CASE
                    WHEN SUM(cantidad) > 0
                    THEN
                        SUM(
                            cantidad
                            * COALESCE(
                                costo_unitario,
                                costo,
                                0
                            )
                        )
                        / SUM(cantidad)
                    ELSE 0
                END
            FROM inventario
            WHERE producto = ?
              AND presentacion = ?
        """, (
            producto,
            presentacion
        )).fetchone()

    lbl_existencia.config(
        text=f"Disponible: {numero(fila[0])}"
    )
    lbl_costo.config(
        text=f"Costo promedio: {moneda(fila[1])}"
    )


def refrescar_catalogos():
    cargar_clientes()
    cargar_bancos()
    cargar_productos()


# ============================================================
# DETALLE DE VENTA
# ============================================================

def seleccionar_lotes(
    conexion,
    producto,
    presentacion,
    requerida
):
    filas = conexion.execute("""
        SELECT
            id,
            cantidad,
            COALESCE(lote, ''),
            COALESCE(
                costo_unitario,
                costo,
                0
            )
        FROM inventario
        WHERE producto = ?
          AND presentacion = ?
          AND cantidad > 0
        ORDER BY
            COALESCE(
                fecha_ingreso,
                fecha,
                ''
            ),
            id
    """, (
        producto,
        presentacion
    )).fetchall()

    disponible = sum(
        float(fila[1] or 0)
        for fila in filas
    )

    if disponible < requerida:
        raise ValueError(
            (
                f"Inventario insuficiente para "
                f"{producto} / {presentacion}. "
                f"Requerido: {requerida:,.2f}; "
                f"disponible: {disponible:,.2f}."
            )
        )

    pendiente = requerida
    resultado = []

    for fila in filas:
        if pendiente <= 0:
            break

        usar = min(
            float(fila[1] or 0),
            pendiente
        )

        resultado.append((
            fila[0],
            usar,
            fila[2],
            float(fila[3] or 0)
        ))
        pendiente -= usar

    return resultado


def mostrar_detalle():
    tabla_detalle.delete(
        *tabla_detalle.get_children()
    )

    for indice, item in enumerate(
        detalle_venta,
        1
    ):
        tabla_detalle.insert(
            "",
            "end",
            iid=str(indice - 1),
            values=(
                indice,
                item["producto"],
                item["presentacion"],
                numero(item["cantidad"]),
                moneda(item["precio"]),
                moneda(
                    item["cantidad"]
                    * item["precio"]
                )
            )
        )

    recalcular_totales()


def agregar_producto():
    producto = combo_producto.get().strip()
    presentacion = (
        combo_presentacion.get().strip()
    )

    if not producto or not presentacion:
        messagebox.showerror(
            "Detalle de venta",
            "Seleccione producto y presentación."
        )
        return

    try:
        cantidad = a_numero(
            entry_cantidad.get(),
            "Cantidad"
        )
        precio = a_numero(
            entry_precio.get(),
            "Precio unitario"
        )
    except ValueError as error:
        messagebox.showerror(
            "Detalle de venta",
            str(error)
        )
        return

    with conectar() as conexion:
        disponible = float(
            conexion.execute("""
                SELECT IFNULL(SUM(cantidad), 0)
                FROM inventario
                WHERE producto = ?
                  AND presentacion = ?
            """, (
                producto,
                presentacion
            )).fetchone()[0] or 0
        )

    agregado = sum(
        item["cantidad"]
        for item in detalle_venta
        if (
            item["producto"] == producto
            and item["presentacion"] == presentacion
        )
    )

    if cantidad + agregado > disponible:
        messagebox.showerror(
            "Inventario",
            (
                f"Stock insuficiente.\n\n"
                f"Disponible: {disponible:,.2f}\n"
                f"Ya agregado: {agregado:,.2f}"
            )
        )
        return

    detalle_venta.append({
        "producto": producto,
        "presentacion": presentacion,
        "cantidad": cantidad,
        "precio": precio
    })

    mostrar_detalle()

    combo_producto.set("")
    combo_presentacion.set("")
    entry_cantidad.delete(0, "end")
    entry_precio.delete(0, "end")
    actualizar_existencia()


def eliminar_linea():
    seleccion = tabla_detalle.selection()

    if not seleccion:
        messagebox.showwarning(
            "Detalle de venta",
            "Seleccione una línea."
        )
        return

    detalle_venta.pop(
        int(seleccion[0])
    )
    mostrar_detalle()


def recalcular_totales(evento=None):
    subtotal = sum(
        item["cantidad"] * item["precio"]
        for item in detalle_venta
    )

    try:
        descuento_pct = a_numero(
            entry_descuento.get() or 0,
            "Descuento",
            permitir_cero=True
        )
        iva_pct = a_numero(
            entry_iva.get() or 0,
            "IVA",
            permitir_cero=True
        )
    except ValueError:
        descuento_pct = 0
        iva_pct = 0

    descuento = subtotal * descuento_pct / 100
    base = subtotal - descuento
    iva = base * iva_pct / 100
    total = base + iva

    lbl_subtotal.config(
        text=moneda(subtotal)
    )
    lbl_descuento.config(
        text=moneda(descuento)
    )
    lbl_iva.config(
        text=moneda(iva)
    )
    lbl_total.config(
        text=moneda(total)
    )


# ============================================================
# VENTA
# ============================================================

def limpiar_venta():
    global detalle_venta

    detalle_venta = []

    lbl_numero.config(
        text=numero_venta()
    )

    entry_fecha.delete(0, "end")
    entry_fecha.insert(0, hoy())

    combo_cliente.set("")
    entry_factura.delete(0, "end")
    combo_pago.set("CONTADO")
    entry_dias.delete(0, "end")
    entry_dias.insert(0, "0")
    combo_banco.set("")
    entry_descuento.delete(0, "end")
    entry_descuento.insert(0, "0")
    entry_iva.delete(0, "end")
    entry_iva.insert(0, "0")
    entry_observaciones.delete(
        "1.0",
        "end"
    )

    combo_producto.set("")
    combo_presentacion.set("")
    entry_cantidad.delete(0, "end")
    entry_precio.delete(0, "end")

    mostrar_detalle()


def cambio_forma_pago(evento=None):
    forma = combo_pago.get().strip()

    if forma == "CRÉDITO":
        entry_dias.config(state="normal")
        if not entry_dias.get().strip():
            entry_dias.insert(0, "30")
        combo_banco.set("")
        combo_banco.config(state="disabled")
    else:
        entry_dias.config(state="normal")
        entry_dias.delete(0, "end")
        entry_dias.insert(0, "0")
        entry_dias.config(state="disabled")
        combo_banco.config(state="readonly")


def registrar_kardex(
    conexion,
    producto,
    presentacion,
    lote,
    salida,
    costo,
    origen,
    observaciones
):
    saldo = float(
        conexion.execute("""
            SELECT IFNULL(
                SUM(entrada - salida),
                0
            )
            FROM kardex
            WHERE producto = ?
              AND presentacion = ?
              AND COALESCE(lote, '') = ?
        """, (
            producto,
            presentacion,
            lote
        )).fetchone()[0] or 0
    )

    saldo_nuevo = saldo - salida

    conexion.execute("""
        INSERT INTO kardex(
            fecha,
            producto,
            presentacion,
            movimiento,
            entrada,
            salida,
            saldo,
            costo_unitario,
            lote,
            origen,
            observaciones
        )
        VALUES (?, ?, ?, ?, 0, ?, ?, ?, ?, ?, ?)
    """, (
        ahora(),
        producto,
        presentacion,
        "SALIDA VENTA",
        salida,
        saldo_nuevo,
        costo,
        lote,
        origen,
        observaciones
    ))


def guardar_venta():
    fecha_venta = entry_fecha.get().strip()

    try:
        datetime.strptime(fecha_venta, "%Y-%m-%d")
        validar_periodo_abierto(fecha_venta)
    except Exception as error:
        messagebox.showerror(
            "Período contable",
            str(error)
        )
        return

    cliente_id = extraer_id_combo(
        combo_cliente.get()
    )

    if not cliente_id:
        messagebox.showerror(
            "Venta",
            "Seleccione un cliente."
        )
        return

    if not detalle_venta:
        messagebox.showerror(
            "Venta",
            "Agregue al menos un producto."
        )
        return

    forma_pago = combo_pago.get().strip()

    if forma_pago not in (
        "CONTADO",
        "CRÉDITO"
    ):
        messagebox.showerror(
            "Venta",
            "Seleccione la forma de pago."
        )
        return

    banco_id = extraer_id_combo(
        combo_banco.get()
    )

    if (
        forma_pago == "CONTADO"
        and not banco_id
    ):
        messagebox.showerror(
            "Venta",
            (
                "Seleccione la cuenta bancaria "
                "que recibirá el pago."
            )
        )
        return

    try:
        datetime.strptime(
            fecha_venta,
            "%Y-%m-%d"
        )

        dias = int(
            entry_dias.get().strip() or 0
        )

        descuento_pct = a_numero(
            entry_descuento.get() or 0,
            "Descuento",
            permitir_cero=True
        )
        iva_pct = a_numero(
            entry_iva.get() or 0,
            "IVA",
            permitir_cero=True
        )
    except ValueError as error:
        messagebox.showerror(
            "Venta",
            str(error)
        )
        return

    if forma_pago == "CRÉDITO" and dias <= 0:
        messagebox.showerror(
            "Venta",
            (
                "Ingrese los días de crédito."
            )
        )
        return

    subtotal = sum(
        item["cantidad"] * item["precio"]
        for item in detalle_venta
    )
    descuento = subtotal * descuento_pct / 100
    base = subtotal - descuento
    iva = base * iva_pct / 100
    total = base + iva

    with conectar() as conexion:
        cliente = conexion.execute("""
            SELECT
                COALESCE(nit, ''),
                nombre
            FROM clientes
            WHERE id = ?
        """, (cliente_id,)).fetchone()

    if not cliente:
        messagebox.showerror(
            "Venta",
            "No se encontró el cliente."
        )
        return

    if not messagebox.askyesno(
        "Confirmar venta",
        (
            f"Venta: {lbl_numero.cget('text')}\n"
            f"Cliente: {cliente[1]}\n"
            f"Forma de pago: {forma_pago}\n"
            f"Total: {moneda(total)}\n\n"
            "¿Desea guardar e integrar la venta?"
        )
    ):
        return

    conexion = conectar()

    try:
        conexion.execute("BEGIN IMMEDIATE")

        cursor = conexion.execute("""
            INSERT INTO ventas_integradas(
                numero,
                fecha,
                cliente_id,
                cliente,
                documento_cliente,
                factura_cliente,
                forma_pago,
                dias_credito,
                vencimiento,
                subtotal,
                descuento,
                iva,
                total,
                costo_total,
                utilidad,
                margen,
                banco_id,
                estado,
                observaciones
            )
            VALUES (
                ?, ?, ?, ?, ?, ?, ?, ?, ?,
                ?, ?, ?, ?, 0, 0, 0, ?,
                'ACTIVA', ?
            )
        """, (
            lbl_numero.cget("text"),
            fecha_venta,
            cliente_id,
            cliente[1],
            cliente[0],
            entry_factura.get().strip(),
            forma_pago,
            dias,
            (
                (
                    datetime.strptime(
                        fecha_venta,
                        "%Y-%m-%d"
                    )
                    + timedelta(days=dias)
                ).strftime("%Y-%m-%d")
                if forma_pago == "CRÉDITO"
                else fecha_venta
            ),
            subtotal,
            descuento,
            iva,
            total,
            banco_id,
            entry_observaciones.get(
                "1.0",
                "end"
            ).strip()
        ))

        venta_id = cursor.lastrowid
        costo_total_venta = 0.0

        factor_descuento = (
            (subtotal - descuento) / subtotal
            if subtotal > 0 else 1
        )

        for item in detalle_venta:
            cantidad_pendiente = item["cantidad"]

            for (
                inventario_id,
                cantidad_usada,
                lote,
                costo_unitario
            ) in seleccionar_lotes(
                conexion,
                item["producto"],
                item["presentacion"],
                cantidad_pendiente
            ):
                conexion.execute("""
                    UPDATE inventario
                    SET cantidad = cantidad - ?
                    WHERE id = ?
                """, (
                    cantidad_usada,
                    inventario_id
                ))

                subtotal_linea = (
                    cantidad_usada
                    * item["precio"]
                    * factor_descuento
                )
                costo_linea = (
                    cantidad_usada
                    * costo_unitario
                )
                utilidad_linea = (
                    subtotal_linea
                    - costo_linea
                )

                costo_total_venta += costo_linea

                conexion.execute("""
                    INSERT INTO detalle_ventas_integradas(
                        venta_id,
                        producto,
                        presentacion,
                        lote,
                        cantidad,
                        precio_unitario,
                        costo_unitario,
                        subtotal,
                        costo_total,
                        utilidad
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    venta_id,
                    item["producto"],
                    item["presentacion"],
                    lote,
                    cantidad_usada,
                    item["precio"],
                    costo_unitario,
                    subtotal_linea,
                    costo_linea,
                    utilidad_linea
                ))

                registrar_kardex(
                    conexion,
                    item["producto"],
                    item["presentacion"],
                    lote,
                    cantidad_usada,
                    costo_unitario,
                    lbl_numero.cget("text"),
                    f"Venta a {cliente[1]}"
                )

        utilidad_total = total - costo_total_venta
        margen = (
            utilidad_total / total * 100
            if total > 0 else 0
        )

        conexion.execute("""
            UPDATE ventas_integradas
            SET costo_total = ?,
                utilidad = ?,
                margen = ?
            WHERE id = ?
        """, (
            costo_total_venta,
            utilidad_total,
            margen,
            venta_id
        ))

        # Compatibilidad con el Dashboard actual.
        conexion.execute("""
            INSERT INTO ventas_cafe(
                fecha,
                cliente,
                producto,
                cantidad,
                precio_unitario,
                total,
                forma_pago
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            fecha_venta,
            cliente[1],
            (
                detalle_venta[0]["producto"]
                if len(detalle_venta) == 1
                else "VENTA VARIOS PRODUCTOS"
            ),
            sum(
                item["cantidad"]
                for item in detalle_venta
            ),
            (
                total
                / sum(
                    item["cantidad"]
                    for item in detalle_venta
                )
            ),
            total,
            forma_pago
        ))

        if forma_pago == "CRÉDITO":
            vencimiento = (
                datetime.strptime(
                    fecha_venta,
                    "%Y-%m-%d"
                )
                + timedelta(days=dias)
            ).strftime("%Y-%m-%d")

            conexion.execute("""
                INSERT INTO cuentas_cobrar(
                    fecha,
                    cliente,
                    concepto,
                    valor,
                    saldo,
                    vencimiento,
                    estado
                )
                VALUES (?, ?, ?, ?, ?, ?, 'PENDIENTE')
            """, (
                fecha_venta,
                cliente[1],
                (
                    f"Venta "
                    f"{lbl_numero.cget('text')}"
                ),
                total,
                total,
                vencimiento
            ))

            conexion.execute("""
                INSERT INTO cuentas_cobrar_v1(
                    fecha,
                    cliente,
                    concepto,
                    valor,
                    estado
                )
                VALUES (?, ?, ?, ?, 'PENDIENTE')
            """, (
                fecha_venta,
                cliente[1],
                (
                    f"Venta "
                    f"{lbl_numero.cget('text')}"
                ),
                total
            ))
        else:
            fila_banco = conexion.execute("""
                SELECT saldo
                FROM bancos
                WHERE id = ?
            """, (banco_id,)).fetchone()

            if not fila_banco:
                raise ValueError(
                    "No se encontró la cuenta bancaria seleccionada."
                )

            saldo_anterior = float(fila_banco[0] or 0)
            saldo_nuevo = saldo_anterior + total

            conexion.execute("""
                UPDATE bancos
                SET saldo = ?
                WHERE id = ?
            """, (
                saldo_nuevo,
                banco_id
            ))

            conexion.execute("""
                INSERT INTO movimientos_bancos(
                    fecha,
                    banco_id,
                    tipo,
                    concepto,
                    valor,
                    saldo_anterior,
                    saldo_nuevo,
                    autorizado_por
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                fecha_venta,
                banco_id,
                "INGRESO VENTA",
                f"Venta {lbl_numero.cget('text')} a {cliente[1]}",
                total,
                saldo_anterior,
                saldo_nuevo,
                os.environ.get(
                    "ERP_USUARIO",
                    "usuario_local"
                )
            ))

        evento_contable = (
            "VENTA_CONTADO"
            if forma_pago == "CONTADO"
            else "VENTA_CREDITO"
        )

        resultado_contable = contabilizar_evento(
            evento=evento_contable,
            valores={
                "total": total,
                "subtotal_sin_iva": base,
                "subtotal": base,
                "descuento": descuento,
                "iva": iva,
                "costo_total": costo_total_venta,
                "utilidad": utilidad_total
            },
            concepto=(
                f"Venta {lbl_numero.cget('text')} "
                f"a {cliente[1]}"
            ),
            modulo_origen="VENTAS",
            tabla_origen="ventas_integradas",
            registro_origen_id=venta_id,
            tercero={
                "tipo_documento": "NIT",
                "numero_documento": (
                    cliente[0]
                    or f"CLIENTE-{cliente_id}"
                ),
                "nombre_razon_social": cliente[1],
                "tipo_tercero": "CLIENTE",
                "origen_modulo": "VENTAS",
                "origen_id": cliente_id
            },
            centro_costo="VENTAS",
            fecha=fecha_venta,
            documento_referencia=(
                entry_factura.get().strip()
                or lbl_numero.cget("text")
            ),
            usuario=os.environ.get(
                "ERP_USUARIO",
                "usuario_local"
            ),
            ruta_db=RUTA_DB,
            conexion_externa=conexion
        )

        registrar_auditoria(
            conexion,
            "REGISTRAR VENTA",
            (
                f"{lbl_numero.cget('text')}; "
                f"cliente {cliente[1]}; "
                f"total {total:.2f}"
            )
        )

        conexion.commit()

        messagebox.showinfo(
            "Venta registrada",
            (
                f"Venta guardada correctamente.\n\n"
                f"Comprobante contable: "
                f"{resultado_contable['consecutivo']}\n"
                f"Total: {moneda(total)}\n"
                f"Costo: {moneda(costo_total_venta)}\n"
                f"Utilidad: {moneda(utilidad_total)}\n"
                f"Margen: {margen:,.2f}%"
            )
        )

        limpiar_venta()
        cargar_historial()
        actualizar_kpis()
        refrescar_catalogos()

    except Exception as error:
        conexion.rollback()

        messagebox.showerror(
            "No fue posible guardar",
            (
                "La operación fue revertida "
                "completamente.\n\n"
                f"{error}"
            )
        )
    finally:
        conexion.close()


# ============================================================
# HISTORIAL Y ANULACIÓN
# ============================================================

def cargar_historial():
    criterio = entry_buscar_venta.get().strip()
    estado = combo_estado_venta.get().strip()

    sql = """
        SELECT
            id,
            numero,
            fecha,
            cliente,
            forma_pago,
            total,
            costo_total,
            utilidad,
            margen,
            estado
        FROM ventas_integradas
        WHERE 1 = 1
    """
    parametros = []

    if criterio:
        patron = f"%{criterio}%"
        sql += """
            AND (
                numero LIKE ?
                OR cliente LIKE ?
                OR factura_cliente LIKE ?
            )
        """
        parametros.extend([
            patron,
            patron,
            patron
        ])

    if estado and estado != "TODOS":
        sql += " AND estado = ?"
        parametros.append(estado)

    sql += " ORDER BY id DESC"

    with conectar() as conexion:
        filas = conexion.execute(
            sql,
            parametros
        ).fetchall()

    tabla_historial.delete(
        *tabla_historial.get_children()
    )

    for fila in filas:
        tabla_historial.insert(
            "",
            "end",
            iid=str(fila[0]),
            values=(
                fila[1],
                fila[2],
                fila[3],
                fila[4],
                moneda(fila[5]),
                moneda(fila[6]),
                moneda(fila[7]),
                f"{float(fila[8] or 0):,.2f}%",
                fila[9]
            ),
            tags=(
                "anulada"
                if fila[9] == "ANULADA"
                else "activa"
            )
        )


def ver_detalle_venta():
    seleccion = tabla_historial.selection()

    if not seleccion:
        messagebox.showwarning(
            "Venta",
            "Seleccione una venta."
        )
        return

    venta_id = int(seleccion[0])

    with conectar() as conexion:
        venta = conexion.execute("""
            SELECT
                numero,
                fecha,
                cliente,
                factura_cliente,
                forma_pago,
                total,
                costo_total,
                utilidad,
                margen,
                estado,
                observaciones
            FROM ventas_integradas
            WHERE id = ?
        """, (venta_id,)).fetchone()

        filas = conexion.execute("""
            SELECT
                producto,
                presentacion,
                lote,
                cantidad,
                precio_unitario,
                costo_unitario,
                subtotal,
                costo_total,
                utilidad
            FROM detalle_ventas_integradas
            WHERE venta_id = ?
            ORDER BY id
        """, (venta_id,)).fetchall()

    top = tk.Toplevel(ventana)
    top.title(
        f"Detalle de venta - {venta[0]}"
    )
    top.geometry("1200x650")
    top.configure(bg=C_FONDO)

    tk.Label(
        top,
        text=f"DETALLE DE VENTA {venta[0]}",
        bg=C_OSCURO,
        fg="white",
        font=("Segoe UI", 16, "bold"),
        pady=14
    ).pack(fill="x")

    resumen = tk.Frame(
        top,
        bg=C_BLANCO
    )
    resumen.pack(
        fill="x",
        padx=15,
        pady=15
    )

    datos = [
        ("Fecha", venta[1]),
        ("Cliente", venta[2]),
        ("Factura", venta[3]),
        ("Forma de pago", venta[4]),
        ("Total", moneda(venta[5])),
        ("Costo", moneda(venta[6])),
        ("Utilidad", moneda(venta[7])),
        ("Margen", f"{float(venta[8] or 0):,.2f}%"),
        ("Estado", venta[9])
    ]

    for indice, (
        titulo,
        valor
    ) in enumerate(datos):
        marco = tk.Frame(
            resumen,
            bg=C_BLANCO
        )
        marco.grid(
            row=indice // 5,
            column=indice % 5,
            sticky="nsew",
            padx=10,
            pady=8
        )

        tk.Label(
            marco,
            text=titulo,
            bg=C_BLANCO,
            fg=C_SUAVE,
            font=("Segoe UI", 8, "bold")
        ).pack(anchor="w")

        tk.Label(
            marco,
            text=valor,
            bg=C_BLANCO,
            fg=C_TEXTO,
            font=("Segoe UI", 10, "bold")
        ).pack(anchor="w")

    columnas = (
        "Producto",
        "Presentacion",
        "Lote",
        "Cantidad",
        "Precio",
        "Costo",
        "Subtotal",
        "CostoTotal",
        "Utilidad"
    )

    tabla = ttk.Treeview(
        top,
        columns=columnas,
        show="headings"
    )

    for columna in columnas:
        tabla.heading(
            columna,
            text=columna
        )

    tabla.pack(
        fill="both",
        expand=True,
        padx=15,
        pady=(0, 15)
    )

    for fila in filas:
        tabla.insert(
            "",
            "end",
            values=(
                fila[0],
                fila[1],
                fila[2],
                numero(fila[3]),
                moneda(fila[4]),
                moneda(fila[5]),
                moneda(fila[6]),
                moneda(fila[7]),
                moneda(fila[8])
            )
        )


def anular_venta():
    seleccion = tabla_historial.selection()

    if not seleccion:
        messagebox.showwarning(
            "Anular venta",
            "Seleccione una venta."
        )
        return

    venta_id = int(seleccion[0])

    with conectar() as conexion:
        venta = conexion.execute("""
            SELECT
                numero,
                cliente,
                forma_pago,
                total,
                banco_id,
                estado,
                fecha
            FROM ventas_integradas
            WHERE id = ?
        """, (venta_id,)).fetchone()

    if not venta:
        return

    if venta[5] == "ANULADA":
        messagebox.showinfo(
            "Anular venta",
            "La venta ya está anulada."
        )
        return

    try:
        validar_periodo_abierto(venta[6])
    except Exception as error:
        messagebox.showerror(
            "Período contable",
            str(error)
        )
        return

    motivo = simpledialog.askstring(
        "Motivo de anulación",
        "Indique el motivo:",
        parent=ventana
    )

    if not motivo:
        return

    if not messagebox.askyesno(
        "Confirmar anulación",
        (
            f"¿Desea anular la venta "
            f"{venta[0]}?\n\n"
            "Se devolverán las existencias y "
            "se revertirán los movimientos."
        )
    ):
        return

    conexion = conectar()

    try:
        conexion.execute("BEGIN IMMEDIATE")

        detalles = conexion.execute("""
            SELECT
                producto,
                presentacion,
                lote,
                cantidad,
                costo_unitario
            FROM detalle_ventas_integradas
            WHERE venta_id = ?
        """, (venta_id,)).fetchall()

        for detalle in detalles:
            registro = conexion.execute("""
                SELECT id
                FROM inventario
                WHERE producto = ?
                  AND presentacion = ?
                  AND COALESCE(lote, '') = ?
                LIMIT 1
            """, (
                detalle[0],
                detalle[1],
                detalle[2]
            )).fetchone()

            if registro:
                conexion.execute("""
                    UPDATE inventario
                    SET cantidad = cantidad + ?
                    WHERE id = ?
                """, (
                    detalle[3],
                    registro[0]
                ))
            else:
                conexion.execute("""
                    INSERT INTO inventario(
                        producto,
                        presentacion,
                        cantidad,
                        lote,
                        costo_unitario,
                        costo,
                        fecha_ingreso,
                        fecha
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    detalle[0],
                    detalle[1],
                    detalle[3],
                    detalle[2],
                    detalle[4],
                    detalle[4],
                    hoy(),
                    hoy()
                ))

            saldo = conexion.execute("""
                SELECT IFNULL(SUM(entrada - salida), 0)
                FROM kardex
                WHERE producto = ?
                  AND presentacion = ?
                  AND COALESCE(lote, '') = ?
            """, (
                detalle[0],
                detalle[1],
                detalle[2]
            )).fetchone()[0]

            conexion.execute("""
                INSERT INTO kardex(
                    fecha,
                    producto,
                    presentacion,
                    movimiento,
                    entrada,
                    salida,
                    saldo,
                    costo_unitario,
                    lote,
                    origen,
                    observaciones
                )
                VALUES (?, ?, ?, 'REVERSIÓN VENTA', ?, 0, ?, ?, ?, ?, ?)
            """, (
                ahora(),
                detalle[0],
                detalle[1],
                detalle[3],
                float(saldo or 0) + float(detalle[3]),
                detalle[4],
                detalle[2],
                venta[0],
                motivo
            ))

        if venta[2] == "CONTADO" and venta[4]:
            conexion.execute("""
                UPDATE bancos
                SET saldo = saldo - ?
                WHERE id = ?
            """, (
                venta[3],
                venta[4]
            ))

        if venta[2] == "CRÉDITO":
            conexion.execute("""
                UPDATE cuentas_cobrar
                SET saldo = 0,
                    estado = 'ANULADA'
                WHERE concepto = ?
            """, (
                f"Venta {venta[0]}",
            ))

            conexion.execute("""
                UPDATE cuentas_cobrar_v1
                SET estado = 'ANULADA'
                WHERE concepto = ?
            """, (
                f"Venta {venta[0]}",
            ))

        conexion.execute("""
            UPDATE ventas_integradas
            SET estado = 'ANULADA',
                anulada_en = ?,
                motivo_anulacion = ?
            WHERE id = ?
        """, (
            ahora(),
            motivo,
            venta_id
        ))

        resultado_reversion = reversar_comprobante_origen(
            modulo_origen="VENTAS",
            tabla_origen="ventas_integradas",
            registro_origen_id=venta_id,
            concepto=(
                f"Reversión de venta {venta[0]}: "
                f"{motivo}"
            ),
            fecha=venta[6],
            documento_referencia=venta[0],
            usuario=os.environ.get(
                "ERP_USUARIO",
                "usuario_local"
            ),
            empresa_codigo="001",
            evento_origen=(
                "VENTA_CONTADO"
                if venta[2] == "CONTADO"
                else "VENTA_CREDITO"
            ),
            evento_reversion="REVERSIÓN_VENTA",
            ruta_db=RUTA_DB,
            conexion_externa=conexion
        )

        registrar_auditoria(
            conexion,
            "ANULAR VENTA",
            f"{venta[0]}; motivo: {motivo}"
        )

        conexion.commit()

        messagebox.showinfo(
            "Anular venta",
            (
                "Venta anulada correctamente.\n\n"
                f"Comprobante de reversión: "
                f"{resultado_reversion['consecutivo']}"
            )
        )

        cargar_historial()
        actualizar_kpis()
        refrescar_catalogos()

    except Exception as error:
        conexion.rollback()
        messagebox.showerror(
            "Anular venta",
            (
                "La operación fue revertida.\n\n"
                f"{error}"
            )
        )
    finally:
        conexion.close()


# ============================================================
# REPORTES
# ============================================================

def exportar_excel():
    carpeta = os.path.join(
        BASE_DIR,
        "reportes"
    )
    os.makedirs(
        carpeta,
        exist_ok=True
    )

    ruta = os.path.join(
        carpeta,
        (
            "ventas_integradas_"
            + datetime.now().strftime(
                "%Y%m%d_%H%M%S"
            )
            + ".xlsx"
        )
    )

    with conectar() as conexion:
        ventas = conexion.execute("""
            SELECT
                numero,
                fecha,
                cliente,
                factura_cliente,
                forma_pago,
                vencimiento,
                subtotal,
                descuento,
                iva,
                total,
                costo_total,
                utilidad,
                margen,
                estado
            FROM ventas_integradas
            ORDER BY id DESC
        """).fetchall()

        detalles = conexion.execute("""
            SELECT
                v.numero,
                d.producto,
                d.presentacion,
                d.lote,
                d.cantidad,
                d.precio_unitario,
                d.costo_unitario,
                d.subtotal,
                d.costo_total,
                d.utilidad
            FROM detalle_ventas_integradas d
            JOIN ventas_integradas v
              ON v.id = d.venta_id
            ORDER BY d.id DESC
        """).fetchall()

    libro = Workbook()
    hoja = libro.active
    hoja.title = "Ventas"

    encabezados = [
        "Número",
        "Fecha",
        "Cliente",
        "Factura",
        "Forma de pago",
        "Vencimiento",
        "Subtotal",
        "Descuento",
        "IVA",
        "Total",
        "Costo",
        "Utilidad",
        "Margen",
        "Estado"
    ]
    hoja.append(encabezados)

    for celda in hoja[1]:
        celda.font = Font(
            bold=True,
            color="FFFFFF"
        )
        celda.fill = PatternFill(
            "solid",
            fgColor="153B5B"
        )
        celda.alignment = Alignment(
            horizontal="center"
        )

    for fila in ventas:
        hoja.append(list(fila))

    hoja.freeze_panes = "A2"
    hoja.auto_filter.ref = hoja.dimensions

    hoja_detalle = libro.create_sheet(
        "Detalle"
    )
    hoja_detalle.append([
        "Venta",
        "Producto",
        "Presentación",
        "Lote",
        "Cantidad",
        "Precio unitario",
        "Costo unitario",
        "Subtotal",
        "Costo total",
        "Utilidad"
    ])

    for celda in hoja_detalle[1]:
        celda.font = Font(
            bold=True,
            color="FFFFFF"
        )
        celda.fill = PatternFill(
            "solid",
            fgColor="0F5C8E"
        )

    for fila in detalles:
        hoja_detalle.append(list(fila))

    libro.save(ruta)

    with conectar() as conexion:
        registrar_auditoria(
            conexion,
            "EXPORTAR VENTAS",
            ruta
        )
        conexion.commit()

    messagebox.showinfo(
        "Exportación",
        f"Reporte generado:\n\n{ruta}"
    )


# ============================================================
# KPIs
# ============================================================

def actualizar_kpis():
    with conectar() as conexion:
        ventas_hoy = float(
            conexion.execute("""
                SELECT IFNULL(SUM(total), 0)
                FROM ventas_integradas
                WHERE fecha = ?
                  AND estado = 'ACTIVA'
            """, (hoy(),)).fetchone()[0] or 0
        )

        ventas_mes = float(
            conexion.execute("""
                SELECT IFNULL(SUM(total), 0)
                FROM ventas_integradas
                WHERE substr(fecha, 1, 7) =
                    substr(?, 1, 7)
                  AND estado = 'ACTIVA'
            """, (hoy(),)).fetchone()[0] or 0
        )

        utilidad_mes = float(
            conexion.execute("""
                SELECT IFNULL(SUM(utilidad), 0)
                FROM ventas_integradas
                WHERE substr(fecha, 1, 7) =
                    substr(?, 1, 7)
                  AND estado = 'ACTIVA'
            """, (hoy(),)).fetchone()[0] or 0
        )

        cartera = float(
            conexion.execute("""
                SELECT IFNULL(SUM(saldo), 0)
                FROM cuentas_cobrar
                WHERE UPPER(estado) = 'PENDIENTE'
            """).fetchone()[0] or 0
        )

    lbl_kpi_hoy.config(
        text=moneda(ventas_hoy)
    )
    lbl_kpi_mes.config(
        text=moneda(ventas_mes)
    )
    lbl_kpi_utilidad.config(
        text=moneda(utilidad_mes)
    )
    lbl_kpi_cartera.config(
        text=moneda(cartera)
    )


def refrescar_todo():
    refrescar_catalogos()
    cargar_historial()
    actualizar_kpis()


# ============================================================
# INTERFAZ
# ============================================================

inicializar_bd()

ventana = tk.Tk()
ventana.title(
    "BME-ERP - Ventas Integradas v1.1 PROTEGIDA"
)
ventana.geometry("1500x900")
ventana.minsize(1180, 720)
ventana.configure(bg=C_FONDO)

try:
    ventana.state("zoomed")
except tk.TclError:
    pass

estilo = ttk.Style()
try:
    estilo.theme_use("clam")
except tk.TclError:
    pass

estilo.configure(
    "Treeview",
    rowheight=28,
    font=("Segoe UI", 9)
)
estilo.configure(
    "Treeview.Heading",
    font=("Segoe UI", 9, "bold")
)

header = tk.Frame(
    ventana,
    bg=C_OSCURO,
    height=82
)
header.pack(fill="x")
header.pack_propagate(False)

tk.Label(
    header,
    text="VENTAS INTEGRADAS · PERÍODOS PROTEGIDOS",
    font=("Segoe UI", 22, "bold"),
    bg=C_OSCURO,
    fg="white"
).pack(
    side="left",
    padx=24,
    pady=20
)

tk.Label(
    header,
    text=(
        "Facturación, inventario, cartera, "
        "bancos y rentabilidad"
    ),
    font=("Segoe UI", 10),
    bg=C_OSCURO,
    fg="#BFDBFE"
).pack(
    side="right",
    padx=24
)

# KPIs
panel_kpi = tk.Frame(
    ventana,
    bg=C_FONDO
)
panel_kpi.pack(
    fill="x",
    padx=18,
    pady=(14, 5)
)

for columna in range(4):
    panel_kpi.grid_columnconfigure(
        columna,
        weight=1
    )


def crear_kpi(columna, titulo, color):
    marco = tk.Frame(
        panel_kpi,
        bg=C_BLANCO,
        highlightbackground=C_BORDE,
        highlightthickness=1
    )
    marco.grid(
        row=0,
        column=columna,
        sticky="ew",
        padx=6
    )

    tk.Frame(
        marco,
        bg=color,
        width=5
    ).pack(
        side="left",
        fill="y"
    )

    interno = tk.Frame(
        marco,
        bg=C_BLANCO
    )
    interno.pack(
        fill="both",
        expand=True,
        padx=14,
        pady=10
    )

    tk.Label(
        interno,
        text=titulo,
        bg=C_BLANCO,
        fg=C_SUAVE,
        font=("Segoe UI", 8, "bold")
    ).pack(anchor="w")

    valor = tk.Label(
        interno,
        text="$0",
        bg=C_BLANCO,
        fg=C_TEXTO,
        font=("Segoe UI", 16, "bold")
    )
    valor.pack(
        anchor="w",
        pady=(3, 0)
    )

    return valor


lbl_kpi_hoy = crear_kpi(
    0,
    "VENTAS DE HOY",
    C_AZUL
)
lbl_kpi_mes = crear_kpi(
    1,
    "VENTAS DEL MES",
    C_VERDE
)
lbl_kpi_utilidad = crear_kpi(
    2,
    "UTILIDAD DEL MES",
    C_MORADO
)
lbl_kpi_cartera = crear_kpi(
    3,
    "CARTERA PENDIENTE",
    C_NARANJA
)

notebook = ttk.Notebook(ventana)
notebook.pack(
    fill="both",
    expand=True,
    padx=18,
    pady=10
)

tab_registro = tk.Frame(
    notebook,
    bg=C_FONDO
)
tab_historial = tk.Frame(
    notebook,
    bg=C_FONDO
)

notebook.add(
    tab_registro,
    text="  Registrar venta  "
)
notebook.add(
    tab_historial,
    text="  Historial y anulación  "
)

# ------------------------------------------------------------
# REGISTRO
# ------------------------------------------------------------

encabezado = tk.LabelFrame(
    tab_registro,
    text="ENCABEZADO DE LA VENTA",
    bg=C_BLANCO,
    fg=C_TEXTO,
    font=("Segoe UI", 10, "bold"),
    padx=12,
    pady=10
)
encabezado.pack(
    fill="x",
    padx=10,
    pady=10
)

for columna in range(7):
    encabezado.grid_columnconfigure(
        columna,
        weight=1
    )

titulos = [
    "Número",
    "Fecha",
    "Cliente",
    "Factura cliente",
    "Forma de pago",
    "Días crédito",
    "Cuenta bancaria"
]

for indice, titulo in enumerate(titulos):
    tk.Label(
        encabezado,
        text=titulo,
        bg=C_BLANCO,
        fg=C_SUAVE
    ).grid(
        row=0,
        column=indice,
        sticky="w"
    )

lbl_numero = tk.Label(
    encabezado,
    text=numero_venta(),
    bg=C_BLANCO,
    fg=C_AZUL,
    font=("Segoe UI", 10, "bold")
)
lbl_numero.grid(
    row=1,
    column=0,
    sticky="w"
)

entry_fecha = ttk.Entry(
    encabezado
)
entry_fecha.grid(
    row=1,
    column=1,
    sticky="ew",
    padx=(0, 8)
)

combo_cliente = ttk.Combobox(
    encabezado,
    state="readonly"
)
combo_cliente.grid(
    row=1,
    column=2,
    sticky="ew",
    padx=(0, 8)
)

entry_factura = ttk.Entry(
    encabezado
)
entry_factura.grid(
    row=1,
    column=3,
    sticky="ew",
    padx=(0, 8)
)

combo_pago = ttk.Combobox(
    encabezado,
    values=[
        "CONTADO",
        "CRÉDITO"
    ],
    state="readonly"
)
combo_pago.grid(
    row=1,
    column=4,
    sticky="ew",
    padx=(0, 8)
)
combo_pago.bind(
    "<<ComboboxSelected>>",
    cambio_forma_pago
)

entry_dias = ttk.Entry(
    encabezado
)
entry_dias.grid(
    row=1,
    column=5,
    sticky="ew",
    padx=(0, 8)
)

combo_banco = ttk.Combobox(
    encabezado,
    state="readonly"
)
combo_banco.grid(
    row=1,
    column=6,
    sticky="ew"
)

tk.Label(
    encabezado,
    text="Descuento %",
    bg=C_BLANCO,
    fg=C_SUAVE
).grid(
    row=2,
    column=0,
    sticky="w",
    pady=(10, 0)
)

entry_descuento = ttk.Entry(
    encabezado
)
entry_descuento.grid(
    row=3,
    column=0,
    sticky="ew",
    padx=(0, 8)
)
entry_descuento.bind(
    "<KeyRelease>",
    recalcular_totales
)

tk.Label(
    encabezado,
    text="IVA %",
    bg=C_BLANCO,
    fg=C_SUAVE
).grid(
    row=2,
    column=1,
    sticky="w",
    pady=(10, 0)
)

entry_iva = ttk.Entry(
    encabezado
)
entry_iva.grid(
    row=3,
    column=1,
    sticky="ew",
    padx=(0, 8)
)
entry_iva.bind(
    "<KeyRelease>",
    recalcular_totales
)

tk.Label(
    encabezado,
    text="Observaciones",
    bg=C_BLANCO,
    fg=C_SUAVE
).grid(
    row=2,
    column=2,
    columnspan=5,
    sticky="w",
    pady=(10, 0)
)

entry_observaciones = tk.Text(
    encabezado,
    height=2,
    relief="solid",
    bd=1
)
entry_observaciones.grid(
    row=3,
    column=2,
    columnspan=5,
    sticky="ew"
)

detalle_frame = tk.LabelFrame(
    tab_registro,
    text="DETALLE DE PRODUCTOS",
    bg=C_BLANCO,
    fg=C_TEXTO,
    font=("Segoe UI", 10, "bold"),
    padx=12,
    pady=10
)
detalle_frame.pack(
    fill="both",
    expand=True,
    padx=10,
    pady=(0, 10)
)

linea = tk.Frame(
    detalle_frame,
    bg=C_BLANCO
)
linea.pack(
    fill="x",
    pady=(0, 8)
)

campos = [
    ("Producto", 0),
    ("Presentación", 1),
    ("Cantidad", 2),
    ("Precio unitario", 3)
]

for texto, columna in campos:
    tk.Label(
        linea,
        text=texto,
        bg=C_BLANCO,
        fg=C_SUAVE
    ).grid(
        row=0,
        column=columna,
        sticky="w",
        padx=4
    )

combo_producto = ttk.Combobox(
    linea,
    state="normal",
    width=34
)
combo_producto.grid(
    row=1,
    column=0,
    sticky="ew",
    padx=4
)
combo_producto.bind(
    "<<ComboboxSelected>>",
    cargar_presentaciones
)

combo_presentacion = ttk.Combobox(
    linea,
    state="normal",
    width=24
)
combo_presentacion.grid(
    row=1,
    column=1,
    sticky="ew",
    padx=4
)
combo_presentacion.bind(
    "<<ComboboxSelected>>",
    actualizar_existencia
)

entry_cantidad = ttk.Entry(
    linea,
    width=16
)
entry_cantidad.grid(
    row=1,
    column=2,
    sticky="ew",
    padx=4
)

entry_precio = ttk.Entry(
    linea,
    width=18
)
entry_precio.grid(
    row=1,
    column=3,
    sticky="ew",
    padx=4
)

tk.Button(
    linea,
    text="Agregar producto",
    command=agregar_producto,
    bg=C_AZUL,
    fg="white",
    relief="flat",
    font=("Segoe UI", 9, "bold"),
    padx=16,
    pady=7
).grid(
    row=1,
    column=4,
    padx=8
)

lbl_existencia = tk.Label(
    linea,
    text="Disponible: 0",
    bg=C_BLANCO,
    fg=C_SUAVE,
    font=("Segoe UI", 8)
)
lbl_existencia.grid(
    row=2,
    column=0,
    sticky="w",
    padx=4,
    pady=(5, 0)
)

lbl_costo = tk.Label(
    linea,
    text="Costo promedio: $0",
    bg=C_BLANCO,
    fg=C_SUAVE,
    font=("Segoe UI", 8)
)
lbl_costo.grid(
    row=2,
    column=1,
    sticky="w",
    padx=4,
    pady=(5, 0)
)

columnas_detalle = (
    "N",
    "Producto",
    "Presentacion",
    "Cantidad",
    "Precio",
    "Subtotal"
)

tabla_detalle = ttk.Treeview(
    detalle_frame,
    columns=columnas_detalle,
    show="headings"
)

for columna in columnas_detalle:
    tabla_detalle.heading(
        columna,
        text=columna
    )

tabla_detalle.pack(
    fill="both",
    expand=True
)

pie_detalle = tk.Frame(
    detalle_frame,
    bg=C_BLANCO
)
pie_detalle.pack(
    fill="x",
    pady=(8, 0)
)

tk.Button(
    pie_detalle,
    text="Eliminar línea seleccionada",
    command=eliminar_linea,
    bg=C_ROJO,
    fg="white",
    relief="flat",
    padx=14,
    pady=6
).pack(side="left")

resumen = tk.Frame(
    pie_detalle,
    bg=C_BLANCO
)
resumen.pack(side="right")

for indice, titulo in enumerate([
    "Subtotal",
    "Descuento",
    "IVA",
    "TOTAL"
]):
    tk.Label(
        resumen,
        text=titulo,
        bg=C_BLANCO,
        fg=C_SUAVE,
        font=("Segoe UI", 8, "bold")
    ).grid(
        row=0,
        column=indice,
        padx=12
    )

lbl_subtotal = tk.Label(
    resumen,
    text="$0",
    bg=C_BLANCO,
    fg=C_TEXTO,
    font=("Segoe UI", 11, "bold")
)
lbl_subtotal.grid(
    row=1,
    column=0,
    padx=12
)

lbl_descuento = tk.Label(
    resumen,
    text="$0",
    bg=C_BLANCO,
    fg=C_TEXTO,
    font=("Segoe UI", 11, "bold")
)
lbl_descuento.grid(
    row=1,
    column=1,
    padx=12
)

lbl_iva = tk.Label(
    resumen,
    text="$0",
    bg=C_BLANCO,
    fg=C_TEXTO,
    font=("Segoe UI", 11, "bold")
)
lbl_iva.grid(
    row=1,
    column=2,
    padx=12
)

lbl_total = tk.Label(
    resumen,
    text="$0",
    bg=C_BLANCO,
    fg=C_AZUL,
    font=("Segoe UI", 14, "bold")
)
lbl_total.grid(
    row=1,
    column=3,
    padx=12
)

acciones = tk.Frame(
    tab_registro,
    bg=C_FONDO
)
acciones.pack(
    fill="x",
    padx=10,
    pady=(0, 10)
)

tk.Button(
    acciones,
    text="GUARDAR E INTEGRAR VENTA",
    command=guardar_venta,
    bg=C_VERDE,
    fg="white",
    relief="flat",
    font=("Segoe UI", 10, "bold"),
    padx=20,
    pady=10
).pack(
    side="left"
)

tk.Button(
    acciones,
    text="Nueva / Limpiar",
    command=limpiar_venta,
    bg=C_OSCURO,
    fg="white",
    relief="flat",
    font=("Segoe UI", 9, "bold"),
    padx=18,
    pady=10
).pack(
    side="left",
    padx=8
)

tk.Button(
    acciones,
    text="Actualizar listas",
    command=refrescar_catalogos,
    bg="#475569",
    fg="white",
    relief="flat",
    font=("Segoe UI", 9, "bold"),
    padx=18,
    pady=10
).pack(
    side="left"
)

# ------------------------------------------------------------
# HISTORIAL
# ------------------------------------------------------------

barra_historial = tk.Frame(
    tab_historial,
    bg=C_BLANCO
)
barra_historial.pack(
    fill="x",
    padx=10,
    pady=10
)

tk.Label(
    barra_historial,
    text="Buscar:",
    bg=C_BLANCO,
    fg=C_TEXTO
).pack(
    side="left",
    padx=(12, 5),
    pady=10
)

entry_buscar_venta = ttk.Entry(
    barra_historial,
    width=28
)
entry_buscar_venta.pack(
    side="left",
    padx=5
)

combo_estado_venta = ttk.Combobox(
    barra_historial,
    values=[
        "TODOS",
        "ACTIVA",
        "ANULADA"
    ],
    state="readonly",
    width=12
)
combo_estado_venta.pack(
    side="left",
    padx=5
)
combo_estado_venta.set("TODOS")

tk.Button(
    barra_historial,
    text="Actualizar",
    command=cargar_historial,
    bg=C_AZUL,
    fg="white",
    relief="flat",
    padx=14,
    pady=6
).pack(
    side="left",
    padx=5
)

tk.Button(
    barra_historial,
    text="Exportar Excel",
    command=exportar_excel,
    bg=C_VERDE,
    fg="white",
    relief="flat",
    padx=14,
    pady=6
).pack(
    side="right",
    padx=8
)

tk.Button(
    barra_historial,
    text="Anular venta",
    command=anular_venta,
    bg=C_ROJO,
    fg="white",
    relief="flat",
    padx=14,
    pady=6
).pack(
    side="right",
    padx=5
)

tk.Button(
    barra_historial,
    text="Ver detalle",
    command=ver_detalle_venta,
    bg=C_OSCURO,
    fg="white",
    relief="flat",
    padx=14,
    pady=6
).pack(
    side="right",
    padx=5
)

columnas_historial = (
    "Numero",
    "Fecha",
    "Cliente",
    "FormaPago",
    "Total",
    "Costo",
    "Utilidad",
    "Margen",
    "Estado"
)

tabla_historial = ttk.Treeview(
    tab_historial,
    columns=columnas_historial,
    show="headings"
)

for columna in columnas_historial:
    tabla_historial.heading(
        columna,
        text=columna
    )

tabla_historial.pack(
    fill="both",
    expand=True,
    padx=10,
    pady=(0, 10)
)
tabla_historial.tag_configure(
    "activa",
    foreground=C_VERDE
)
tabla_historial.tag_configure(
    "anulada",
    foreground=C_ROJO
)

barra_estado = tk.Frame(
    ventana,
    bg=C_BLANCO,
    height=28
)
barra_estado.pack(fill="x")

tk.Label(
    barra_estado,
    text=f"Base de datos: {RUTA_DB}",
    bg=C_BLANCO,
    fg=C_SUAVE,
    font=("Segoe UI", 8)
).pack(
    side="left",
    padx=12
)

tk.Label(
    barra_estado,
    text="BME-ERP Ventas Integradas v1.1 PROTEGIDA",
    bg=C_BLANCO,
    fg=C_SUAVE,
    font=("Segoe UI", 8)
).pack(
    side="right",
    padx=12
)

limpiar_venta()
cambio_forma_pago()
refrescar_todo()

ventana.mainloop()
