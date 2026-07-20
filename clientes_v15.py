import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
from datetime import datetime
from openpyxl import Workbook, load_workbook

# =====================================
# CONEXION
# =====================================

RUTA_DB = r"C:\Users\jrive\visual\erp_cafe.db"

conexion = sqlite3.connect(RUTA_DB)
cursor = conexion.cursor()
cliente_id_seleccionado = None

# =====================================
# TABLA
# =====================================

cursor.execute("""
CREATE TABLE IF NOT EXISTS clientes
(
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    fecha_registro TEXT,
    nit TEXT,
    nombre TEXT,
    telefono TEXT,
    ciudad TEXT,
    correo TEXT
)
""")

conexion.commit()
# =====================================
# CAMPO ESTADO
# =====================================

try:

    cursor.execute(
        """
        ALTER TABLE clientes
        ADD COLUMN estado TEXT
        """
    )

except:

    pass

cursor.execute(
    """
    UPDATE clientes
    SET estado='ACTIVO'
    WHERE estado IS NULL
    """
)

conexion.commit()

# =====================================
# FUNCIONES
# =====================================

def activar_desactivar_cliente():

    seleccionado = tabla.selection()

    if not seleccionado:
        messagebox.showwarning(
            "ERP-027",
            "Seleccione un cliente."
        )
        return

    datos = tabla.item(seleccionado[0], "values")

    cliente_id = datos[0]
    estado = datos[6]

    if estado == "ACTIVO":

        respuesta = messagebox.askyesno(
            "Confirmar",
            f"¿Desea desactivar este cliente?\n\n{datos[2]}"
        )

        if not respuesta:
            return

        nuevo_estado = "INACTIVO"

    else:

        respuesta = messagebox.askyesno(
            "Confirmar",
            f"¿Desea activar este cliente?\n\n{datos[2]}"
        )

        if not respuesta:
            return

        nuevo_estado = "ACTIVO"

    cursor.execute(
        """
        UPDATE clientes
        SET estado = ?
        WHERE id = ?
        """,
        (nuevo_estado, cliente_id)
    )

    conexion.commit()

    mostrar_clientes()

    messagebox.showinfo(
        "ERP-027",
        f"Cliente cambiado a {nuevo_estado}."
    )


def guardar_cliente():

    try:

        nit = entry_nit.get().strip()
        nombre = entry_nombre.get().strip()
        telefono = entry_telefono.get().strip()
        ciudad = entry_ciudad.get().strip()
        correo = entry_correo.get().strip()

        if nombre == "":
            messagebox.showerror(
                "Error",
                "Debe ingresar el nombre."
            )
            return
        cursor.execute(
            """
            SELECT COUNT(*)
            FROM clientes
            WHERE nombre = ?
            """,
            (nombre,)
        )

        existe_nombre = cursor.fetchone()[0]

        if existe_nombre > 0:

            messagebox.showerror(
                "Error",
                "Ya existe un cliente con ese nombre."
            )
            return
        if nit != "":

            cursor.execute(
                """
                SELECT COUNT(*)
                FROM clientes
                WHERE nit = ?
                """,
                (nit,)
            )

            existe_nit = cursor.fetchone()[0]

            if existe_nit > 0:

                messagebox.showerror(
                    "Error",
                    "Ya existe un cliente con ese NIT."
                )
                return

        fecha = datetime.now().strftime(
            "%Y-%m-%d %H:%M:%S"
        )

        cursor.execute("""
        INSERT INTO clientes
(
    fecha_registro,
    nit,
    nombre,
    telefono,
    ciudad,
    correo
)
VALUES (?,?,?,?,?,?)
        """,
        (
            fecha,
            nit,
            nombre,
            telefono,
            ciudad,
            correo
        ))

        conexion.commit()

        informacion(...)
        entry_nit.delete(0, tk.END)
        entry_nombre.delete(0, tk.END)
        entry_telefono.delete(0, tk.END)
        entry_ciudad.delete(0, tk.END)
        entry_correo.delete(0, tk.END)

        mostrar_clientes()

    except Exception as e:

        messagebox.showerror(
            "Error",
            str(e)
        )

# =====================================

def mostrar_clientes():

    tabla.delete(*tabla.get_children())

    # ==============================
    # INDICADORES
    # ==============================

    cursor.execute("SELECT COUNT(*) FROM clientes")
    total = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM clientes WHERE estado='ACTIVO'")
    activos = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM clientes WHERE estado='INACTIVO'")
    inactivos = cursor.fetchone()[0]

    lbl_indicadores.config(
        text=f"Clientes: {total}     🟢 Activos: {activos}     🔴 Inactivos: {inactivos}"
    )

    # ==============================
    # CONSULTA
    # ==============================

    sql = """
        SELECT
            id,
            IFNULL(nit,''),
            nombre,
            telefono,
            ciudad,
            correo,
            estado
        FROM clientes
    """

    if filtro_estado.get() == "ACTIVO":
        sql += " WHERE estado='ACTIVO'"

    elif filtro_estado.get() == "INACTIVO":
        sql += " WHERE estado='INACTIVO'"

    sql += " ORDER BY nombre"

    cursor.execute(sql)

    registros = cursor.fetchall()

    for fila in registros:

        if fila[6] == "ACTIVO":
            etiqueta = "activo"
        else:
            etiqueta = "inactivo"

        tabla.insert(
            "",
            tk.END,
            values=fila,
            tags=(etiqueta,)
        )
    # =====================================

def buscar_clientes():

    texto = entry_busqueda.get().strip()

    tabla.delete(*tabla.get_children())

    cursor.execute(
        """
        SELECT
            id,
            IFNULL(nit,''),
            nombre,
            telefono,
            ciudad,
            correo,
            estado
        FROM clientes
        WHERE
            nit LIKE ?
            OR nombre LIKE ?
        ORDER BY nombre
        """,
        (
            "%" + texto + "%",
            "%" + texto + "%"
        )
    )

    registros = cursor.fetchall()
    for fila in registros:
    
        print("FILA:", fila)

        tabla.insert(
            "",
            tk.END,
            values=(
                fila[0],
                fila[1],
                fila[2],
                fila[3],
                fila[4],
                fila[5],
                fila[6]
            )
    )   
    # =====================================

def seleccionar_cliente(event):

    global cliente_id_seleccionado

    item = tabla.focus()

    if item == "":
        return

    datos = tabla.item(item, "values")

    cliente_id_seleccionado = datos[0]

    entry_nit.delete(0, tk.END)
    entry_nombre.delete(0, tk.END)
    entry_telefono.delete(0, tk.END)
    entry_ciudad.delete(0, tk.END)
    entry_correo.delete(0, tk.END)

    entry_nit.insert(0, datos[1])
    entry_nombre.insert(0, datos[2])
    entry_telefono.insert(0, datos[3])
    entry_ciudad.insert(0, datos[4])
    entry_correo.insert(0, datos[5])  
      # =====================================

def actualizar_cliente():
    print("ENTRO A ACTUALIZAR")

    global cliente_id_seleccionado

    try:

        if cliente_id_seleccionado is None:

            messagebox.showerror(
                "Error",
                "Debe seleccionar un cliente."
            )
            return

        nit = entry_nit.get().strip()
        nombre = entry_nombre.get().strip()
        telefono = entry_telefono.get().strip()
        ciudad = entry_ciudad.get().strip()
        correo = entry_correo.get().strip()
        print("ID:", cliente_id_seleccionado)

        cursor.execute(
            """
            UPDATE clientes
            SET
                nit=?,
                nombre=?,
                telefono=?,
                ciudad=?,
                correo=?
            WHERE id=?
            """,
            (
                nit,
                nombre,
                telefono,
                ciudad,
                correo,
                cliente_id_seleccionado
            )
        )

        conexion.commit()

        messagebox.showinfo(
            "Éxito",
            "Cliente actualizado."
        )

        mostrar_clientes()

    except Exception as e:

        messagebox.showerror(
            "Error",
            str(e)
        )
        # =====================================

def eliminar_cliente():

    global cliente_id_seleccionado

    try:

        if cliente_id_seleccionado is None:

            messagebox.showerror(
                "Error",
                "Debe seleccionar un cliente."
            )
            return

        respuesta = messagebox.askyesno(
            "Confirmar",
            "¿Desea eliminar este cliente?"
        )

        if not respuesta:
            return

        cursor.execute(
            """
            DELETE FROM clientes
            WHERE id = ?
            """,
            (cliente_id_seleccionado,)
        )

        conexion.commit()

        entry_nit.delete(0, tk.END)
        entry_nombre.delete(0, tk.END)
        entry_telefono.delete(0, tk.END)
        entry_ciudad.delete(0, tk.END)
        entry_correo.delete(0, tk.END)

        cliente_id_seleccionado = None

        mostrar_clientes()

        messagebox.showinfo(
            "Éxito",
            "Cliente eliminado correctamente."
        )

    except Exception as e:

        messagebox.showerror(
            "Error",
            str(e)
        )
        # =====================================

def limpiar_formulario():

    global cliente_id_seleccionado

    entry_nit.delete(0, tk.END)
    entry_nombre.delete(0, tk.END)
    entry_telefono.delete(0, tk.END)
    entry_ciudad.delete(0, tk.END)
    entry_correo.delete(0, tk.END)

    cliente_id_seleccionado = None
    # =====================================

def exportar_excel():

    try:

        wb = Workbook()
        ws = wb.active

        ws.title = "Clientes"

        ws.append([
            "ID",
            "Fecha Registro",
            "NIT",
            "Nombre",
            "Telefono",
            "Ciudad",
            "Correo"
        ])

        cursor.execute("""
        SELECT
            id,
            fecha_registro,
            nit,
            nombre,
            telefono,
            ciudad,
            correo
        FROM clientes
        ORDER BY nombre
        """)

        registros = cursor.fetchall()

        for fila in registros:
            ws.append(fila)

        archivo = r"C:\Users\jrive\visual\clientes.xlsx"

        wb.save(archivo)

        messagebox.showinfo(
            "Éxito",
            f"Archivo exportado:\n{archivo}"
        )

    except Exception as e:

        messagebox.showerror(
            "Error",
            str(e)
        )
     # =====================================

def importar_excel():

    try:

        archivo = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Archivos Excel", "*.xlsx")]
        )

        if not archivo:
            return

        wb = load_workbook(archivo)
        ws = wb.active

        importados = 0

        for fila in ws.iter_rows(min_row=2, values_only=True):

            fecha_registro = fila[1]
            nit = str(fila[2] or "").strip()
            nombre = str(fila[3] or "").strip()
            telefono = str(fila[4] or "").strip()
            ciudad = str(fila[5] or "").strip()
            correo = str(fila[6] or "").strip()

            if nombre == "":
                continue

            cursor.execute(
                "SELECT COUNT(*) FROM clientes WHERE nombre=?",
                (nombre,)
            )

            if cursor.fetchone()[0] > 0:
                continue

            cursor.execute(
                "SELECT COUNT(*) FROM clientes WHERE nit=?",
                (nit,)
            )

            if nit != "" and cursor.fetchone()[0] > 0:
                continue

            cursor.execute(
                """
                INSERT INTO clientes
                (
                    fecha_registro,
                    nit,
                    nombre,
                    telefono,
                    ciudad,
                    correo
                )
                VALUES (?,?,?,?,?,?)
                """,
                (
                    fecha_registro,
                    nit,
                    nombre,
                    telefono,
                    ciudad,
                    correo
                )
            )

            importados += 1

        conexion.commit()

        mostrar_clientes()

        messagebox.showinfo(
            "Éxito",
            f"Clientes importados: {importados}"
        )

    except Exception as e:

        messagebox.showerror(
            "Error",
            str(e)
        )   

# =====================================
# VENTANA
# =====================================

ventana = tk.Tk()
ventana.configure(bg="#F4F6F9")

ventana.title(
    "ERP Café Alto de la Cruz - Clientes"
)

ventana.geometry("1100x650")
# =====================================
# FILTRO DE CLIENTES
# =====================================

filtro_estado = tk.StringVar(value="TODOS")

frame_filtro = tk.Frame(ventana)
frame_filtro.pack(pady=5)
# =====================================
# INDICADORES
# =====================================

lbl_indicadores = tk.Label(
    ventana,
    text="",
    font=("Arial", 10, "bold"),
    fg="blue"
)

lbl_indicadores.pack(pady=3)

tk.Label(
    frame_filtro,
    text="Mostrar:"
).pack(side="left", padx=5)
tk.Radiobutton(
    frame_filtro,
    text="Todos",
    variable=filtro_estado,
    value="TODOS",
    command=mostrar_clientes
).pack(side="left")

tk.Radiobutton(
    frame_filtro,
    text="Activos",
    variable=filtro_estado,
    value="ACTIVO",
    command=mostrar_clientes
).pack(side="left")
tk.Radiobutton(
    frame_filtro,
    text="Inactivos",
    variable=filtro_estado,
    value="INACTIVO",
    command=mostrar_clientes
).pack(side="left")


# =====================================
# TITULO
# =====================================

titulo = tk.Label(
    ventana,
    text="GESTIÓN DE CLIENTES",
    font=("Segoe UI", 22, "bold"),
    fg="#1F4E79",
    bg="#F4F6F9"
)


titulo.pack(pady=10)

# =====================================
# FORMULARIO
# =====================================

frame = tk.Frame(
    ventana,
    bg="#F4F6F9"
)

frame.pack(pady=10)

tk.Label(
    frame,
    text="NIT"
).grid(row=0,column=0,padx=10,pady=5)

entry_nit = tk.Entry(
    frame,
    width=20
)

entry_nit.grid(row=1,column=0)

tk.Label(
    frame,
    text="Nombre"
).grid(row=0,column=1,padx=10,pady=5)

entry_nombre = tk.Entry(
    frame,
    width=30
)

entry_nombre.grid(row=1,column=1)

tk.Label(
    frame,
    text="Telefono"
).grid(row=0,column=2,padx=10,pady=5)

entry_telefono = tk.Entry(
    frame,
    width=20
)

entry_telefono.grid(row=1,column=2)

tk.Label(
    frame,
    text="Ciudad"
).grid(row=0,column=3,padx=10,pady=5)

entry_ciudad = tk.Entry(
    frame,
    width=20
)

entry_ciudad.grid(row=1,column=3)

tk.Label(
    frame,
    text="Correo"
).grid(row=0,column=4,padx=10,pady=5)

entry_correo = tk.Entry(
    frame,
    width=30
)

entry_correo.grid(row=1,column=4)
# =====================================
# BUSQUEDA
# =====================================

frame_busqueda = tk.Frame(ventana)

frame_busqueda.pack(pady=5)

tk.Label(
    frame_busqueda,
    text="Buscar Cliente"
).pack(side="left", padx=5)

entry_busqueda = tk.Entry(
    frame_busqueda,
    width=40
)

entry_busqueda.pack(side="left", padx=5)
tk.Button(
    frame_busqueda,
    text="Buscar",
    command=buscar_clientes
).pack(side="left", padx=5)


tk.Button(
    ventana,
    text="Guardar Cliente",
    width=25,
    command=guardar_cliente
).pack(pady=10)
tk.Button(
    ventana,
    text="Actualizar Cliente",
    width=25,
    command=actualizar_cliente
).pack(pady=5)
tk.Button(
    ventana,
    text="Eliminar Cliente",
    width=25,
    command=eliminar_cliente
).pack(pady=5)
tk.Button(
    ventana,
    text="Limpiar Formulario",
    width=25,
    command=limpiar_formulario
).pack(pady=5)
tk.Button(
    ventana,
    text="Exportar Excel",
    width=25,
    command=exportar_excel
).pack(pady=5)
tk.Button(
    ventana,
    text="Importar Excel",
    width=25,
    command=importar_excel
).pack(pady=5)
tk.Button(
    ventana,
    text="Activar / Desactivar",
    width=25,
    command=activar_desactivar_cliente
).pack(pady=5)

# =====================================
# TABLA
# =====================================

tabla = ttk.Treeview(
    ventana,
 columns=(
    "ID",
    "NIT",
    "Nombre",
    "Telefono",
    "Ciudad",
    "Correo",
    "Estado"
),   
    show="headings"
)

tabla.heading("ID", text="ID")
tabla.heading("NIT", text="NIT")
tabla.heading("Nombre", text="Nombre")
tabla.heading("Telefono", text="Teléfono")
tabla.heading("Ciudad", text="Ciudad")
tabla.heading("Correo", text="Correo")
tabla.heading("Estado", text="Estado")
tabla.column("ID", width=70)
tabla.column("NIT", width=120)
tabla.column("Nombre", width=220)
tabla.column("Telefono", width=130)
tabla.column("Ciudad", width=150)
tabla.column("Correo", width=250)
tabla.column("Estado", width=200)
tabla.column("Estado", anchor="center")
tabla.tag_configure(
    "activo",
    foreground="darkgreen"
)

tabla.tag_configure(
    "inactivo",
    foreground="red"
)

tabla.pack(
    fill="both",
    expand=True,
    padx=10,
    pady=10
)
tabla.bind(
    "<<TreeviewSelect>>",
    seleccionar_cliente
)

# =====================================
# CARGA INICIAL
# =====================================

mostrar_clientes()

ventana.mainloop()

conexion.close()