
"""
SIGA ERP - MRP Profesional v1
Archivo: mrp_produccion_v1.py
"""

import os
import sqlite3
import tkinter as tk
from datetime import datetime
from pathlib import Path
from tkinter import messagebox, ttk

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent
RUTA_DB = BASE_DIR / "erp_cafe.db"

C_FONDO = "#EEF3F8"
C_BLANCO = "#FFFFFF"
C_OSCURO = "#153B5B"
C_AZUL = "#0F5C8E"
C_VERDE = "#15803D"
C_NARANJA = "#C56A00"
C_ROJO = "#B42318"
C_GRIS = "#64748B"
C_TEXTO = "#1F2937"
C_BORDE = "#D7E0E8"

resultados_actuales = []
ultimo_mrp_id = None


def conectar():
    if not RUTA_DB.exists():
        raise FileNotFoundError(
            f"No se encontró la base de datos:\n{RUTA_DB}"
        )
    con = sqlite3.connect(RUTA_DB, timeout=30)
    con.row_factory = sqlite3.Row
    con.execute("PRAGMA foreign_keys = ON")
    con.execute("PRAGMA busy_timeout = 10000")
    return con


def usuario():
    return (
        os.environ.get("ERP_USUARIO", "").strip()
        or os.environ.get("USERNAME", "usuario_local")
    )


def hoy():
    return datetime.now().strftime("%Y-%m-%d")


def moneda(valor):
    return f"${float(valor or 0):,.2f}"


def numero(valor):
    return f"{float(valor or 0):,.4f}"


def validar_fecha(texto, nombre):
    texto = texto.strip()
    if not texto:
        return ""
    try:
        datetime.strptime(texto, "%Y-%m-%d")
    except ValueError as error:
        raise ValueError(
            f"{nombre} debe tener formato AAAA-MM-DD."
        ) from error
    return texto


def numero_mrp(cursor):
    fila = cursor.execute("""
        SELECT IFNULL(MAX(id), 0) + 1 AS siguiente
        FROM mrp_ejecuciones
    """).fetchone()

    return (
        f"MRP-{datetime.now().strftime('%Y%m%d')}-"
        f"{int(fila['siguiente']):05d}"
    )


def estados_incluidos():
    estados = []

    if var_aprobada.get():
        estados.append("APROBADA")

    if var_liberada.get():
        estados.append("LIBERADA")

    if var_proceso.get():
        estados.append("EN PROCESO")

    if not estados:
        raise ValueError(
            "Seleccione al menos un estado de orden."
        )

    return estados


def ejecutar_mrp():
    global resultados_actuales, ultimo_mrp_id

    try:
        desde = validar_fecha(entry_desde.get(), "Desde")
        hasta = validar_fecha(entry_hasta.get(), "Hasta")

        if desde and hasta and desde > hasta:
            raise ValueError(
                "La fecha Desde no puede ser mayor que Hasta."
            )

        estados = estados_incluidos()

    except ValueError as error:
        messagebox.showerror(
            "MRP",
            str(error)
        )
        return

    marcadores = ",".join("?" for _ in estados)

    sql = f"""
        SELECT
            r.componente,
            r.presentacion,
            r.tipo_componente,
            r.unidad,
            SUM(
                CASE
                    WHEN o.cantidad_programada > 0
                    THEN r.cantidad_teorica
                         * (
                             (o.cantidad_programada - o.cantidad_producida)
                             / o.cantidad_programada
                           )
                    ELSE 0
                END
            ) AS requerida,
            AVG(r.costo_unitario_estandar) AS costo_unitario
        FROM ordenes_requerimientos_v2 r
        INNER JOIN ordenes_produccion_v2 o
            ON o.id=r.orden_id
        WHERE o.estado IN ({marcadores})
    """

    params = list(estados)

    if desde:
        sql += " AND date(o.fecha_programada) >= date(?)"
        params.append(desde)

    if hasta:
        sql += " AND date(o.fecha_programada) <= date(?)"
        params.append(hasta)

    sql += """
        GROUP BY
            r.componente,
            r.presentacion,
            r.tipo_componente,
            r.unidad
        ORDER BY
            r.componente,
            r.presentacion
    """

    con = conectar()

    try:
        cur = con.cursor()
        cur.execute("BEGIN IMMEDIATE")

        filas = cur.execute(sql, params).fetchall()

        resultados = []

        for fila in filas:
            inventario = cur.execute("""
                SELECT IFNULL(SUM(cantidad), 0) AS disponible
                FROM inventario
                WHERE producto=?
                  AND presentacion=?
            """, (
                fila["componente"],
                fila["presentacion"],
            )).fetchone()

            reservado = cur.execute("""
                SELECT IFNULL(SUM(cantidad_reservada), 0) AS reservado
                FROM reservas_materiales_mrp
                WHERE componente=?
                  AND presentacion=?
                  AND estado='ACTIVA'
            """, (
                fila["componente"],
                fila["presentacion"],
            )).fetchone()

            requerida = float(fila["requerida"] or 0)
            disponible = float(
                inventario["disponible"] or 0
            )
            reservado_valor = float(
                reservado["reservado"] or 0
            )
            disponible_neto = disponible - reservado_valor
            faltante = max(0, requerida - disponible_neto)
            costo_unitario = float(
                fila["costo_unitario"] or 0
            )
            costo_proyectado = requerida * costo_unitario
            estado = "FALTANTE" if faltante > 0 else "OK"

            resultados.append({
                "componente": fila["componente"],
                "presentacion": fila["presentacion"],
                "tipo": fila["tipo_componente"],
                "unidad": fila["unidad"],
                "requerida": requerida,
                "disponible": disponible,
                "reservado": reservado_valor,
                "disponible_neto": disponible_neto,
                "faltante": faltante,
                "costo_unitario": costo_unitario,
                "costo_proyectado": costo_proyectado,
                "estado": estado,
            })

        total_costo = sum(
            item["costo_proyectado"]
            for item in resultados
        )
        total_faltantes = sum(
            1
            for item in resultados
            if item["faltante"] > 0
        )

        numero = numero_mrp(cur)

        cab = cur.execute("""
            INSERT INTO mrp_ejecuciones(
                numero,
                fecha_desde,
                fecha_hasta,
                estados_incluidos,
                costo_total_proyectado,
                componentes_totales,
                componentes_faltantes,
                usuario,
                observaciones
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            numero,
            desde,
            hasta,
            ", ".join(estados),
            total_costo,
            len(resultados),
            total_faltantes,
            usuario(),
            txt_observaciones.get(
                "1.0",
                "end"
            ).strip(),
        ))

        mrp_id = int(cab.lastrowid)

        for item in resultados:
            cur.execute("""
                INSERT INTO mrp_resultados(
                    mrp_id,
                    componente,
                    presentacion,
                    tipo_componente,
                    unidad,
                    cantidad_requerida,
                    inventario_disponible,
                    inventario_reservado,
                    disponible_neto,
                    faltante,
                    costo_unitario_estandar,
                    costo_proyectado,
                    estado
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                mrp_id,
                item["componente"],
                item["presentacion"],
                item["tipo"],
                item["unidad"],
                item["requerida"],
                item["disponible"],
                item["reservado"],
                item["disponible_neto"],
                item["faltante"],
                item["costo_unitario"],
                item["costo_proyectado"],
                item["estado"],
            ))

            if item["faltante"] > 0:
                prioridad = (
                    "ALTA"
                    if item["faltante"]
                    >= item["requerida"] * 0.5
                    else "NORMAL"
                )

                cur.execute("""
                    INSERT INTO sugerencias_compra_mrp(
                        mrp_id,
                        componente,
                        presentacion,
                        cantidad_sugerida,
                        unidad,
                        costo_unitario_estimado,
                        costo_total_estimado,
                        prioridad,
                        estado,
                        observaciones
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'PENDIENTE', ?)
                """, (
                    mrp_id,
                    item["componente"],
                    item["presentacion"],
                    item["faltante"],
                    item["unidad"],
                    item["costo_unitario"],
                    item["faltante"]
                    * item["costo_unitario"],
                    prioridad,
                    "Generado automáticamente por MRP",
                ))

        con.commit()

        resultados_actuales = resultados
        ultimo_mrp_id = mrp_id

        cargar_tabla(resultados_actuales)
        cargar_sugerencias()
        cargar_historial()

        messagebox.showinfo(
            "MRP",
            (
                f"Ejecución {numero} generada correctamente.\n\n"
                f"Componentes: {len(resultados)}\n"
                f"Faltantes: {total_faltantes}\n"
                f"Costo proyectado: {moneda(total_costo)}"
            )
        )

    except Exception as error:
        con.rollback()
        messagebox.showerror(
            "MRP",
            str(error)
        )

    finally:
        con.close()


def cargar_tabla(filas):
    tabla_resultados.delete(
        *tabla_resultados.get_children()
    )

    total_requerido = 0.0
    total_faltante = 0.0
    total_costo = 0.0

    for i, item in enumerate(filas):
        total_requerido += item["requerida"]
        total_faltante += item["faltante"]
        total_costo += item["costo_proyectado"]

        tabla_resultados.insert(
            "",
            "end",
            iid=str(i),
            values=(
                i + 1,
                item["componente"],
                item["presentacion"],
                item["tipo"],
                numero(item["requerida"]),
                item["unidad"],
                numero(item["disponible"]),
                numero(item["reservado"]),
                numero(item["disponible_neto"]),
                numero(item["faltante"]),
                moneda(item["costo_unitario"]),
                moneda(item["costo_proyectado"]),
                item["estado"],
            ),
            tags=("faltante" if item["estado"] == "FALTANTE" else "ok",)
        )

    lbl_componentes.config(text=str(len(filas)))
    lbl_requerido.config(text=numero(total_requerido))
    lbl_faltante.config(text=numero(total_faltante))
    lbl_costo.config(text=moneda(total_costo))


def reservar_materiales():
    estados = estados_incluidos()
    marcadores = ",".join("?" for _ in estados)

    con = conectar()

    try:
        cur = con.cursor()
        cur.execute("BEGIN IMMEDIATE")

        ordenes = cur.execute(f"""
            SELECT id, numero
            FROM ordenes_produccion_v2
            WHERE estado IN ({marcadores})
            ORDER BY fecha_programada, id
        """, estados).fetchall()

        reservas_creadas = 0

        for orden in ordenes:
            reqs = cur.execute("""
                SELECT
                    id,
                    componente,
                    presentacion,
                    cantidad_teorica,
                    cantidad_consumida,
                    unidad
                FROM ordenes_requerimientos_v2
                WHERE orden_id=?
            """, (orden["id"],)).fetchall()

            for req in reqs:
                pendiente = max(
                    0,
                    float(req["cantidad_teorica"] or 0)
                    - float(req["cantidad_consumida"] or 0)
                )

                cur.execute("""
                    INSERT INTO reservas_materiales_mrp(
                        orden_id,
                        requerimiento_id,
                        componente,
                        presentacion,
                        cantidad_reservada,
                        unidad,
                        estado,
                        observaciones
                    )
                    VALUES (?, ?, ?, ?, ?, ?, 'ACTIVA', ?)
                    ON CONFLICT(orden_id, requerimiento_id)
                    DO UPDATE SET
                        cantidad_reservada=excluded.cantidad_reservada,
                        unidad=excluded.unidad,
                        estado='ACTIVA',
                        liberada_en='',
                        observaciones=excluded.observaciones
                """, (
                    orden["id"],
                    req["id"],
                    req["componente"],
                    req["presentacion"],
                    pendiente,
                    req["unidad"],
                    f"Reserva MRP para {orden['numero']}",
                ))

                cur.execute("""
                    UPDATE ordenes_requerimientos_v2
                    SET cantidad_reservada=?
                    WHERE id=?
                """, (
                    pendiente,
                    req["id"],
                ))

                reservas_creadas += 1

        con.commit()

        messagebox.showinfo(
            "MRP",
            f"Reservas teóricas actualizadas: {reservas_creadas}"
        )

        ejecutar_mrp()

    except Exception as error:
        con.rollback()
        messagebox.showerror(
            "MRP",
            str(error)
        )

    finally:
        con.close()


def liberar_reservas():
    if not messagebox.askyesno(
        "MRP",
        "¿Desea liberar todas las reservas activas?"
    ):
        return

    with conectar() as con:
        con.execute("BEGIN IMMEDIATE")
        con.execute("""
            UPDATE reservas_materiales_mrp
            SET estado='LIBERADA',
                liberada_en=CURRENT_TIMESTAMP
            WHERE estado='ACTIVA'
        """)
        con.execute("""
            UPDATE ordenes_requerimientos_v2
            SET cantidad_reservada=0
        """)
        con.commit()

    messagebox.showinfo(
        "MRP",
        "Reservas liberadas correctamente."
    )


def cargar_sugerencias():
    tabla_sugerencias.delete(
        *tabla_sugerencias.get_children()
    )

    with conectar() as con:
        filas = con.execute("""
            SELECT
                id,
                componente,
                presentacion,
                cantidad_sugerida,
                unidad,
                costo_unitario_estimado,
                costo_total_estimado,
                prioridad,
                estado
            FROM sugerencias_compra_mrp
            WHERE mrp_id=COALESCE(?, mrp_id)
            ORDER BY
                CASE prioridad
                    WHEN 'ALTA' THEN 1
                    WHEN 'NORMAL' THEN 2
                    ELSE 3
                END,
                componente
        """, (ultimo_mrp_id,)).fetchall()

    for fila in filas:
        tabla_sugerencias.insert(
            "",
            "end",
            iid=str(fila["id"]),
            values=(
                fila["componente"],
                fila["presentacion"],
                numero(fila["cantidad_sugerida"]),
                fila["unidad"],
                moneda(fila["costo_unitario_estimado"]),
                moneda(fila["costo_total_estimado"]),
                fila["prioridad"],
                fila["estado"],
            )
        )


def marcar_sugerencia():
    sel = tabla_sugerencias.selection()

    if not sel:
        messagebox.showwarning(
            "Sugerencias",
            "Seleccione una sugerencia."
        )
        return

    sugerencia_id = int(sel[0])

    with conectar() as con:
        con.execute("""
            UPDATE sugerencias_compra_mrp
            SET estado='REVISADA'
            WHERE id=?
        """, (sugerencia_id,))
        con.commit()

    cargar_sugerencias()


def cargar_historial():
    tabla_historial.delete(
        *tabla_historial.get_children()
    )

    with conectar() as con:
        filas = con.execute("""
            SELECT
                id,
                numero,
                fecha_desde,
                fecha_hasta,
                estados_incluidos,
                costo_total_proyectado,
                componentes_totales,
                componentes_faltantes,
                usuario,
                creado_en
            FROM mrp_ejecuciones
            ORDER BY id DESC
        """).fetchall()

    for fila in filas:
        tabla_historial.insert(
            "",
            "end",
            values=(
                fila["numero"],
                fila["fecha_desde"],
                fila["fecha_hasta"],
                fila["estados_incluidos"],
                fila["componentes_totales"],
                fila["componentes_faltantes"],
                moneda(fila["costo_total_proyectado"]),
                fila["usuario"],
                fila["creado_en"],
            )
        )


def exportar_excel():
    if not resultados_actuales:
        messagebox.showwarning(
            "Exportar",
            "Primero ejecute el MRP."
        )
        return

    carpeta = BASE_DIR / "reportes"
    carpeta.mkdir(parents=True, exist_ok=True)

    ruta = carpeta / (
        "mrp_"
        + datetime.now().strftime("%Y%m%d_%H%M%S")
        + ".xlsx"
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "MRP"

    ws.merge_cells("A1:M1")
    ws["A1"] = "SIGA ERP - MRP PROFESIONAL"
    ws["A1"].font = Font(
        bold=True,
        color="FFFFFF",
        size=16
    )
    ws["A1"].fill = PatternFill(
        "solid",
        fgColor="153B5B"
    )

    ws.append([])
    ws.append([
        "N",
        "Componente",
        "Presentación",
        "Tipo",
        "Requerida",
        "Unidad",
        "Disponible",
        "Reservado",
        "Disponible neto",
        "Faltante",
        "Costo unitario",
        "Costo proyectado",
        "Estado",
    ])

    for celda in ws[3]:
        celda.font = Font(
            bold=True,
            color="FFFFFF"
        )
        celda.fill = PatternFill(
            "solid",
            fgColor="0F5C8E"
        )

    for i, item in enumerate(resultados_actuales, 1):
        ws.append([
            i,
            item["componente"],
            item["presentacion"],
            item["tipo"],
            item["requerida"],
            item["unidad"],
            item["disponible"],
            item["reservado"],
            item["disponible_neto"],
            item["faltante"],
            item["costo_unitario"],
            item["costo_proyectado"],
            item["estado"],
        ])

    for fila in range(4, ws.max_row + 1):
        ws.cell(fila, 11).number_format = '$#,##0.00'
        ws.cell(fila, 12).number_format = '$#,##0.00'

    for columna, ancho in {
        1: 8,
        2: 28,
        3: 20,
        4: 22,
        5: 14,
        6: 10,
        7: 14,
        8: 14,
        9: 16,
        10: 14,
        11: 16,
        12: 18,
        13: 12,
    }.items():
        ws.column_dimensions[
            get_column_letter(columna)
        ].width = ancho

    wb.save(ruta)

    messagebox.showinfo(
        "Exportar",
        f"Archivo generado:\n\n{ruta}"
    )


ventana = tk.Tk()
ventana.title("SIGA ERP - MRP Profesional v1")
ventana.geometry("1600x920")
ventana.minsize(1250, 760)
ventana.configure(bg=C_FONDO)

try:
    ventana.state("zoomed")
except tk.TclError:
    pass

estilo = ttk.Style()
try:
    estilo.theme_use("clam")
except tk.TclError:
    pass

estilo.configure(
    "Treeview",
    rowheight=27,
    font=("Segoe UI", 9)
)
estilo.configure(
    "Treeview.Heading",
    font=("Segoe UI", 9, "bold")
)

header = tk.Frame(
    ventana,
    bg=C_OSCURO,
    height=86
)
header.pack(fill="x")
header.pack_propagate(False)

tk.Label(
    header,
    text="MRP PROFESIONAL v1",
    bg=C_OSCURO,
    fg="white",
    font=("Segoe UI", 21, "bold")
).pack(anchor="w", padx=25, pady=(15, 0))

tk.Label(
    header,
    text=(
        "Planeación de materiales, faltantes, reservas "
        "y sugerencias de compra"
    ),
    bg=C_OSCURO,
    fg="#BFDBFE",
    font=("Segoe UI", 9)
).pack(anchor="w", padx=26, pady=(3, 0))

filtros = tk.Frame(
    ventana,
    bg=C_BLANCO,
    highlightbackground=C_BORDE,
    highlightthickness=1
)
filtros.pack(
    fill="x",
    padx=15,
    pady=(12, 7)
)

tk.Label(
    filtros,
    text="Desde",
    bg=C_BLANCO,
    fg=C_GRIS
).grid(row=0, column=0, sticky="w", padx=5, pady=(8, 0))

entry_desde = ttk.Entry(filtros, width=14)
entry_desde.grid(row=1, column=0, padx=5, pady=(0, 8))

tk.Label(
    filtros,
    text="Hasta",
    bg=C_BLANCO,
    fg=C_GRIS
).grid(row=0, column=1, sticky="w", padx=5, pady=(8, 0))

entry_hasta = ttk.Entry(filtros, width=14)
entry_hasta.grid(row=1, column=1, padx=5, pady=(0, 8))

var_aprobada = tk.BooleanVar(value=True)
var_liberada = tk.BooleanVar(value=True)
var_proceso = tk.BooleanVar(value=True)

tk.Checkbutton(
    filtros,
    text="APROBADA",
    variable=var_aprobada,
    bg=C_BLANCO
).grid(row=1, column=2, padx=5)

tk.Checkbutton(
    filtros,
    text="LIBERADA",
    variable=var_liberada,
    bg=C_BLANCO
).grid(row=1, column=3, padx=5)

tk.Checkbutton(
    filtros,
    text="EN PROCESO",
    variable=var_proceso,
    bg=C_BLANCO
).grid(row=1, column=4, padx=5)

tk.Button(
    filtros,
    text="Ejecutar MRP",
    command=ejecutar_mrp,
    bg=C_AZUL,
    fg="white",
    relief="flat",
    font=("Segoe UI", 9, "bold"),
    padx=16,
    pady=7
).grid(row=1, column=5, padx=5)

tk.Button(
    filtros,
    text="Reservar materiales",
    command=reservar_materiales,
    bg=C_VERDE,
    fg="white",
    relief="flat",
    padx=16,
    pady=7
).grid(row=1, column=6, padx=5)

tk.Button(
    filtros,
    text="Liberar reservas",
    command=liberar_reservas,
    bg=C_NARANJA,
    fg="white",
    relief="flat",
    padx=16,
    pady=7
).grid(row=1, column=7, padx=5)

tk.Button(
    filtros,
    text="Exportar Excel",
    command=exportar_excel,
    bg=C_GRIS,
    fg="white",
    relief="flat",
    padx=16,
    pady=7
).grid(row=1, column=8, padx=5)

tk.Label(
    filtros,
    text="Observaciones",
    bg=C_BLANCO,
    fg=C_GRIS
).grid(row=0, column=9, sticky="w", padx=5, pady=(8, 0))

txt_observaciones = tk.Text(
    filtros,
    height=2,
    width=35,
    relief="solid",
    bd=1
)
txt_observaciones.grid(
    row=1,
    column=9,
    padx=5,
    pady=(0, 8)
)

kpis = tk.Frame(ventana, bg=C_FONDO)
kpis.pack(fill="x", padx=15, pady=(0, 7))

for i in range(4):
    kpis.columnconfigure(i, weight=1)


def tarjeta(columna, titulo, color):
    marco = tk.Frame(
        kpis,
        bg=C_BLANCO,
        highlightbackground=C_BORDE,
        highlightthickness=1
    )
    marco.grid(
        row=0,
        column=columna,
        sticky="ew",
        padx=4
    )

    tk.Frame(
        marco,
        bg=color,
        width=5
    ).pack(side="left", fill="y")

    contenido = tk.Frame(
        marco,
        bg=C_BLANCO
    )
    contenido.pack(
        fill="both",
        expand=True,
        padx=12,
        pady=8
    )

    tk.Label(
        contenido,
        text=titulo,
        bg=C_BLANCO,
        fg=C_GRIS,
        font=("Segoe UI", 8, "bold")
    ).pack(anchor="w")

    valor = tk.Label(
        contenido,
        text="0",
        bg=C_BLANCO,
        fg=C_TEXTO,
        font=("Segoe UI", 15, "bold")
    )
    valor.pack(anchor="w")
    return valor


lbl_componentes = tarjeta(0, "COMPONENTES", C_AZUL)
lbl_requerido = tarjeta(1, "TOTAL REQUERIDO", C_GRIS)
lbl_faltante = tarjeta(2, "TOTAL FALTANTE", C_ROJO)
lbl_costo = tarjeta(3, "COSTO PROYECTADO", C_VERDE)

notebook = ttk.Notebook(ventana)
notebook.pack(
    fill="both",
    expand=True,
    padx=15,
    pady=(0, 15)
)

tab_resultados = tk.Frame(notebook, bg=C_FONDO)
tab_sugerencias = tk.Frame(notebook, bg=C_FONDO)
tab_historial = tk.Frame(notebook, bg=C_FONDO)

notebook.add(tab_resultados, text="  Resultados MRP  ")
notebook.add(tab_sugerencias, text="  Sugerencias de compra  ")
notebook.add(tab_historial, text="  Historial  ")

columnas = (
    "N",
    "Componente",
    "Presentación",
    "Tipo",
    "Requerida",
    "Unidad",
    "Disponible",
    "Reservado",
    "Disponible neto",
    "Faltante",
    "Costo unitario",
    "Costo proyectado",
    "Estado"
)

tabla_resultados = ttk.Treeview(
    tab_resultados,
    columns=columnas,
    show="headings"
)

for columna in columnas:
    tabla_resultados.heading(
        columna,
        text=columna
    )

tabla_resultados.pack(
    fill="both",
    expand=True,
    padx=10,
    pady=10
)

tabla_resultados.tag_configure(
    "ok",
    foreground=C_VERDE
)
tabla_resultados.tag_configure(
    "faltante",
    foreground=C_ROJO
)

columnas_sug = (
    "Componente",
    "Presentación",
    "Cantidad sugerida",
    "Unidad",
    "Costo unitario",
    "Costo total",
    "Prioridad",
    "Estado"
)

tabla_sugerencias = ttk.Treeview(
    tab_sugerencias,
    columns=columnas_sug,
    show="headings"
)

for columna in columnas_sug:
    tabla_sugerencias.heading(
        columna,
        text=columna
    )

tabla_sugerencias.pack(
    fill="both",
    expand=True,
    padx=10,
    pady=10
)

tk.Button(
    tab_sugerencias,
    text="Marcar sugerencia como revisada",
    command=marcar_sugerencia,
    bg=C_VERDE,
    fg="white",
    relief="flat",
    padx=16,
    pady=7
).pack(anchor="e", padx=10, pady=(0, 10))

columnas_hist = (
    "Número",
    "Desde",
    "Hasta",
    "Estados",
    "Componentes",
    "Faltantes",
    "Costo proyectado",
    "Usuario",
    "Creado en"
)

tabla_historial = ttk.Treeview(
    tab_historial,
    columns=columnas_hist,
    show="headings"
)

for columna in columnas_hist:
    tabla_historial.heading(
        columna,
        text=columna
    )

tabla_historial.pack(
    fill="both",
    expand=True,
    padx=10,
    pady=10
)

entry_desde.insert(0, hoy())
entry_hasta.insert(0, hoy())

cargar_sugerencias()
cargar_historial()

ventana.mainloop()
