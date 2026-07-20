import sqlite3
from openpyxl import Workbook

# =====================================
# BASE DE DATOS
# =====================================

RUTA_DB = r"C:\Users\jrive\visual\erp_cafe.db"

# =====================================
# FUNCION VALOR
# =====================================

def valor(sql):

    try:

        con = sqlite3.connect(RUTA_DB)
        cur = con.cursor()

        cur.execute(sql)

        dato = cur.fetchone()

        con.close()

        if dato is None:
            return 0

        if dato[0] is None:
            return 0

        return dato[0]

    except:

        return 0

# =====================================
# INDICADORES
# =====================================

ventas = valor(
    "SELECT IFNULL(SUM(total),0) FROM ventas_cafe"
)

cartera = valor(
    "SELECT IFNULL(SUM(valor),0) FROM cuentas_cobrar_v1 WHERE estado='Pendiente'"
)

bancos = valor(
    "SELECT IFNULL(SUM(saldo),0) FROM bancos"
)

cxp = valor(
    "SELECT IFNULL(SUM(saldo),0) FROM cuentas_pagar WHERE estado='Pendiente'"
)

produccion = valor(
    "SELECT IFNULL(SUM(cafe_tostado),0) FROM produccion_cafe"
)

inventario = valor(
    "SELECT IFNULL(SUM(saldo_kg),0) FROM almacen_pergamino"
)

nomina = valor(
    "SELECT IFNULL(SUM(neto_pagar),0) FROM nomina"
)

prestaciones = valor(
    "SELECT IFNULL(SUM(total_prestaciones),0) FROM prestaciones"
)

costo_personal = (
    nomina +
    prestaciones
)

utilidad = ventas * 0.40

# =====================================
# EXCEL
# =====================================

wb = Workbook()

ws = wb.active

ws.title = "Resumen Ejecutivo"

ws["A1"] = "INDICADOR"
ws["B1"] = "VALOR"

datos = [

    ("Ventas Totales", ventas),
    ("Cartera", cartera),
    ("Bancos", bancos),
    ("Cuentas por Pagar", cxp),
    ("Produccion Kg", produccion),
    ("Inventario Kg", inventario),
    ("Nomina Acumulada", nomina),
    ("Prestaciones Acumuladas", prestaciones),
    ("Costo Total Personal", costo_personal),
    ("Utilidad Estimada", utilidad)

]

fila = 2

for indicador, valor_indicador in datos:

    ws.cell(
        row=fila,
        column=1,
        value=indicador
    )

    ws.cell(
        row=fila,
        column=2,
        value=valor_indicador
    )

    fila += 1

archivo = r"C:\Users\jrive\visual\Reporte_Ejecutivo.xlsx"

wb.save(archivo)

print()
print("REPORTE GENERADO")
print(archivo)
print()