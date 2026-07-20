import tkinter as tk
from tkinter import ttk, messagebox
import sqlite3
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ============================================================
# BME-ERP
# COSTOS INDUSTRIALES v1.0
# ============================================================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
RUTA_DB = os.path.join(BASE_DIR, "erp_cafe.db")

C_FONDO = "#EEF3F8"
C_AZUL = "#0F5C8E"
C_OSCURO = "#153B5B"
C_VERDE = "#15803D"
C_NARANJA = "#C56A00"
C_ROJO = "#B42318"
C_MORADO = "#7C3AED"
C_TEXTO = "#1F2937"
C_SUAVE = "#64748B"
C_BLANCO = "#FFFFFF"
C_BORDE = "#D7E0E8"

centro_seleccionado_id = None
concepto_seleccionado_id = None


# ============================================================
# BASE DE DATOS
# ============================================================

def conectar():
    conexion = sqlite3.connect(RUTA_DB, timeout=20)
    conexion.execute("PRAGMA foreign_keys = ON")
    return conexion


def columnas_tabla(conexion, tabla):
    return {
        fila[1]
        for fila in conexion.execute(
            f"PRAGMA table_info({tabla})"
        ).fetchall()
    }


def agregar_columna_si_falta(
    conexion,
    tabla,
    columna,
    definicion
):
    if columna not in columnas_tabla(conexion, tabla):
        conexion.execute(
            f"ALTER TABLE {tabla} "
            f"ADD COLUMN {columna} {definicion}"
        )


def inicializar_bd():
    with conectar() as conexion:
        conexion.executescript("""
            CREATE TABLE IF NOT EXISTS centros_costos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                codigo TEXT NOT NULL UNIQUE,
                nombre TEXT NOT NULL,
                tipo TEXT NOT NULL DEFAULT 'PRODUCCIÓN',
                estado TEXT NOT NULL DEFAULT 'ACTIVO',
                observaciones TEXT DEFAULT '',
                creado_en TEXT DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS conceptos_costos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                codigo TEXT NOT NULL UNIQUE,
                nombre TEXT NOT NULL,
                clasificacion TEXT NOT NULL,
                comportamiento TEXT NOT NULL,
                unidad TEXT DEFAULT 'MES',
                valor_base REAL DEFAULT 0,
                centro_costo_id INTEGER,
                estado TEXT NOT NULL DEFAULT 'ACTIVO',
                observaciones TEXT DEFAULT '',
                creado_en TEXT DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(centro_costo_id)
                    REFERENCES centros_costos(id)
            );

            CREATE TABLE IF NOT EXISTS costos_periodicos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                periodo TEXT NOT NULL,
                concepto_id INTEGER NOT NULL,
                centro_costo_id INTEGER,
                valor REAL NOT NULL DEFAULT 0,
                base_distribucion TEXT DEFAULT 'PRODUCCIÓN',
                observaciones TEXT DEFAULT '',
                creado_en TEXT DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(concepto_id)
                    REFERENCES conceptos_costos(id),
                FOREIGN KEY(centro_costo_id)
                    REFERENCES centros_costos(id)
            );

            CREATE TABLE IF NOT EXISTS precios_referencia (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                producto TEXT NOT NULL,
                presentacion TEXT NOT NULL,
                precio_venta REAL NOT NULL DEFAULT 0,
                fecha_actualizacion TEXT DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(producto, presentacion)
            );

            CREATE TABLE IF NOT EXISTS auditoria_erp (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                fecha_hora TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                usuario TEXT,
                rol TEXT,
                accion TEXT NOT NULL,
                detalle TEXT,
                modulo TEXT
            );

            CREATE INDEX IF NOT EXISTS idx_costos_periodo
            ON costos_periodicos(periodo);

            CREATE INDEX IF NOT EXISTS idx_costos_concepto
            ON costos_periodicos(concepto_id);
        """)

        # Compatibilidad con Producción v1 y v2.
        columnas_op = columnas_tabla(
            conexion,
            "ordenes_produccion"
        )

        if columnas_op:
            for columna, definicion in [
                ("merma", "REAL DEFAULT 0"),
                ("desperdicio", "REAL DEFAULT 0"),
                ("reproceso", "REAL DEFAULT 0"),
                ("calidad", "TEXT DEFAULT 'PENDIENTE'")
            ]:
                agregar_columna_si_falta(
                    conexion,
                    "ordenes_produccion",
                    columna,
                    definicion
                )

        # Centros iniciales.
        centros_iniciales = [
            ("CC-001", "Producción general", "PRODUCCIÓN"),
            ("CC-002", "Tostión", "PRODUCCIÓN"),
            ("CC-003", "Molido", "PRODUCCIÓN"),
            ("CC-004", "Empaque", "PRODUCCIÓN"),
            ("CC-005", "Mantenimiento", "APOYO"),
            ("CC-006", "Administración planta", "ADMINISTRATIVO")
        ]

        for codigo, nombre, tipo in centros_iniciales:
            conexion.execute("""
                INSERT OR IGNORE INTO centros_costos(
                    codigo, nombre, tipo, estado
                )
                VALUES (?, ?, ?, 'ACTIVO')
            """, (codigo, nombre, tipo))

        conceptos_iniciales = [
            (
                "CIF-001",
                "Energía indirecta",
                "CIF",
                "VARIABLE",
                "MES"
            ),
            (
                "CIF-002",
                "Agua de planta",
                "CIF",
                "VARIABLE",
                "MES"
            ),
            (
                "CIF-003",
                "Mantenimiento",
                "CIF",
                "MIXTO",
                "MES"
            ),
            (
                "CIF-004",
                "Depreciación maquinaria",
                "CIF",
                "FIJO",
                "MES"
            ),
            (
                "CF-001",
                "Arriendo planta",
                "GASTO FIJO",
                "FIJO",
                "MES"
            ),
            (
                "CF-002",
                "Vigilancia",
                "GASTO FIJO",
                "FIJO",
                "MES"
            ),
            (
                "CF-003",
                "Internet y comunicaciones",
                "GASTO FIJO",
                "FIJO",
                "MES"
            )
        ]

        centro_general = conexion.execute("""
            SELECT id
            FROM centros_costos
            WHERE codigo = 'CC-001'
        """).fetchone()

        centro_id = centro_general[0] if centro_general else None

        for (
            codigo,
            nombre,
            clasificacion,
            comportamiento,
            unidad
        ) in conceptos_iniciales:
            conexion.execute("""
                INSERT OR IGNORE INTO conceptos_costos(
                    codigo, nombre, clasificacion,
                    comportamiento, unidad,
                    centro_costo_id, estado
                )
                VALUES (?, ?, ?, ?, ?, ?, 'ACTIVO')
            """, (
                codigo,
                nombre,
                clasificacion,
                comportamiento,
                unidad,
                centro_id
            ))

        conexion.commit()


# ============================================================
# UTILIDADES
# ============================================================

def hoy():
    return datetime.now().strftime("%Y-%m-%d")


def periodo_actual():
    return datetime.now().strftime("%Y-%m")


def moneda(valor):
    return f"${float(valor or 0):,.0f}"


def numero(valor, decimales=2):
    return f"{float(valor or 0):,.{decimales}f}"


def a_numero(valor, nombre, permitir_cero=True):
    texto = str(valor).strip().replace(",", "")

    try:
        dato = float(texto)
    except ValueError:
        raise ValueError(
            f"{nombre} debe ser un número."
        )

    if permitir_cero:
        if dato < 0:
            raise ValueError(
                f"{nombre} no puede ser negativo."
            )
    elif dato <= 0:
        raise ValueError(
            f"{nombre} debe ser mayor que cero."
        )

    return dato


def registrar_auditoria(
    conexion,
    accion,
    detalle=""
):
    conexion.execute("""
        INSERT INTO auditoria_erp(
            usuario, rol, accion, detalle, modulo
        )
        VALUES (?, ?, ?, ?, 'Costos Industriales')
    """, (
        os.environ.get(
            "ERP_USUARIO",
            "usuario_local"
        ),
        os.environ.get(
            "ERP_ROL",
            "OPERADOR"
        ),
        accion,
        detalle
    ))


def tabla_existe(conexion, tabla):
    return conexion.execute("""
        SELECT COUNT(*)
        FROM sqlite_master
        WHERE type = 'table'
          AND name = ?
    """, (tabla,)).fetchone()[0] > 0


# ============================================================
# CENTROS DE COSTOS
# ============================================================

def limpiar_centro():
    global centro_seleccionado_id

    centro_seleccionado_id = None
    entry_centro_codigo.delete(0, "end")
    entry_centro_nombre.delete(0, "end")
    combo_centro_tipo.set("PRODUCCIÓN")
    combo_centro_estado.set("ACTIVO")
    entry_centro_obs.delete("1.0", "end")
    lbl_modo_centro.config(
        text="NUEVO CENTRO",
        fg=C_VERDE
    )


def guardar_centro():
    global centro_seleccionado_id

    codigo = entry_centro_codigo.get().strip().upper()
    nombre = entry_centro_nombre.get().strip()
    tipo = combo_centro_tipo.get().strip().upper()
    estado = combo_centro_estado.get().strip().upper()
    observaciones = entry_centro_obs.get(
        "1.0",
        "end"
    ).strip()

    if not codigo or not nombre:
        messagebox.showerror(
            "Centro de costos",
            "Ingrese código y nombre."
        )
        return

    conexion = conectar()

    try:
        conexion.execute("BEGIN IMMEDIATE")

        if centro_seleccionado_id is None:
            conexion.execute("""
                INSERT INTO centros_costos(
                    codigo, nombre, tipo,
                    estado, observaciones
                )
                VALUES (?, ?, ?, ?, ?)
            """, (
                codigo,
                nombre,
                tipo,
                estado,
                observaciones
            ))
            accion = "CREAR CENTRO DE COSTOS"
        else:
            conexion.execute("""
                UPDATE centros_costos
                SET codigo = ?,
                    nombre = ?,
                    tipo = ?,
                    estado = ?,
                    observaciones = ?
                WHERE id = ?
            """, (
                codigo,
                nombre,
                tipo,
                estado,
                observaciones,
                centro_seleccionado_id
            ))
            accion = "ACTUALIZAR CENTRO DE COSTOS"

        registrar_auditoria(
            conexion,
            accion,
            f"{codigo} - {nombre}"
        )
        conexion.commit()

        messagebox.showinfo(
            "Centro de costos",
            "Información guardada correctamente."
        )

        limpiar_centro()
        cargar_centros()
        cargar_listas()

    except sqlite3.IntegrityError:
        conexion.rollback()
        messagebox.showerror(
            "Centro de costos",
            "El código ya está registrado."
        )
    except Exception as error:
        conexion.rollback()
        messagebox.showerror(
            "Centro de costos",
            f"No fue posible guardar.\n\n{error}"
        )
    finally:
        conexion.close()


def cargar_centros():
    with conectar() as conexion:
        filas = conexion.execute("""
            SELECT
                id, codigo, nombre,
                tipo, estado
            FROM centros_costos
            ORDER BY codigo
        """).fetchall()

    tabla_centros.delete(
        *tabla_centros.get_children()
    )

    for fila in filas:
        tabla_centros.insert(
            "",
            "end",
            iid=str(fila[0]),
            values=(
                fila[1],
                fila[2],
                fila[3],
                fila[4]
            ),
            tags=(
                "inactivo"
                if fila[4] == "INACTIVO"
                else "activo"
            )
        )


def seleccionar_centro(evento=None):
    global centro_seleccionado_id

    seleccion = tabla_centros.selection()
    if not seleccion:
        return

    centro_seleccionado_id = int(
        seleccion[0]
    )

    with conectar() as conexion:
        fila = conexion.execute("""
            SELECT
                codigo, nombre,
                tipo, estado,
                observaciones
            FROM centros_costos
            WHERE id = ?
        """, (
            centro_seleccionado_id,
        )).fetchone()

    if not fila:
        return

    entry_centro_codigo.delete(0, "end")
    entry_centro_codigo.insert(0, fila[0])

    entry_centro_nombre.delete(0, "end")
    entry_centro_nombre.insert(0, fila[1])

    combo_centro_tipo.set(fila[2])
    combo_centro_estado.set(fila[3])

    entry_centro_obs.delete("1.0", "end")
    entry_centro_obs.insert(
        "1.0",
        fila[4] or ""
    )

    lbl_modo_centro.config(
        text=f"EDITANDO: {fila[0]}",
        fg=C_NARANJA
    )


# ============================================================
# CONCEPTOS Y COSTOS PERIÓDICOS
# ============================================================

def cargar_listas():
    with conectar() as conexion:
        centros = conexion.execute("""
            SELECT id, codigo, nombre
            FROM centros_costos
            WHERE estado = 'ACTIVO'
            ORDER BY codigo
        """).fetchall()

        conceptos = conexion.execute("""
            SELECT id, codigo, nombre
            FROM conceptos_costos
            WHERE estado = 'ACTIVO'
            ORDER BY codigo
        """).fetchall()

    valores_centros = [
        f"{fila[0]} | {fila[1]} | {fila[2]}"
        for fila in centros
    ]
    valores_conceptos = [
        f"{fila[0]} | {fila[1]} | {fila[2]}"
        for fila in conceptos
    ]

    combo_concepto_centro["values"] = (
        valores_centros
    )
    combo_periodico_centro["values"] = (
        valores_centros
    )
    combo_periodico_concepto["values"] = (
        valores_conceptos
    )


def extraer_id_combo(valor):
    if not valor:
        return None

    try:
        return int(
            str(valor).split("|")[0].strip()
        )
    except (ValueError, IndexError):
        return None


def limpiar_concepto():
    global concepto_seleccionado_id

    concepto_seleccionado_id = None
    entry_concepto_codigo.delete(0, "end")
    entry_concepto_nombre.delete(0, "end")
    combo_clasificacion.set("CIF")
    combo_comportamiento.set("FIJO")
    combo_unidad.set("MES")
    entry_valor_base.delete(0, "end")
    entry_valor_base.insert(0, "0")
    combo_concepto_centro.set("")
    combo_concepto_estado.set("ACTIVO")
    entry_concepto_obs.delete("1.0", "end")
    lbl_modo_concepto.config(
        text="NUEVO CONCEPTO",
        fg=C_VERDE
    )


def guardar_concepto():
    global concepto_seleccionado_id

    codigo = entry_concepto_codigo.get().strip().upper()
    nombre = entry_concepto_nombre.get().strip()
    clasificacion = combo_clasificacion.get().strip()
    comportamiento = combo_comportamiento.get().strip()
    unidad = combo_unidad.get().strip()
    centro_id = extraer_id_combo(
        combo_concepto_centro.get()
    )
    estado = combo_concepto_estado.get().strip()
    observaciones = entry_concepto_obs.get(
        "1.0",
        "end"
    ).strip()

    if not codigo or not nombre:
        messagebox.showerror(
            "Concepto",
            "Ingrese código y nombre."
        )
        return

    try:
        valor_base = a_numero(
            entry_valor_base.get() or 0,
            "Valor base"
        )
    except ValueError as error:
        messagebox.showerror(
            "Concepto",
            str(error)
        )
        return

    conexion = conectar()

    try:
        conexion.execute("BEGIN IMMEDIATE")

        if concepto_seleccionado_id is None:
            conexion.execute("""
                INSERT INTO conceptos_costos(
                    codigo, nombre, clasificacion,
                    comportamiento, unidad,
                    valor_base, centro_costo_id,
                    estado, observaciones
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                codigo,
                nombre,
                clasificacion,
                comportamiento,
                unidad,
                valor_base,
                centro_id,
                estado,
                observaciones
            ))
            accion = "CREAR CONCEPTO DE COSTO"
        else:
            conexion.execute("""
                UPDATE conceptos_costos
                SET codigo = ?,
                    nombre = ?,
                    clasificacion = ?,
                    comportamiento = ?,
                    unidad = ?,
                    valor_base = ?,
                    centro_costo_id = ?,
                    estado = ?,
                    observaciones = ?
                WHERE id = ?
            """, (
                codigo,
                nombre,
                clasificacion,
                comportamiento,
                unidad,
                valor_base,
                centro_id,
                estado,
                observaciones,
                concepto_seleccionado_id
            ))
            accion = "ACTUALIZAR CONCEPTO DE COSTO"

        registrar_auditoria(
            conexion,
            accion,
            f"{codigo} - {nombre}"
        )
        conexion.commit()

        messagebox.showinfo(
            "Concepto",
            "Concepto guardado correctamente."
        )

        limpiar_concepto()
        cargar_conceptos()
        cargar_listas()

    except sqlite3.IntegrityError:
        conexion.rollback()
        messagebox.showerror(
            "Concepto",
            "El código ya está registrado."
        )
    except Exception as error:
        conexion.rollback()
        messagebox.showerror(
            "Concepto",
            f"No fue posible guardar.\n\n{error}"
        )
    finally:
        conexion.close()


def cargar_conceptos():
    with conectar() as conexion:
        filas = conexion.execute("""
            SELECT
                cc.id,
                cc.codigo,
                cc.nombre,
                cc.clasificacion,
                cc.comportamiento,
                cc.unidad,
                cc.valor_base,
                COALESCE(c.codigo || ' - ' || c.nombre, ''),
                cc.estado
            FROM conceptos_costos cc
            LEFT JOIN centros_costos c
              ON c.id = cc.centro_costo_id
            ORDER BY cc.codigo
        """).fetchall()

    tabla_conceptos.delete(
        *tabla_conceptos.get_children()
    )

    for fila in filas:
        tabla_conceptos.insert(
            "",
            "end",
            iid=str(fila[0]),
            values=(
                fila[1],
                fila[2],
                fila[3],
                fila[4],
                fila[5],
                moneda(fila[6]),
                fila[7],
                fila[8]
            )
        )


def seleccionar_concepto(evento=None):
    global concepto_seleccionado_id

    seleccion = tabla_conceptos.selection()
    if not seleccion:
        return

    concepto_seleccionado_id = int(
        seleccion[0]
    )

    with conectar() as conexion:
        fila = conexion.execute("""
            SELECT
                codigo, nombre,
                clasificacion,
                comportamiento,
                unidad, valor_base,
                centro_costo_id,
                estado, observaciones
            FROM conceptos_costos
            WHERE id = ?
        """, (
            concepto_seleccionado_id,
        )).fetchone()

        centro = None
        if fila and fila[6]:
            centro = conexion.execute("""
                SELECT id, codigo, nombre
                FROM centros_costos
                WHERE id = ?
            """, (fila[6],)).fetchone()

    if not fila:
        return

    entry_concepto_codigo.delete(0, "end")
    entry_concepto_codigo.insert(0, fila[0])

    entry_concepto_nombre.delete(0, "end")
    entry_concepto_nombre.insert(0, fila[1])

    combo_clasificacion.set(fila[2])
    combo_comportamiento.set(fila[3])
    combo_unidad.set(fila[4])

    entry_valor_base.delete(0, "end")
    entry_valor_base.insert(0, fila[5])

    if centro:
        combo_concepto_centro.set(
            f"{centro[0]} | {centro[1]} | {centro[2]}"
        )
    else:
        combo_concepto_centro.set("")

    combo_concepto_estado.set(fila[7])

    entry_concepto_obs.delete("1.0", "end")
    entry_concepto_obs.insert(
        "1.0",
        fila[8] or ""
    )

    lbl_modo_concepto.config(
        text=f"EDITANDO: {fila[0]}",
        fg=C_NARANJA
    )


def registrar_costo_periodico():
    periodo = entry_periodo.get().strip()
    concepto_id = extraer_id_combo(
        combo_periodico_concepto.get()
    )
    centro_id = extraer_id_combo(
        combo_periodico_centro.get()
    )
    base = combo_base_distribucion.get().strip()
    observaciones = entry_periodico_obs.get(
        "1.0",
        "end"
    ).strip()

    try:
        datetime.strptime(
            periodo + "-01",
            "%Y-%m-%d"
        )
    except ValueError:
        messagebox.showerror(
            "Costo periódico",
            "El periodo debe tener formato AAAA-MM."
        )
        return

    if not concepto_id:
        messagebox.showerror(
            "Costo periódico",
            "Seleccione un concepto."
        )
        return

    try:
        valor = a_numero(
            entry_periodico_valor.get(),
            "Valor",
            permitir_cero=False
        )
    except ValueError as error:
        messagebox.showerror(
            "Costo periódico",
            str(error)
        )
        return

    with conectar() as conexion:
        conexion.execute("""
            INSERT INTO costos_periodicos(
                periodo, concepto_id,
                centro_costo_id, valor,
                base_distribucion,
                observaciones
            )
            VALUES (?, ?, ?, ?, ?, ?)
        """, (
            periodo,
            concepto_id,
            centro_id,
            valor,
            base,
            observaciones
        ))

        registrar_auditoria(
            conexion,
            "REGISTRAR COSTO PERIÓDICO",
            f"Periodo {periodo}; valor {valor:.2f}"
        )
        conexion.commit()

    entry_periodico_valor.delete(0, "end")
    entry_periodico_obs.delete("1.0", "end")

    cargar_costos_periodicos()
    actualizar_indicadores()


def cargar_costos_periodicos():
    periodo = entry_periodo.get().strip()

    with conectar() as conexion:
        filas = conexion.execute("""
            SELECT
                cp.id,
                cp.periodo,
                c.codigo,
                c.nombre,
                c.clasificacion,
                c.comportamiento,
                COALESCE(cc.codigo || ' - ' || cc.nombre, ''),
                cp.valor,
                cp.base_distribucion
            FROM costos_periodicos cp
            JOIN conceptos_costos c
              ON c.id = cp.concepto_id
            LEFT JOIN centros_costos cc
              ON cc.id = cp.centro_costo_id
            WHERE cp.periodo = ?
            ORDER BY cp.id DESC
        """, (periodo,)).fetchall()

    tabla_periodicos.delete(
        *tabla_periodicos.get_children()
    )

    for fila in filas:
        tabla_periodicos.insert(
            "",
            "end",
            iid=str(fila[0]),
            values=(
                fila[1],
                fila[2],
                fila[3],
                fila[4],
                fila[5],
                fila[6],
                moneda(fila[7]),
                fila[8]
            )
        )


def eliminar_costo_periodico():
    seleccion = tabla_periodicos.selection()

    if not seleccion:
        messagebox.showwarning(
            "Costo periódico",
            "Seleccione un registro."
        )
        return

    costo_id = int(seleccion[0])

    if not messagebox.askyesno(
        "Eliminar costo",
        "¿Desea eliminar el costo seleccionado?"
    ):
        return

    with conectar() as conexion:
        fila = conexion.execute("""
            SELECT periodo, valor
            FROM costos_periodicos
            WHERE id = ?
        """, (costo_id,)).fetchone()

        conexion.execute("""
            DELETE FROM costos_periodicos
            WHERE id = ?
        """, (costo_id,))

        registrar_auditoria(
            conexion,
            "ELIMINAR COSTO PERIÓDICO",
            (
                f"Periodo {fila[0]}; "
                f"valor {float(fila[1]):.2f}"
                if fila else str(costo_id)
            )
        )
        conexion.commit()

    cargar_costos_periodicos()
    actualizar_indicadores()


# ============================================================
# COSTEO POR ORDEN
# ============================================================

def calcular_cif_periodo(
    conexion,
    periodo
):
    return float(
        conexion.execute("""
            SELECT IFNULL(SUM(cp.valor), 0)
            FROM costos_periodicos cp
            JOIN conceptos_costos c
              ON c.id = cp.concepto_id
            WHERE cp.periodo = ?
              AND c.clasificacion IN (
                  'CIF',
                  'GASTO FIJO'
              )
        """, (periodo,)).fetchone()[0] or 0
    )


def produccion_total_periodo(
    conexion,
    periodo
):
    return float(
        conexion.execute("""
            SELECT IFNULL(
                SUM(cantidad_producida),
                0
            )
            FROM ordenes_produccion
            WHERE substr(fecha, 1, 7) = ?
              AND estado IN (
                  'CERRADA',
                  'PARCIAL'
              )
        """, (periodo,)).fetchone()[0] or 0
    )


def obtener_precio_venta(
    conexion,
    producto,
    presentacion
):
    precio = conexion.execute("""
        SELECT precio_venta
        FROM precios_referencia
        WHERE producto = ?
          AND presentacion = ?
    """, (
        producto,
        presentacion
    )).fetchone()

    if precio:
        return float(precio[0] or 0)

    if tabla_existe(
        conexion,
        "ventas_cafe"
    ):
        columnas = columnas_tabla(
            conexion,
            "ventas_cafe"
        )

        if {
            "producto",
            "precio_unitario"
        }.issubset(columnas):
            fila = conexion.execute("""
                SELECT IFNULL(
                    AVG(precio_unitario),
                    0
                )
                FROM ventas_cafe
                WHERE producto = ?
            """, (producto,)).fetchone()

            return float(fila[0] or 0)

    return 0.0


def cargar_ordenes_costeadas():
    periodo = entry_periodo_orden.get().strip()
    producto = entry_buscar_producto.get().strip()

    sql = """
        SELECT
            id,
            numero,
            fecha,
            producto_terminado,
            presentacion,
            cantidad_producida,
            lote_producto,
            costo_materiales,
            mano_obra,
            energia,
            gas,
            costos_indirectos,
            costo_total,
            costo_unitario,
            estado,
            COALESCE(merma, 0),
            COALESCE(desperdicio, 0),
            COALESCE(reproceso, 0)
        FROM ordenes_produccion
        WHERE estado IN (
            'CERRADA',
            'PARCIAL'
        )
    """
    parametros = []

    if periodo:
        sql += " AND substr(fecha, 1, 7) = ?"
        parametros.append(periodo)

    if producto:
        sql += """
            AND (
                producto_terminado LIKE ?
                OR presentacion LIKE ?
                OR numero LIKE ?
                OR lote_producto LIKE ?
            )
        """
        patron = f"%{producto}%"
        parametros.extend([
            patron,
            patron,
            patron,
            patron
        ])

    sql += " ORDER BY id DESC"

    with conectar() as conexion:
        filas = conexion.execute(
            sql,
            parametros
        ).fetchall()

        cif_periodo = calcular_cif_periodo(
            conexion,
            periodo
        ) if periodo else 0

        produccion_periodo = (
            produccion_total_periodo(
                conexion,
                periodo
            )
            if periodo else 0
        )

    tabla_ordenes.delete(
        *tabla_ordenes.get_children()
    )

    for fila in filas:
        cantidad = float(fila[5] or 0)
        cif_asignado = (
            cif_periodo
            * cantidad
            / produccion_periodo
            if produccion_periodo > 0
            else 0
        )

        costo_base = float(fila[12] or 0)
        costo_total_gerencial = (
            costo_base
            + cif_asignado
        )
        costo_unitario_gerencial = (
            costo_total_gerencial / cantidad
            if cantidad > 0 else 0
        )

        rendimiento = (
            cantidad
            / (
                cantidad
                + float(fila[15] or 0)
                + float(fila[16] or 0)
            )
            * 100
            if (
                cantidad
                + float(fila[15] or 0)
                + float(fila[16] or 0)
            ) > 0
            else 0
        )

        tabla_ordenes.insert(
            "",
            "end",
            iid=str(fila[0]),
            values=(
                fila[1],
                fila[2],
                fila[3],
                fila[4],
                numero(cantidad),
                fila[6],
                moneda(fila[7]),
                moneda(fila[8]),
                moneda(
                    float(fila[9] or 0)
                    + float(fila[10] or 0)
                ),
                moneda(fila[11]),
                moneda(cif_asignado),
                moneda(costo_total_gerencial),
                moneda(costo_unitario_gerencial),
                f"{rendimiento:,.2f}%",
                fila[14]
            )
        )

    actualizar_indicadores()


def ver_detalle_orden():
    seleccion = tabla_ordenes.selection()

    if not seleccion:
        messagebox.showwarning(
            "Detalle de costos",
            "Seleccione una orden."
        )
        return

    orden_id = int(seleccion[0])

    with conectar() as conexion:
        orden = conexion.execute("""
            SELECT
                numero, fecha,
                producto_terminado,
                presentacion,
                cantidad_programada,
                cantidad_producida,
                lote_producto,
                responsable,
                costo_materiales,
                mano_obra,
                energia,
                gas,
                costos_indirectos,
                costo_total,
                costo_unitario,
                estado,
                COALESCE(merma, 0),
                COALESCE(desperdicio, 0),
                COALESCE(reproceso, 0)
            FROM ordenes_produccion
            WHERE id = ?
        """, (orden_id,)).fetchone()

        consumos = conexion.execute("""
            SELECT
                producto,
                presentacion,
                lote,
                cantidad,
                costo_unitario,
                costo_total
            FROM consumos_produccion
            WHERE orden_id = ?
            ORDER BY costo_total DESC
        """, (orden_id,)).fetchall()

        periodo = str(orden[1])[:7]
        cif_periodo = calcular_cif_periodo(
            conexion,
            periodo
        )
        produccion_periodo = (
            produccion_total_periodo(
                conexion,
                periodo
            )
        )

        cif_asignado = (
            cif_periodo
            * float(orden[5] or 0)
            / produccion_periodo
            if produccion_periodo > 0
            else 0
        )

        precio = obtener_precio_venta(
            conexion,
            orden[2],
            orden[3]
        )

    costo_total_gerencial = (
        float(orden[13] or 0)
        + cif_asignado
    )
    costo_unitario_gerencial = (
        costo_total_gerencial
        / float(orden[5] or 0)
        if float(orden[5] or 0) > 0
        else 0
    )

    utilidad_unitaria = (
        precio - costo_unitario_gerencial
    )
    margen = (
        utilidad_unitaria / precio * 100
        if precio > 0 else 0
    )

    top = tk.Toplevel(ventana)
    top.title(
        f"Costeo detallado - {orden[0]}"
    )
    top.geometry("1180x720")
    top.configure(bg=C_FONDO)

    tk.Label(
        top,
        text=(
            f"COSTEO DETALLADO - {orden[0]}"
        ),
        bg=C_OSCURO,
        fg="white",
        font=("Segoe UI", 16, "bold"),
        pady=14
    ).pack(fill="x")

    resumen = tk.Frame(
        top,
        bg=C_BLANCO,
        highlightbackground=C_BORDE,
        highlightthickness=1
    )
    resumen.pack(
        fill="x",
        padx=15,
        pady=15
    )

    datos = [
        (
            "Producto",
            f"{orden[2]} / {orden[3]}"
        ),
        (
            "Lote",
            orden[6]
        ),
        (
            "Cantidad producida",
            numero(orden[5])
        ),
        (
            "Materiales",
            moneda(orden[8])
        ),
        (
            "Mano de obra",
            moneda(orden[9])
        ),
        (
            "Energía + gas",
            moneda(
                float(orden[10] or 0)
                + float(orden[11] or 0)
            )
        ),
        (
            "Indirectos registrados",
            moneda(orden[12])
        ),
        (
            "CIF distribuido",
            moneda(cif_asignado)
        ),
        (
            "Costo total gerencial",
            moneda(costo_total_gerencial)
        ),
        (
            "Costo unitario gerencial",
            moneda(costo_unitario_gerencial)
        ),
        (
            "Precio de venta",
            moneda(precio)
        ),
        (
            "Margen estimado",
            f"{margen:,.2f}%"
        )
    ]

    for indice, (
        titulo,
        valor
    ) in enumerate(datos):
        fila = indice // 4
        columna = indice % 4

        marco = tk.Frame(
            resumen,
            bg=C_BLANCO
        )
        marco.grid(
            row=fila,
            column=columna,
            sticky="nsew",
            padx=12,
            pady=10
        )

        tk.Label(
            marco,
            text=titulo,
            bg=C_BLANCO,
            fg=C_SUAVE,
            font=("Segoe UI", 8, "bold")
        ).pack(anchor="w")

        tk.Label(
            marco,
            text=valor,
            bg=C_BLANCO,
            fg=C_TEXTO,
            font=("Segoe UI", 11, "bold")
        ).pack(anchor="w")

    for columna in range(4):
        resumen.grid_columnconfigure(
            columna,
            weight=1
        )

    columnas = (
        "Producto",
        "Presentacion",
        "Lote",
        "Cantidad",
        "CostoUnitario",
        "CostoTotal"
    )
    tabla = ttk.Treeview(
        top,
        columns=columnas,
        show="headings"
    )

    encabezados = {
        "Producto": "Materia prima / insumo",
        "Presentacion": "Presentación",
        "Lote": "Lote",
        "Cantidad": "Cantidad",
        "CostoUnitario": "Costo unitario",
        "CostoTotal": "Costo total"
    }

    for columna in columnas:
        tabla.heading(
            columna,
            text=encabezados[columna]
        )

    tabla.column(
        "Producto",
        width=280
    )
    tabla.column(
        "Presentacion",
        width=150
    )
    tabla.column(
        "Lote",
        width=140
    )
    tabla.column(
        "Cantidad",
        width=110,
        anchor="e"
    )
    tabla.column(
        "CostoUnitario",
        width=130,
        anchor="e"
    )
    tabla.column(
        "CostoTotal",
        width=130,
        anchor="e"
    )

    tabla.pack(
        fill="both",
        expand=True,
        padx=15,
        pady=(0, 15)
    )

    for fila in consumos:
        tabla.insert(
            "",
            "end",
            values=(
                fila[0],
                fila[1],
                fila[2],
                numero(fila[3]),
                moneda(fila[4]),
                moneda(fila[5])
            )
        )


# ============================================================
# PRECIOS Y RENTABILIDAD
# ============================================================

def guardar_precio_referencia():
    producto = entry_precio_producto.get().strip()
    presentacion = (
        entry_precio_presentacion.get().strip()
    )

    if not producto or not presentacion:
        messagebox.showerror(
            "Precio de referencia",
            "Ingrese producto y presentación."
        )
        return

    try:
        precio = a_numero(
            entry_precio_venta.get(),
            "Precio de venta",
            permitir_cero=False
        )
    except ValueError as error:
        messagebox.showerror(
            "Precio de referencia",
            str(error)
        )
        return

    with conectar() as conexion:
        conexion.execute("""
            INSERT INTO precios_referencia(
                producto,
                presentacion,
                precio_venta,
                fecha_actualizacion
            )
            VALUES (?, ?, ?, ?)
            ON CONFLICT(producto, presentacion)
            DO UPDATE SET
                precio_venta = excluded.precio_venta,
                fecha_actualizacion =
                    excluded.fecha_actualizacion
        """, (
            producto,
            presentacion,
            precio,
            hoy()
        ))

        registrar_auditoria(
            conexion,
            "ACTUALIZAR PRECIO DE REFERENCIA",
            (
                f"{producto} / {presentacion}; "
                f"{precio:.2f}"
            )
        )
        conexion.commit()

    cargar_rentabilidad()

    messagebox.showinfo(
        "Precio de referencia",
        "Precio guardado correctamente."
    )


def cargar_rentabilidad():
    with conectar() as conexion:
        filas = conexion.execute("""
            SELECT
                op.producto_terminado,
                op.presentacion,
                COUNT(*) AS ordenes,
                SUM(op.cantidad_producida),
                SUM(op.costo_total),
                CASE
                    WHEN SUM(op.cantidad_producida) > 0
                    THEN
                        SUM(op.costo_total)
                        / SUM(op.cantidad_producida)
                    ELSE 0
                END AS costo_promedio,
                COALESCE(pr.precio_venta, 0)
            FROM ordenes_produccion op
            LEFT JOIN precios_referencia pr
              ON pr.producto = op.producto_terminado
             AND pr.presentacion = op.presentacion
            WHERE op.estado IN (
                'CERRADA',
                'PARCIAL'
            )
            GROUP BY
                op.producto_terminado,
                op.presentacion
            ORDER BY
                op.producto_terminado,
                op.presentacion
        """).fetchall()

    tabla_rentabilidad.delete(
        *tabla_rentabilidad.get_children()
    )

    for fila in filas:
        costo = float(fila[5] or 0)
        precio = float(fila[6] or 0)
        utilidad = precio - costo
        margen = (
            utilidad / precio * 100
            if precio > 0 else 0
        )

        tabla_rentabilidad.insert(
            "",
            "end",
            values=(
                fila[0],
                fila[1],
                fila[2],
                numero(fila[3]),
                moneda(fila[4]),
                moneda(costo),
                moneda(precio),
                moneda(utilidad),
                f"{margen:,.2f}%"
            ),
            tags=(
                "positivo"
                if margen >= 0 else "negativo"
            )
        )


# ============================================================
# INDICADORES Y REPORTES
# ============================================================

def actualizar_indicadores():
    periodo = entry_periodo_orden.get().strip()

    with conectar() as conexion:
        costo_produccion = float(
            conexion.execute("""
                SELECT IFNULL(
                    SUM(costo_total),
                    0
                )
                FROM ordenes_produccion
                WHERE substr(fecha, 1, 7) = ?
                  AND estado IN (
                      'CERRADA',
                      'PARCIAL'
                  )
            """, (periodo,)).fetchone()[0] or 0
        )

        produccion = float(
            conexion.execute("""
                SELECT IFNULL(
                    SUM(cantidad_producida),
                    0
                )
                FROM ordenes_produccion
                WHERE substr(fecha, 1, 7) = ?
                  AND estado IN (
                      'CERRADA',
                      'PARCIAL'
                  )
            """, (periodo,)).fetchone()[0] or 0
        )

        costos_periodicos = float(
            conexion.execute("""
                SELECT IFNULL(
                    SUM(valor),
                    0
                )
                FROM costos_periodicos
                WHERE periodo = ?
            """, (periodo,)).fetchone()[0] or 0
        )

        ordenes = conexion.execute("""
            SELECT COUNT(*)
            FROM ordenes_produccion
            WHERE substr(fecha, 1, 7) = ?
              AND estado IN (
                  'CERRADA',
                  'PARCIAL'
              )
        """, (periodo,)).fetchone()[0]

    costo_total = (
        costo_produccion
        + costos_periodicos
    )

    costo_promedio = (
        costo_total / produccion
        if produccion > 0 else 0
    )

    lbl_kpi_costo_total.config(
        text=moneda(costo_total)
    )
    lbl_kpi_costo_unitario.config(
        text=moneda(costo_promedio)
    )
    lbl_kpi_costos_periodicos.config(
        text=moneda(costos_periodicos)
    )
    lbl_kpi_ordenes.config(
        text=str(ordenes)
    )


def exportar_excel():
    periodo = entry_periodo_orden.get().strip()
    carpeta = os.path.join(
        BASE_DIR,
        "reportes"
    )
    os.makedirs(
        carpeta,
        exist_ok=True
    )

    ruta = os.path.join(
        carpeta,
        (
            f"costos_industriales_"
            f"{periodo}_"
            f"{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            f".xlsx"
        )
    )

    with conectar() as conexion:
        ordenes = conexion.execute("""
            SELECT
                numero, fecha,
                producto_terminado,
                presentacion,
                cantidad_programada,
                cantidad_producida,
                lote_producto,
                responsable,
                costo_materiales,
                mano_obra,
                energia,
                gas,
                costos_indirectos,
                costo_total,
                costo_unitario,
                estado,
                COALESCE(merma, 0),
                COALESCE(desperdicio, 0),
                COALESCE(reproceso, 0)
            FROM ordenes_produccion
            WHERE substr(fecha, 1, 7) = ?
            ORDER BY id DESC
        """, (periodo,)).fetchall()

        periodicos = conexion.execute("""
            SELECT
                cp.periodo,
                c.codigo,
                c.nombre,
                c.clasificacion,
                c.comportamiento,
                COALESCE(
                    cc.codigo || ' - ' || cc.nombre,
                    ''
                ),
                cp.valor,
                cp.base_distribucion,
                cp.observaciones
            FROM costos_periodicos cp
            JOIN conceptos_costos c
              ON c.id = cp.concepto_id
            LEFT JOIN centros_costos cc
              ON cc.id = cp.centro_costo_id
            WHERE cp.periodo = ?
            ORDER BY cp.id
        """, (periodo,)).fetchall()

        consumos = conexion.execute("""
            SELECT
                op.numero,
                cp.producto,
                cp.presentacion,
                cp.lote,
                cp.cantidad,
                cp.costo_unitario,
                cp.costo_total
            FROM consumos_produccion cp
            JOIN ordenes_produccion op
              ON op.id = cp.orden_id
            WHERE substr(op.fecha, 1, 7) = ?
            ORDER BY op.numero, cp.id
        """, (periodo,)).fetchall()

    libro = Workbook()
    hoja = libro.active
    hoja.title = "Órdenes costeadas"

    encabezados_ordenes = [
        "Orden",
        "Fecha",
        "Producto",
        "Presentación",
        "Programada",
        "Producida",
        "Lote",
        "Responsable",
        "Materiales",
        "Mano de obra",
        "Energía",
        "Gas",
        "Indirectos",
        "Costo total",
        "Costo unitario",
        "Estado",
        "Merma",
        "Desperdicio",
        "Reproceso"
    ]
    hoja.append(encabezados_ordenes)

    for celda in hoja[1]:
        celda.font = Font(
            bold=True,
            color="FFFFFF"
        )
        celda.fill = PatternFill(
            "solid",
            fgColor="153B5B"
        )
        celda.alignment = Alignment(
            horizontal="center"
        )

    for fila in ordenes:
        hoja.append(list(fila))

    hoja.freeze_panes = "A2"
    hoja.auto_filter.ref = hoja.dimensions

    hoja_periodicos = libro.create_sheet(
        "Costos periódicos"
    )
    hoja_periodicos.append([
        "Periodo",
        "Código",
        "Concepto",
        "Clasificación",
        "Comportamiento",
        "Centro",
        "Valor",
        "Base de distribución",
        "Observaciones"
    ])

    for celda in hoja_periodicos[1]:
        celda.font = Font(
            bold=True,
            color="FFFFFF"
        )
        celda.fill = PatternFill(
            "solid",
            fgColor="0F5C8E"
        )

    for fila in periodicos:
        hoja_periodicos.append(list(fila))

    hoja_consumos = libro.create_sheet(
        "Consumos"
    )
    hoja_consumos.append([
        "Orden",
        "Producto",
        "Presentación",
        "Lote",
        "Cantidad",
        "Costo unitario",
        "Costo total"
    ])

    for celda in hoja_consumos[1]:
        celda.font = Font(
            bold=True,
            color="FFFFFF"
        )
        celda.fill = PatternFill(
            "solid",
            fgColor="15803D"
        )

    for fila in consumos:
        hoja_consumos.append(list(fila))

    libro.save(ruta)

    with conectar() as conexion:
        registrar_auditoria(
            conexion,
            "EXPORTAR COSTOS A EXCEL",
            ruta
        )
        conexion.commit()

    messagebox.showinfo(
        "Exportación",
        (
            "Reporte generado correctamente:\n\n"
            f"{ruta}"
        )
    )


def refrescar_todo():
    cargar_centros()
    cargar_conceptos()
    cargar_listas()
    cargar_costos_periodicos()
    cargar_ordenes_costeadas()
    cargar_rentabilidad()
    actualizar_indicadores()


# ============================================================
# INTERFAZ
# ============================================================

inicializar_bd()

ventana = tk.Tk()
ventana.title(
    "BME-ERP - Costos Industriales v1.0"
)
ventana.geometry("1500x900")
ventana.minsize(1180, 720)
ventana.configure(bg=C_FONDO)

try:
    ventana.state("zoomed")
except tk.TclError:
    pass

estilo = ttk.Style()
try:
    estilo.theme_use("clam")
except tk.TclError:
    pass

estilo.configure(
    "Treeview",
    rowheight=28,
    font=("Segoe UI", 9)
)
estilo.configure(
    "Treeview.Heading",
    font=("Segoe UI", 9, "bold")
)

header = tk.Frame(
    ventana,
    bg=C_OSCURO,
    height=82
)
header.pack(fill="x")
header.pack_propagate(False)

tk.Label(
    header,
    text="COSTOS INDUSTRIALES",
    font=("Segoe UI", 22, "bold"),
    bg=C_OSCURO,
    fg="white"
).pack(
    side="left",
    padx=24,
    pady=20
)

tk.Label(
    header,
    text=(
        "Costeo por orden, CIF, "
        "centros de costos y rentabilidad"
    ),
    font=("Segoe UI", 10),
    bg=C_OSCURO,
    fg="#BFDBFE"
).pack(
    side="right",
    padx=24
)

# KPIs
panel_kpi = tk.Frame(
    ventana,
    bg=C_FONDO
)
panel_kpi.pack(
    fill="x",
    padx=18,
    pady=(14, 5)
)

for columna in range(4):
    panel_kpi.grid_columnconfigure(
        columna,
        weight=1
    )


def crear_kpi(
    columna,
    titulo,
    color
):
    marco = tk.Frame(
        panel_kpi,
        bg=C_BLANCO,
        highlightbackground=C_BORDE,
        highlightthickness=1
    )
    marco.grid(
        row=0,
        column=columna,
        sticky="ew",
        padx=6
    )

    tk.Frame(
        marco,
        bg=color,
        width=5
    ).pack(
        side="left",
        fill="y"
    )

    interno = tk.Frame(
        marco,
        bg=C_BLANCO
    )
    interno.pack(
        fill="both",
        expand=True,
        padx=14,
        pady=10
    )

    tk.Label(
        interno,
        text=titulo,
        bg=C_BLANCO,
        fg=C_SUAVE,
        font=("Segoe UI", 8, "bold")
    ).pack(anchor="w")

    valor = tk.Label(
        interno,
        text="$0",
        bg=C_BLANCO,
        fg=C_TEXTO,
        font=("Segoe UI", 16, "bold")
    )
    valor.pack(
        anchor="w",
        pady=(3, 0)
    )

    return valor


lbl_kpi_costo_total = crear_kpi(
    0,
    "COSTO TOTAL DEL PERIODO",
    C_AZUL
)
lbl_kpi_costo_unitario = crear_kpi(
    1,
    "COSTO PROMEDIO UNITARIO",
    C_VERDE
)
lbl_kpi_costos_periodicos = crear_kpi(
    2,
    "COSTOS FIJOS Y CIF",
    C_NARANJA
)
lbl_kpi_ordenes = crear_kpi(
    3,
    "ÓRDENES COSTEADAS",
    C_MORADO
)

notebook = ttk.Notebook(ventana)
notebook.pack(
    fill="both",
    expand=True,
    padx=18,
    pady=10
)

tab_ordenes = tk.Frame(
    notebook,
    bg=C_FONDO
)
tab_periodicos = tk.Frame(
    notebook,
    bg=C_FONDO
)
tab_centros = tk.Frame(
    notebook,
    bg=C_FONDO
)
tab_conceptos = tk.Frame(
    notebook,
    bg=C_FONDO
)
tab_rentabilidad = tk.Frame(
    notebook,
    bg=C_FONDO
)

notebook.add(
    tab_ordenes,
    text="  Costeo por orden  "
)
notebook.add(
    tab_periodicos,
    text="  Costos fijos y CIF  "
)
notebook.add(
    tab_centros,
    text="  Centros de costos  "
)
notebook.add(
    tab_conceptos,
    text="  Conceptos  "
)
notebook.add(
    tab_rentabilidad,
    text="  Rentabilidad  "
)

# ------------------------------------------------------------
# COSTEO POR ORDEN
# ------------------------------------------------------------

filtros_orden = tk.Frame(
    tab_ordenes,
    bg=C_BLANCO
)
filtros_orden.pack(
    fill="x",
    padx=10,
    pady=10
)

tk.Label(
    filtros_orden,
    text="Periodo:",
    bg=C_BLANCO,
    fg=C_TEXTO
).pack(
    side="left",
    padx=(12, 5),
    pady=10
)

entry_periodo_orden = ttk.Entry(
    filtros_orden,
    width=10
)
entry_periodo_orden.pack(
    side="left",
    padx=5
)
entry_periodo_orden.insert(
    0,
    periodo_actual()
)

tk.Label(
    filtros_orden,
    text="Buscar:",
    bg=C_BLANCO,
    fg=C_TEXTO
).pack(
    side="left",
    padx=(15, 5)
)

entry_buscar_producto = ttk.Entry(
    filtros_orden,
    width=28
)
entry_buscar_producto.pack(
    side="left",
    padx=5
)

tk.Button(
    filtros_orden,
    text="Actualizar",
    command=cargar_ordenes_costeadas,
    bg=C_AZUL,
    fg="white",
    relief="flat",
    padx=14,
    pady=6
).pack(
    side="left",
    padx=5
)

tk.Button(
    filtros_orden,
    text="Exportar Excel",
    command=exportar_excel,
    bg=C_VERDE,
    fg="white",
    relief="flat",
    font=("Segoe UI", 9, "bold"),
    padx=14,
    pady=6
).pack(
    side="right",
    padx=12
)

tk.Button(
    filtros_orden,
    text="Ver detalle",
    command=ver_detalle_orden,
    bg=C_OSCURO,
    fg="white",
    relief="flat",
    font=("Segoe UI", 9, "bold"),
    padx=14,
    pady=6
).pack(
    side="right",
    padx=5
)

columnas_orden = (
    "Orden",
    "Fecha",
    "Producto",
    "Presentacion",
    "Cantidad",
    "Lote",
    "Materiales",
    "ManoObra",
    "EnergiaGas",
    "Indirectos",
    "CIF",
    "CostoTotal",
    "CostoUnitario",
    "Rendimiento",
    "Estado"
)

tabla_ordenes = ttk.Treeview(
    tab_ordenes,
    columns=columnas_orden,
    show="headings"
)

encabezados_orden = {
    "Orden": "Orden",
    "Fecha": "Fecha",
    "Producto": "Producto",
    "Presentacion": "Presentación",
    "Cantidad": "Cantidad",
    "Lote": "Lote",
    "Materiales": "Materiales",
    "ManoObra": "Mano de obra",
    "EnergiaGas": "Energía + gas",
    "Indirectos": "Indirectos OP",
    "CIF": "CIF distribuido",
    "CostoTotal": "Costo total",
    "CostoUnitario": "Costo unitario",
    "Rendimiento": "Rendimiento",
    "Estado": "Estado"
}

for columna in columnas_orden:
    tabla_ordenes.heading(
        columna,
        text=encabezados_orden[columna]
    )

anchos_orden = {
    "Orden": 155,
    "Fecha": 95,
    "Producto": 210,
    "Presentacion": 130,
    "Cantidad": 95,
    "Lote": 155,
    "Materiales": 110,
    "ManoObra": 105,
    "EnergiaGas": 110,
    "Indirectos": 105,
    "CIF": 105,
    "CostoTotal": 115,
    "CostoUnitario": 115,
    "Rendimiento": 100,
    "Estado": 90
}

for columna in columnas_orden:
    tabla_ordenes.column(
        columna,
        width=anchos_orden[columna],
        anchor=(
            "w"
            if columna in (
                "Producto",
                "Presentacion"
            )
            else "center"
        )
    )

tabla_ordenes.pack(
    fill="both",
    expand=True,
    padx=10,
    pady=(0, 10)
)

# ------------------------------------------------------------
# COSTOS PERIÓDICOS
# ------------------------------------------------------------

form_periodico = tk.LabelFrame(
    tab_periodicos,
    text="REGISTRO DE COSTOS FIJOS Y CIF",
    bg=C_BLANCO,
    fg=C_TEXTO,
    font=("Segoe UI", 10, "bold"),
    padx=12,
    pady=10
)
form_periodico.pack(
    fill="x",
    padx=10,
    pady=10
)

for columna in range(5):
    form_periodico.grid_columnconfigure(
        columna,
        weight=1
    )

campos_periodicos = [
    ("Periodo", 0),
    ("Concepto", 1),
    ("Centro de costos", 2),
    ("Valor", 3),
    ("Base distribución", 4)
]

for texto, columna in campos_periodicos:
    tk.Label(
        form_periodico,
        text=texto,
        bg=C_BLANCO,
        fg=C_SUAVE
    ).grid(
        row=0,
        column=columna,
        sticky="w"
    )

entry_periodo = ttk.Entry(
    form_periodico
)
entry_periodo.grid(
    row=1,
    column=0,
    sticky="ew",
    padx=(0, 8)
)
entry_periodo.insert(
    0,
    periodo_actual()
)

combo_periodico_concepto = ttk.Combobox(
    form_periodico,
    state="readonly"
)
combo_periodico_concepto.grid(
    row=1,
    column=1,
    sticky="ew",
    padx=(0, 8)
)

combo_periodico_centro = ttk.Combobox(
    form_periodico,
    state="readonly"
)
combo_periodico_centro.grid(
    row=1,
    column=2,
    sticky="ew",
    padx=(0, 8)
)

entry_periodico_valor = ttk.Entry(
    form_periodico
)
entry_periodico_valor.grid(
    row=1,
    column=3,
    sticky="ew",
    padx=(0, 8)
)

combo_base_distribucion = ttk.Combobox(
    form_periodico,
    values=[
        "PRODUCCIÓN",
        "ÓRDENES",
        "HORAS",
        "MANUAL"
    ],
    state="readonly"
)
combo_base_distribucion.grid(
    row=1,
    column=4,
    sticky="ew"
)
combo_base_distribucion.set(
    "PRODUCCIÓN"
)

tk.Label(
    form_periodico,
    text="Observaciones",
    bg=C_BLANCO,
    fg=C_SUAVE
).grid(
    row=2,
    column=0,
    sticky="w",
    pady=(10, 0)
)

entry_periodico_obs = tk.Text(
    form_periodico,
    height=2,
    relief="solid",
    bd=1
)
entry_periodico_obs.grid(
    row=3,
    column=0,
    columnspan=4,
    sticky="ew",
    padx=(0, 8)
)

tk.Button(
    form_periodico,
    text="Registrar costo",
    command=registrar_costo_periodico,
    bg=C_VERDE,
    fg="white",
    relief="flat",
    font=("Segoe UI", 9, "bold"),
    padx=18,
    pady=8
).grid(
    row=3,
    column=4,
    sticky="ew"
)

barra_periodicos = tk.Frame(
    tab_periodicos,
    bg=C_BLANCO
)
barra_periodicos.pack(
    fill="x",
    padx=10,
    pady=(0, 5)
)

tk.Button(
    barra_periodicos,
    text="Actualizar periodo",
    command=cargar_costos_periodicos,
    bg=C_AZUL,
    fg="white",
    relief="flat",
    padx=14,
    pady=6
).pack(
    side="left",
    padx=8,
    pady=8
)

tk.Button(
    barra_periodicos,
    text="Eliminar seleccionado",
    command=eliminar_costo_periodico,
    bg=C_ROJO,
    fg="white",
    relief="flat",
    padx=14,
    pady=6
).pack(
    side="right",
    padx=8,
    pady=8
)

columnas_periodicos = (
    "Periodo",
    "Codigo",
    "Concepto",
    "Clasificacion",
    "Comportamiento",
    "Centro",
    "Valor",
    "Base"
)

tabla_periodicos = ttk.Treeview(
    tab_periodicos,
    columns=columnas_periodicos,
    show="headings"
)

for columna in columnas_periodicos:
    tabla_periodicos.heading(
        columna,
        text=columna
    )

tabla_periodicos.pack(
    fill="both",
    expand=True,
    padx=10,
    pady=(0, 10)
)

# ------------------------------------------------------------
# CENTROS
# ------------------------------------------------------------

form_centro = tk.LabelFrame(
    tab_centros,
    text="CENTRO DE COSTOS",
    bg=C_BLANCO,
    fg=C_TEXTO,
    font=("Segoe UI", 10, "bold"),
    padx=12,
    pady=10
)
form_centro.pack(
    fill="x",
    padx=10,
    pady=10
)

for columna in range(4):
    form_centro.grid_columnconfigure(
        columna,
        weight=1
    )

for texto, columna in [
    ("Código", 0),
    ("Nombre", 1),
    ("Tipo", 2),
    ("Estado", 3)
]:
    tk.Label(
        form_centro,
        text=texto,
        bg=C_BLANCO,
        fg=C_SUAVE
    ).grid(
        row=0,
        column=columna,
        sticky="w"
    )

entry_centro_codigo = ttk.Entry(
    form_centro
)
entry_centro_codigo.grid(
    row=1,
    column=0,
    sticky="ew",
    padx=(0, 8)
)

entry_centro_nombre = ttk.Entry(
    form_centro
)
entry_centro_nombre.grid(
    row=1,
    column=1,
    sticky="ew",
    padx=(0, 8)
)

combo_centro_tipo = ttk.Combobox(
    form_centro,
    values=[
        "PRODUCCIÓN",
        "APOYO",
        "ADMINISTRATIVO",
        "COMERCIAL"
    ],
    state="readonly"
)
combo_centro_tipo.grid(
    row=1,
    column=2,
    sticky="ew",
    padx=(0, 8)
)

combo_centro_estado = ttk.Combobox(
    form_centro,
    values=[
        "ACTIVO",
        "INACTIVO"
    ],
    state="readonly"
)
combo_centro_estado.grid(
    row=1,
    column=3,
    sticky="ew"
)

lbl_modo_centro = tk.Label(
    form_centro,
    text="NUEVO CENTRO",
    bg=C_BLANCO,
    fg=C_VERDE,
    font=("Segoe UI", 9, "bold")
)
lbl_modo_centro.grid(
    row=2,
    column=0,
    sticky="w",
    pady=(10, 0)
)

tk.Label(
    form_centro,
    text="Observaciones",
    bg=C_BLANCO,
    fg=C_SUAVE
).grid(
    row=2,
    column=1,
    sticky="w",
    pady=(10, 0)
)

entry_centro_obs = tk.Text(
    form_centro,
    height=2,
    relief="solid",
    bd=1
)
entry_centro_obs.grid(
    row=3,
    column=1,
    columnspan=3,
    sticky="ew"
)

botones_centro = tk.Frame(
    form_centro,
    bg=C_BLANCO
)
botones_centro.grid(
    row=3,
    column=0,
    sticky="w"
)

tk.Button(
    botones_centro,
    text="Guardar",
    command=guardar_centro,
    bg=C_VERDE,
    fg="white",
    relief="flat",
    padx=14,
    pady=6
).pack(
    side="left",
    padx=(0, 5)
)

tk.Button(
    botones_centro,
    text="Limpiar",
    command=limpiar_centro,
    bg=C_OSCURO,
    fg="white",
    relief="flat",
    padx=14,
    pady=6
).pack(side="left")

tabla_centros = ttk.Treeview(
    tab_centros,
    columns=(
        "Codigo",
        "Nombre",
        "Tipo",
        "Estado"
    ),
    show="headings"
)

for columna in (
    "Codigo",
    "Nombre",
    "Tipo",
    "Estado"
):
    tabla_centros.heading(
        columna,
        text=columna
    )

tabla_centros.pack(
    fill="both",
    expand=True,
    padx=10,
    pady=(0, 10)
)
tabla_centros.bind(
    "<<TreeviewSelect>>",
    seleccionar_centro
)
tabla_centros.tag_configure(
    "activo",
    foreground=C_VERDE
)
tabla_centros.tag_configure(
    "inactivo",
    foreground=C_ROJO
)

# ------------------------------------------------------------
# CONCEPTOS
# ------------------------------------------------------------

form_concepto = tk.LabelFrame(
    tab_conceptos,
    text="CONCEPTO DE COSTO",
    bg=C_BLANCO,
    fg=C_TEXTO,
    font=("Segoe UI", 10, "bold"),
    padx=12,
    pady=10
)
form_concepto.pack(
    fill="x",
    padx=10,
    pady=10
)

for columna in range(6):
    form_concepto.grid_columnconfigure(
        columna,
        weight=1
    )

titulos_concepto = [
    "Código",
    "Nombre",
    "Clasificación",
    "Comportamiento",
    "Unidad",
    "Valor base"
]

for indice, titulo in enumerate(
    titulos_concepto
):
    tk.Label(
        form_concepto,
        text=titulo,
        bg=C_BLANCO,
        fg=C_SUAVE
    ).grid(
        row=0,
        column=indice,
        sticky="w"
    )

entry_concepto_codigo = ttk.Entry(
    form_concepto
)
entry_concepto_codigo.grid(
    row=1,
    column=0,
    sticky="ew",
    padx=(0, 8)
)

entry_concepto_nombre = ttk.Entry(
    form_concepto
)
entry_concepto_nombre.grid(
    row=1,
    column=1,
    sticky="ew",
    padx=(0, 8)
)

combo_clasificacion = ttk.Combobox(
    form_concepto,
    values=[
        "MATERIA PRIMA",
        "MANO DE OBRA",
        "CIF",
        "GASTO FIJO",
        "GASTO VARIABLE"
    ],
    state="readonly"
)
combo_clasificacion.grid(
    row=1,
    column=2,
    sticky="ew",
    padx=(0, 8)
)

combo_comportamiento = ttk.Combobox(
    form_concepto,
    values=[
        "FIJO",
        "VARIABLE",
        "MIXTO"
    ],
    state="readonly"
)
combo_comportamiento.grid(
    row=1,
    column=3,
    sticky="ew",
    padx=(0, 8)
)

combo_unidad = ttk.Combobox(
    form_concepto,
    values=[
        "MES",
        "HORA",
        "KG",
        "UND",
        "ORDEN"
    ],
    state="readonly"
)
combo_unidad.grid(
    row=1,
    column=4,
    sticky="ew",
    padx=(0, 8)
)

entry_valor_base = ttk.Entry(
    form_concepto
)
entry_valor_base.grid(
    row=1,
    column=5,
    sticky="ew"
)

tk.Label(
    form_concepto,
    text="Centro de costos",
    bg=C_BLANCO,
    fg=C_SUAVE
).grid(
    row=2,
    column=0,
    sticky="w",
    pady=(10, 0)
)

combo_concepto_centro = ttk.Combobox(
    form_concepto,
    state="readonly"
)
combo_concepto_centro.grid(
    row=3,
    column=0,
    columnspan=2,
    sticky="ew",
    padx=(0, 8)
)

tk.Label(
    form_concepto,
    text="Estado",
    bg=C_BLANCO,
    fg=C_SUAVE
).grid(
    row=2,
    column=2,
    sticky="w",
    pady=(10, 0)
)

combo_concepto_estado = ttk.Combobox(
    form_concepto,
    values=[
        "ACTIVO",
        "INACTIVO"
    ],
    state="readonly"
)
combo_concepto_estado.grid(
    row=3,
    column=2,
    sticky="ew",
    padx=(0, 8)
)

tk.Label(
    form_concepto,
    text="Observaciones",
    bg=C_BLANCO,
    fg=C_SUAVE
).grid(
    row=2,
    column=3,
    sticky="w",
    pady=(10, 0)
)

entry_concepto_obs = tk.Text(
    form_concepto,
    height=2,
    relief="solid",
    bd=1
)
entry_concepto_obs.grid(
    row=3,
    column=3,
    columnspan=3,
    sticky="ew"
)

lbl_modo_concepto = tk.Label(
    form_concepto,
    text="NUEVO CONCEPTO",
    bg=C_BLANCO,
    fg=C_VERDE,
    font=("Segoe UI", 9, "bold")
)
lbl_modo_concepto.grid(
    row=4,
    column=0,
    sticky="w",
    pady=(10, 0)
)

botones_concepto = tk.Frame(
    form_concepto,
    bg=C_BLANCO
)
botones_concepto.grid(
    row=4,
    column=4,
    columnspan=2,
    sticky="e",
    pady=(10, 0)
)

tk.Button(
    botones_concepto,
    text="Guardar concepto",
    command=guardar_concepto,
    bg=C_VERDE,
    fg="white",
    relief="flat",
    padx=14,
    pady=6
).pack(
    side="left",
    padx=4
)

tk.Button(
    botones_concepto,
    text="Limpiar",
    command=limpiar_concepto,
    bg=C_OSCURO,
    fg="white",
    relief="flat",
    padx=14,
    pady=6
).pack(
    side="left",
    padx=4
)

columnas_concepto = (
    "Codigo",
    "Nombre",
    "Clasificacion",
    "Comportamiento",
    "Unidad",
    "ValorBase",
    "Centro",
    "Estado"
)

tabla_conceptos = ttk.Treeview(
    tab_conceptos,
    columns=columnas_concepto,
    show="headings"
)

for columna in columnas_concepto:
    tabla_conceptos.heading(
        columna,
        text=columna
    )

tabla_conceptos.pack(
    fill="both",
    expand=True,
    padx=10,
    pady=(0, 10)
)
tabla_conceptos.bind(
    "<<TreeviewSelect>>",
    seleccionar_concepto
)

# ------------------------------------------------------------
# RENTABILIDAD
# ------------------------------------------------------------

form_precio = tk.LabelFrame(
    tab_rentabilidad,
    text="PRECIO DE VENTA DE REFERENCIA",
    bg=C_BLANCO,
    fg=C_TEXTO,
    font=("Segoe UI", 10, "bold"),
    padx=12,
    pady=10
)
form_precio.pack(
    fill="x",
    padx=10,
    pady=10
)

for columna in range(4):
    form_precio.grid_columnconfigure(
        columna,
        weight=1
    )

for texto, columna in [
    ("Producto", 0),
    ("Presentación", 1),
    ("Precio de venta", 2)
]:
    tk.Label(
        form_precio,
        text=texto,
        bg=C_BLANCO,
        fg=C_SUAVE
    ).grid(
        row=0,
        column=columna,
        sticky="w"
    )

entry_precio_producto = ttk.Entry(
    form_precio
)
entry_precio_producto.grid(
    row=1,
    column=0,
    sticky="ew",
    padx=(0, 8)
)

entry_precio_presentacion = ttk.Entry(
    form_precio
)
entry_precio_presentacion.grid(
    row=1,
    column=1,
    sticky="ew",
    padx=(0, 8)
)

entry_precio_venta = ttk.Entry(
    form_precio
)
entry_precio_venta.grid(
    row=1,
    column=2,
    sticky="ew",
    padx=(0, 8)
)

tk.Button(
    form_precio,
    text="Guardar precio",
    command=guardar_precio_referencia,
    bg=C_VERDE,
    fg="white",
    relief="flat",
    font=("Segoe UI", 9, "bold"),
    padx=16,
    pady=7
).grid(
    row=1,
    column=3,
    sticky="ew"
)

columnas_rentabilidad = (
    "Producto",
    "Presentacion",
    "Ordenes",
    "Produccion",
    "CostoTotal",
    "CostoPromedio",
    "PrecioVenta",
    "Utilidad",
    "Margen"
)

tabla_rentabilidad = ttk.Treeview(
    tab_rentabilidad,
    columns=columnas_rentabilidad,
    show="headings"
)

for columna in columnas_rentabilidad:
    tabla_rentabilidad.heading(
        columna,
        text=columna
    )

tabla_rentabilidad.pack(
    fill="both",
    expand=True,
    padx=10,
    pady=(0, 10)
)
tabla_rentabilidad.tag_configure(
    "positivo",
    foreground=C_VERDE
)
tabla_rentabilidad.tag_configure(
    "negativo",
    foreground=C_ROJO
)

# Barra inferior
barra = tk.Frame(
    ventana,
    bg=C_BLANCO,
    height=28
)
barra.pack(fill="x")

tk.Label(
    barra,
    text=f"Base de datos: {RUTA_DB}",
    bg=C_BLANCO,
    fg=C_SUAVE,
    font=("Segoe UI", 8)
).pack(
    side="left",
    padx=12
)

tk.Label(
    barra,
    text="BME-ERP Costos Industriales v1.0",
    bg=C_BLANCO,
    fg=C_SUAVE,
    font=("Segoe UI", 8)
).pack(
    side="right",
    padx=12
)

# Inicio
limpiar_centro()
limpiar_concepto()
refrescar_todo()

ventana.mainloop()
