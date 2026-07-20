# ==========================================================
# ERP CAFÉ ALTO DE LA CRUZ
# MÓDULO INVENTARIO v1.0
# SECCIÓN 1: BASE DE DATOS Y FUNCIONES
# ==========================================================
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog

import sqlite3
from datetime import datetime

from PIL import Image, ImageTk

import os
import csv
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ==========================================================
# CONFIGURACIÓN
# ==========================================================

RUTA_DB = r"C:\Users\jrive\visual\erp_cafe.db"

conexion = sqlite3.connect(RUTA_DB)
cursor = conexion.cursor()

registro_editando = None


# ==========================================================
# BASE DE DATOS
# ==========================================================

def inicializar_bd():
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS inventario (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            producto TEXT,
            presentacion TEXT,
            cantidad REAL DEFAULT 0,
            lote TEXT DEFAULT '',
            costo_unitario REAL DEFAULT 0,
            fecha_ingreso TEXT DEFAULT '',
            numero_despacho TEXT DEFAULT '',
            stock_minimo INTEGER DEFAULT 0,
            costo REAL DEFAULT 0,
            fecha TEXT DEFAULT '',
            despacho TEXT DEFAULT ''
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS kardex (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha TEXT,
            producto TEXT,
            presentacion TEXT,
            movimiento TEXT,
            entrada REAL DEFAULT 0,
            salida REAL DEFAULT 0,
            saldo REAL DEFAULT 0,
            costo_unitario REAL DEFAULT 0,
            lote TEXT DEFAULT '',
            origen TEXT DEFAULT '',
            observaciones TEXT DEFAULT ''
        )
    """)

    conexion.commit()


# ==========================================================
# UTILIDADES
# ==========================================================

def obtener_fecha_actual():
    return datetime.now().strftime("%Y-%m-%d")


def obtener_fecha_hora():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def convertir_numero(valor, defecto=0):
    try:
        return float(valor)
    except:
        return defecto


def obtener_id_seleccionado():
    seleccionado = tabla.selection()

    if not seleccionado:
        messagebox.showwarning("Atención", "Seleccione un registro.")
        return None

    valores = tabla.item(seleccionado[0], "values")
    return valores[0]


def limpiar_formulario():
    global registro_editando

    registro_editando = None

    combo_producto.set("")
    combo_presentacion.set("")
    txt_cantidad.delete(0, tk.END)
    txt_stock.delete(0, tk.END)
    txt_costo.delete(0, tk.END)
    txt_lote.delete(0, tk.END)
    txt_fecha.delete(0, tk.END)
    txt_despacho.delete(0, tk.END)

    lbl_valor.config(text="VALOR DEL INVENTARIO : $ 0.00")


def calcular_valor_formulario(event=None):
    cantidad = convertir_numero(txt_cantidad.get())
    costo = convertir_numero(txt_costo.get())
    valor = cantidad * costo

    lbl_valor.config(text=f"VALOR DEL INVENTARIO : $ {valor:,.0f}")


# ==========================================================
# INDICADORES
# ==========================================================

def actualizar_indicadores():
    cursor.execute("SELECT COUNT(*) FROM inventario")
    referencias = cursor.fetchone()[0]

    cursor.execute("SELECT IFNULL(SUM(cantidad),0) FROM inventario")
    unidades = cursor.fetchone()[0]

    cursor.execute("""
        SELECT IFNULL(SUM(cantidad * COALESCE(costo_unitario, costo, 0)),0)
        FROM inventario
    """)
    valor_total = cursor.fetchone()[0]

    cursor.execute("""
        SELECT COUNT(*)
        FROM inventario
        WHERE cantidad <= stock_minimo
          AND cantidad > 0
          AND stock_minimo > 0
    """)
    stock_bajo = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM inventario WHERE cantidad = 0")
    agotados = cursor.fetchone()[0]

    lbl_referencias.config(text=str(referencias))
    lbl_unidades.config(text=f"{unidades:,.0f}")
    lbl_valor_total.config(text=f"${valor_total:,.0f}")
    lbl_stock.config(text=str(stock_bajo))
    lbl_agotados.config(text=str(agotados))


# ==========================================================
# CARGAR INVENTARIO
# ==========================================================

def cargar_inventario():
    for item in tabla.get_children():
        tabla.delete(item)

    cursor.execute("""
        SELECT
            id,
            producto,
            presentacion,
            cantidad,
            COALESCE(lote, ''),
            COALESCE(costo_unitario, costo, 0),
            cantidad * COALESCE(costo_unitario, costo, 0),
            COALESCE(fecha_ingreso, fecha, ''),
            stock_minimo
        FROM inventario
        ORDER BY producto, presentacion
    """)

    registros = cursor.fetchall()

    for fila in registros:
        id_reg, producto, presentacion, cantidad, lote, costo, valor, fecha, stock_minimo = fila

        tag = "normal"

        if cantidad == 0:
            tag = "agotado"
        elif stock_minimo and cantidad <= stock_minimo:
            tag = "bajo"

        tabla.insert("", tk.END, values=(
            id_reg,
            producto,
            presentacion,
            f"{cantidad:,.0f}",
            lote,
            f"{costo:,.0f}",
            f"{valor:,.0f}",
            fecha
        ), tags=(tag,))

    actualizar_indicadores()


# ==========================================================
# BUSCAR
# ==========================================================

def buscar_inventario():
    criterio = simpledialog.askstring(
        "Buscar inventario",
        "Ingrese producto, presentación o lote:"
    )

    if criterio is None:
        return

    criterio = criterio.strip()

    if criterio == "":
        cargar_inventario()
        return

    for item in tabla.get_children():
        tabla.delete(item)

    cursor.execute("""
        SELECT
            id,
            producto,
            presentacion,
            cantidad,
            COALESCE(lote, ''),
            COALESCE(costo_unitario, costo, 0),
            cantidad * COALESCE(costo_unitario, costo, 0),
            COALESCE(fecha_ingreso, fecha, ''),
            stock_minimo
        FROM inventario
        WHERE producto LIKE ?
           OR presentacion LIKE ?
           OR lote LIKE ?
        ORDER BY producto, presentacion
    """, (
        f"%{criterio}%",
        f"%{criterio}%",
        f"%{criterio}%"
    ))

    registros = cursor.fetchall()

    for fila in registros:
        id_reg, producto, presentacion, cantidad, lote, costo, valor, fecha, stock_minimo = fila

        tag = "normal"

        if cantidad == 0:
            tag = "agotado"
        elif stock_minimo and cantidad <= stock_minimo:
            tag = "bajo"

        tabla.insert("", tk.END, values=(
            id_reg,
            producto,
            presentacion,
            f"{cantidad:,.0f}",
            lote,
            f"{costo:,.0f}",
            f"{valor:,.0f}",
            fecha
        ), tags=(tag,))

    if len(registros) == 0:
        messagebox.showinfo("Buscar", "No se encontraron resultados.")


# ==========================================================
# KARDEX
# ==========================================================

def registrar_kardex(producto, presentacion, lote, movimiento, entrada, salida, costo, origen, observacion):
    cursor.execute("""
        SELECT IFNULL(SUM(entrada - salida), 0)
        FROM kardex
        WHERE producto = ?
          AND presentacion = ?
          AND COALESCE(lote, '') = ?
    """, (producto, presentacion, lote))

    saldo_anterior = cursor.fetchone()[0]
    saldo_nuevo = saldo_anterior + entrada - salida

    cursor.execute("""
        INSERT INTO kardex (
            fecha,
            producto,
            presentacion,
            movimiento,
            entrada,
            salida,
            saldo,
            costo_unitario,
            lote,
            origen,
            observaciones
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        obtener_fecha_hora(),
        producto,
        presentacion,
        movimiento,
        entrada,
        salida,
        saldo_nuevo,
        costo,
        lote,
        origen,
        observacion
    ))


# ==========================================================
# AGREGAR INVENTARIO
# ==========================================================

def agregar_inventario():
    producto = combo_producto.get().strip()
    presentacion = combo_presentacion.get().strip()

    if producto == "":
        messagebox.showerror("Error", "Seleccione un producto.")
        return

    if presentacion == "":
        messagebox.showerror("Error", "Seleccione una presentación.")
        return

    try:
        cantidad = float(txt_cantidad.get())
        stock_minimo = int(txt_stock.get() or 0)
        costo = float(txt_costo.get() or 0)
    except:
        messagebox.showerror("Error", "Cantidad, stock mínimo y costo deben ser números.")
        return

    if cantidad <= 0:
        messagebox.showerror("Error", "La cantidad debe ser mayor que cero.")
        return

    lote = txt_lote.get().strip()
    fecha = txt_fecha.get().strip() or obtener_fecha_actual()
    despacho = txt_despacho.get().strip()

    try:
        cursor.execute("""
            SELECT id
            FROM inventario
            WHERE producto = ?
              AND presentacion = ?
              AND COALESCE(lote, '') = ?
        """, (producto, presentacion, lote))

        existe = cursor.fetchone()

        if existe:
            id_inventario = existe[0]

            cursor.execute("""
                UPDATE inventario
                SET cantidad = cantidad + ?,
                    stock_minimo = ?,
                    costo = ?,
                    costo_unitario = ?,
                    fecha = ?,
                    fecha_ingreso = ?,
                    despacho = ?,
                    numero_despacho = ?
                WHERE id = ?
            """, (
                cantidad,
                stock_minimo,
                costo,
                costo,
                fecha,
                fecha,
                despacho,
                despacho,
                id_inventario
            ))

            operacion = "ACTUALIZACIÓN INVENTARIO"

        else:
            cursor.execute("""
                INSERT INTO inventario (
                    producto,
                    presentacion,
                    cantidad,
                    stock_minimo,
                    costo,
                    costo_unitario,
                    lote,
                    fecha,
                    fecha_ingreso,
                    despacho,
                    numero_despacho
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                producto,
                presentacion,
                cantidad,
                stock_minimo,
                costo,
                costo,
                lote,
                fecha,
                fecha,
                despacho,
                despacho
            ))

            operacion = "ENTRADA INVENTARIO"

        registrar_kardex(
            producto,
            presentacion,
            lote,
            operacion,
            cantidad,
            0,
            costo,
            despacho,
            "Entrada registrada desde módulo Inventario"
        )

        conexion.commit()
        limpiar_formulario()
        cargar_inventario()

        messagebox.showinfo("Éxito", "Inventario actualizado correctamente.")

    except Exception as e:
        conexion.rollback()
        messagebox.showerror("Error", f"No se pudo agregar inventario:\n{e}")


# ==========================================================
# EDITAR INVENTARIO
# ==========================================================

def editar_inventario():
    id_registro = obtener_id_seleccionado()

    if id_registro is None:
        return

    cursor.execute("""
        SELECT
            producto,
            presentacion,
            cantidad,
            stock_minimo,
            COALESCE(costo_unitario, costo, 0),
            COALESCE(lote, ''),
            COALESCE(fecha_ingreso, fecha, ''),
            COALESCE(numero_despacho, despacho, '')
        FROM inventario
        WHERE id = ?
    """, (id_registro,))

    fila = cursor.fetchone()

    if not fila:
        messagebox.showerror("Error", "No se encontró el registro.")
        return

    nueva_cantidad = simpledialog.askfloat(
        "Editar cantidad",
        "Nueva cantidad:",
        initialvalue=fila[2]
    )

    if nueva_cantidad is None:
        return

    nuevo_costo = simpledialog.askfloat(
        "Editar costo",
        "Nuevo costo unitario:",
        initialvalue=fila[4]
    )

    if nuevo_costo is None:
        return

    try:
        cursor.execute("""
            UPDATE inventario
            SET cantidad = ?,
                costo = ?,
                costo_unitario = ?
            WHERE id = ?
        """, (
            nueva_cantidad,
            nuevo_costo,
            nuevo_costo,
            id_registro
        ))

        conexion.commit()
        cargar_inventario()

        messagebox.showinfo("Éxito", "Registro editado correctamente.")

    except Exception as e:
        conexion.rollback()
        messagebox.showerror("Error", f"No se pudo editar:\n{e}")


# ==========================================================
# ELIMINAR INVENTARIO
# ==========================================================

def eliminar_inventario():
    id_registro = obtener_id_seleccionado()

    if id_registro is None:
        return

    confirmar = messagebox.askyesno(
        "Confirmar eliminación",
        "¿Está seguro de eliminar este registro del inventario?"
    )

    if not confirmar:
        return

    try:
        cursor.execute("""
            SELECT
                producto,
                presentacion,
                cantidad,
                COALESCE(lote, ''),
                COALESCE(costo_unitario, costo, 0)
            FROM inventario
            WHERE id = ?
        """, (id_registro,))

        fila = cursor.fetchone()

        if fila:
            producto, presentacion, cantidad, lote, costo = fila

            registrar_kardex(
                producto,
                presentacion,
                lote,
                "ELIMINACIÓN INVENTARIO",
                0,
                cantidad,
                costo,
                "ELIMINACIÓN",
                "Registro eliminado desde módulo Inventario"
            )

        cursor.execute("DELETE FROM inventario WHERE id = ?", (id_registro,))
        conexion.commit()

        cargar_inventario()
        messagebox.showinfo("Éxito", "Registro eliminado correctamente.")

    except Exception as e:
        conexion.rollback()
        messagebox.showerror("Error", f"No se pudo eliminar:\n{e}")


# ==========================================================
# REGISTRAR SALIDA
# ==========================================================

def registrar_salida_inventario():
    id_registro = obtener_id_seleccionado()

    if id_registro is None:
        return

    cursor.execute("""
        SELECT
            producto,
            presentacion,
            cantidad,
            COALESCE(lote, ''),
            COALESCE(costo_unitario, costo, 0)
        FROM inventario
        WHERE id = ?
    """, (id_registro,))

    fila = cursor.fetchone()

    if not fila:
        messagebox.showerror("Error", "No se encontró el producto.")
        return

    producto, presentacion, existencia_actual, lote, costo = fila

    cantidad_salida = simpledialog.askfloat(
        "Registrar salida",
        f"Existencia actual: {existencia_actual}\nCantidad a retirar:"
    )

    if cantidad_salida is None:
        return

    if cantidad_salida <= 0:
        messagebox.showerror("Error", "La salida debe ser mayor que cero.")
        return

    if cantidad_salida > existencia_actual:
        messagebox.showerror("Error", "No hay inventario suficiente.")
        return

    motivo = simpledialog.askstring(
        "Motivo de salida",
        "Motivo: venta, muestra, merma, baja, ajuste:"
    )

    if motivo is None or motivo.strip() == "":
        motivo = "SALIDA INVENTARIO"

    saldo_nuevo = existencia_actual - cantidad_salida

    try:
        cursor.execute("""
            UPDATE inventario
            SET cantidad = ?
            WHERE id = ?
        """, (saldo_nuevo, id_registro))

        registrar_kardex(
            producto,
            presentacion,
            lote,
            "SALIDA INVENTARIO",
            0,
            cantidad_salida,
            costo,
            "SALIDA MANUAL",
            motivo
        )

        conexion.commit()
        cargar_inventario()

        messagebox.showinfo("Éxito", "Salida registrada correctamente.")

    except Exception as e:
        conexion.rollback()
        messagebox.showerror("Error", f"No se pudo registrar la salida:\n{e}")
        # ==========================================================
# SECCIÓN 2A
# INTERFAZ PRINCIPAL
# ==========================================================

inicializar_bd()

ventana = tk.Tk()
ventana.title("ERP CAFÉ ALTO DE LA CRUZ")
ventana.state("zoomed")
ventana.configure(bg="#E9EEF4")

# ==========================================================
# ESTILOS
# ==========================================================

style = ttk.Style()
style.theme_use("clam")

style.configure(
    "Treeview",
    font=("Segoe UI",10),
    rowheight=28
)

style.configure(
    "Treeview.Heading",
    font=("Segoe UI",10,"bold")
)

# ==========================================================
# HEADER
# ==========================================================

frame_header = tk.Frame(
    ventana,
    bg="#0F4C81",
    height=90
)

frame_header.pack(fill="x")

tk.Label(
    frame_header,
    text="ERP CAFÉ ALTO DE LA CRUZ",
    bg="#0F4C81",
    fg="white",
    font=("Segoe UI",22,"bold")
).pack(pady=(12,0))

tk.Label(
    frame_header,
    text="MÓDULO INVENTARIO",
    bg="#0F4C81",
    fg="white",
    font=("Segoe UI",12)
).pack()

# ==========================================================
# FRAME REGISTRO
# ==========================================================

frame_registro = tk.LabelFrame(
    ventana,
    text="REGISTRO DE INVENTARIO",
    bg="white",
    padx=20,
    pady=15,
    font=("Segoe UI",11,"bold")
)

frame_registro.pack(fill="x",padx=15,pady=10)

# ==========================================================
# PRODUCTO
# ==========================================================

tk.Label(
    frame_registro,
    text="Producto",
    bg="white",
    font=("Segoe UI",10,"bold")
).grid(row=0,column=0,sticky="w")

combo_producto = ttk.Combobox(
    frame_registro,
    width=40,
    state="readonly"
)

combo_producto["values"]=(
    "Café Especial",
    "Café Premium",
    "Café Tradicional",
    "Premium",
    "Tradicional"
)

combo_producto.grid(
    row=0,
    column=1,
    padx=10,
    pady=5
)

# ==========================================================
# PRESENTACIÓN
# ==========================================================

tk.Label(
    frame_registro,
    text="Presentación",
    bg="white",
    font=("Segoe UI",10,"bold")
).grid(row=0,column=2,sticky="w")

combo_presentacion=ttk.Combobox(
    frame_registro,
    width=20,
    state="readonly"
)

combo_presentacion["values"]=(
    "125 g",
    "250 g",
    "500 g",
    "1000 g"
)

combo_presentacion.grid(
    row=0,
    column=3,
    padx=10,
    pady=5
)

# ==========================================================
# CANTIDAD
# ==========================================================

tk.Label(
    frame_registro,
    text="Cantidad",
    bg="white",
    font=("Segoe UI",10,"bold")
).grid(row=1,column=0,sticky="w")

txt_cantidad=tk.Entry(
    frame_registro,
    width=30
)

txt_cantidad.grid(
    row=1,
    column=1,
    padx=10,
    pady=5
)

txt_cantidad.bind(
    "<KeyRelease>",
    calcular_valor_formulario
)

# ==========================================================
# STOCK
# ==========================================================

tk.Label(
    frame_registro,
    text="Stock mínimo",
    bg="white",
    font=("Segoe UI",10,"bold")
).grid(row=1,column=2,sticky="w")

txt_stock=tk.Entry(
    frame_registro,
    width=20
)

txt_stock.grid(
    row=1,
    column=3,
    padx=10,
    pady=5
)

# ==========================================================
# COSTO
# ==========================================================

tk.Label(
    frame_registro,
    text="Costo Unitario",
    bg="white",
    font=("Segoe UI",10,"bold")
).grid(row=2,column=0,sticky="w")

txt_costo=tk.Entry(
    frame_registro,
    width=30
)

txt_costo.grid(
    row=2,
    column=1,
    padx=10,
    pady=5
)

txt_costo.bind(
    "<KeyRelease>",
    calcular_valor_formulario
)

# ==========================================================
# LOTE
# ==========================================================

tk.Label(
    frame_registro,
    text="Lote",
    bg="white",
    font=("Segoe UI",10,"bold")
).grid(row=2,column=2,sticky="w")

txt_lote=tk.Entry(
    frame_registro,
    width=20
)

txt_lote.grid(
    row=2,
    column=3,
    padx=10,
    pady=5
)

# ==========================================================
# FECHA
# ==========================================================

tk.Label(
    frame_registro,
    text="Fecha",
    bg="white",
    font=("Segoe UI",10,"bold")
).grid(row=3,column=0,sticky="w")

txt_fecha=tk.Entry(
    frame_registro,
    width=30
)

txt_fecha.grid(
    row=3,
    column=1,
    padx=10,
    pady=5
)

# ==========================================================
# DESPACHO
# ==========================================================

tk.Label(
    frame_registro,
    text="No. Despacho",
    bg="white",
    font=("Segoe UI",10,"bold")
).grid(row=3,column=2,sticky="w")

txt_despacho=tk.Entry(
    frame_registro,
    width=20
)

txt_despacho.grid(
    row=3,
    column=3,
    padx=10,
    pady=5
)

# ==========================================================
# VALOR
# ==========================================================

lbl_valor=tk.Label(
    frame_registro,
    text="VALOR DEL INVENTARIO : $0",
    bg="white",
    fg="#0F4C81",
    font=("Segoe UI",12,"bold")
)

lbl_valor.grid(
    row=4,
    column=0,
    columnspan=4,
    pady=15
)

# ==========================================================
# BOTONES
# ==========================================================

frame_botones=tk.Frame(
    frame_registro,
    bg="white"
)

frame_botones.grid(
    row=5,
    column=0,
    columnspan=4,
    pady=10
)

tk.Button(
    frame_botones,
    text="AGREGAR INVENTARIO",
    width=22,
    bg="#0F4C81",
    fg="white",
    font=("Segoe UI",10,"bold"),
    command=agregar_inventario
).pack(side="left",padx=8)

tk.Button(
    frame_botones,
    text="LIMPIAR",
    width=14,
    command=limpiar_formulario
).pack(side="left",padx=8)

tk.Button(
    frame_botones,
    text="CANCELAR",
    width=14,
    command=limpiar_formulario
).pack(side="left",padx=8)
# ==========================================================
# SECCIÓN 2B
# ACCIONES, TREEVIEW E INDICADORES
# ==========================================================

# ==========================================================
# FRAME ACCIONES
# ==========================================================

frame_acciones = tk.LabelFrame(
    ventana,
    text="ACCIONES",
    bg="white",
    padx=15,
    pady=10,
    font=("Segoe UI", 11, "bold")
)

frame_acciones.pack(
    fill="x",
    padx=15,
    pady=(0, 10)
)

tk.Button(
    frame_acciones,
    text="Actualizar",
    width=14,
    bg="#0F4C81",
    fg="white",
    font=("Segoe UI", 9, "bold"),
    command=cargar_inventario
).pack(side="left", padx=6)

tk.Button(
    frame_acciones,
    text="Buscar",
    width=14,
    font=("Segoe UI", 9),
    command=buscar_inventario
).pack(side="left", padx=6)

tk.Button(
    frame_acciones,
    text="Editar",
    width=14,
    font=("Segoe UI", 9),
    command=editar_inventario
).pack(side="left", padx=6)

tk.Button(
    frame_acciones,
    text="Eliminar",
    width=14,
    font=("Segoe UI", 9),
    command=eliminar_inventario
).pack(side="left", padx=6)

tk.Button(
    frame_acciones,
    text="Salida",
    width=14,
    font=("Segoe UI", 9),
    command=registrar_salida_inventario
).pack(side="left", padx=6)

tk.Button(
    frame_acciones,
    text="Exportar Excel",
    width=16,
    font=("Segoe UI", 9),
    command=lambda: messagebox.showinfo(
        "Pendiente",
        "La exportación a Excel se activará en la Sección 3."
    )
).pack(side="left", padx=6)

tk.Button(
    frame_acciones,
    text="Imprimir",
    width=14,
    font=("Segoe UI", 9),
    command=lambda: messagebox.showinfo(
        "Pendiente",
        "La impresión se activará en la Sección 3."
    )
).pack(side="left", padx=6)


# ==========================================================
# FRAME TABLA
# ==========================================================

frame_tabla = tk.LabelFrame(
    ventana,
    text="INVENTARIO ACTUAL",
    bg="white",
    font=("Segoe UI", 11, "bold")
)

frame_tabla.pack(
    fill="both",
    expand=True,
    padx=15,
    pady=(0, 10)
)

# ==========================================================
# TREEVIEW
# ==========================================================

columnas = (
    "ID",
    "Producto",
    "Presentación",
    "Cantidad",
    "Lote",
    "Costo",
    "Valor",
    "Fecha"
)

tabla = ttk.Treeview(
    frame_tabla,
    columns=columnas,
    show="headings"
)

for col in columnas:
    tabla.heading(col, text=col)

tabla.column(
    "ID",
    width=60,
    anchor="center"
)

tabla.column(
    "Producto",
    width=250,
    anchor="w"
)

tabla.column(
    "Presentación",
    width=120,
    anchor="center"
)

tabla.column(
    "Cantidad",
    width=110,
    anchor="center"
)

tabla.column(
    "Lote",
    width=140,
    anchor="center"
)

tabla.column(
    "Costo",
    width=130,
    anchor="e"
)

tabla.column(
    "Valor",
    width=150,
    anchor="e"
)

tabla.column(
    "Fecha",
    width=140,
    anchor="center"
)

# ==========================================================
# COLORES DEL TREEVIEW
# ==========================================================

tabla.tag_configure(
    "normal",
    background="white"
)

tabla.tag_configure(
    "bajo",
    background="#FFF3CD"
)

tabla.tag_configure(
    "agotado",
    background="#F8D7DA"
)

# ==========================================================
# SCROLL
# ==========================================================

scroll_y = ttk.Scrollbar(
    frame_tabla,
    orient="vertical",
    command=tabla.yview
)

scroll_x = ttk.Scrollbar(
    frame_tabla,
    orient="horizontal",
    command=tabla.xview
)

tabla.configure(
    yscrollcommand=scroll_y.set,
    xscrollcommand=scroll_x.set
)

tabla.pack(
    side="left",
    fill="both",
    expand=True
)

scroll_y.pack(
    side="right",
    fill="y"
)

scroll_x.pack(
    side="bottom",
    fill="x"
)

# ==========================================================
# FRAME INDICADORES
# ==========================================================

frame_indicadores = tk.Frame(
    ventana,
    bg="#E9EEF4",
    height=90
)

frame_indicadores.pack(
    fill="x",
    padx=15,
    pady=(0, 10)
)

frame_indicadores.pack_propagate(False)

for i in range(5):
    frame_indicadores.grid_columnconfigure(i, weight=1)


def crear_card_indicador(columna, titulo, color, tamano=22):
    card = tk.Frame(
        frame_indicadores,
        bg="white",
        bd=1,
        relief="solid"
    )

    card.grid(
        row=0,
        column=columna,
        padx=8,
        pady=8,
        sticky="nsew"
    )

    tk.Label(
        card,
        text=titulo,
        bg="white",
        fg="#555555",
        font=("Segoe UI", 10, "bold")
    ).pack(pady=(10, 2))

    label = tk.Label(
        card,
        text="0",
        bg="white",
        fg=color,
        font=("Segoe UI", tamano, "bold")
    )

    label.pack()

    return label


lbl_referencias = crear_card_indicador(
    0,
    "REFERENCIAS",
    "#0F4C81"
)

lbl_unidades = crear_card_indicador(
    1,
    "UNIDADES",
    "#0F4C81"
)

lbl_valor_total = crear_card_indicador(
    2,
    "VALOR INVENTARIO",
    "#0F4C81",
    18
)

lbl_stock = crear_card_indicador(
    3,
    "STOCK BAJO",
    "#C47A00"
)

lbl_agotados = crear_card_indicador(
    4,
    "AGOTADOS",
    "red"
)
# ==========================================================
# SECCIÓN 3
# EXPORTAR, IMPRIMIR, DOBLE CLIC E INICIO
# ==========================================================

def exportar_excel():
    archivo = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Archivo Excel", "*.xlsx")],
        initialfile=f"Inventario_{obtener_fecha_actual()}.xlsx"
    )

    if not archivo:
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    ws.merge_cells("A1:H1")
    ws["A1"] = "ERP CAFÉ ALTO DE LA CRUZ"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:H2")
    ws["A2"] = f"REPORTE DE INVENTARIO - {obtener_fecha_hora()}"
    ws["A2"].alignment = Alignment(horizontal="center")

    encabezados = ["ID", "Producto", "Presentación", "Cantidad", "Lote", "Costo", "Valor", "Fecha"]
    ws.append([])
    ws.append(encabezados)

    fila_encabezado = 4
    for celda in ws[fila_encabezado]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="0F4C81")
        celda.alignment = Alignment(horizontal="center")

    for item in tabla.get_children():
        ws.append(list(tabla.item(item, "values")))

    fila_total = ws.max_row + 2
    ws[f"A{fila_total}"] = "TOTAL REFERENCIAS"
    ws[f"B{fila_total}"] = lbl_referencias.cget("text")

    ws[f"D{fila_total}"] = "TOTAL UNIDADES"
    ws[f"E{fila_total}"] = lbl_unidades.cget("text")

    ws[f"G{fila_total}"] = "VALOR INVENTARIO"
    ws[f"H{fila_total}"] = lbl_valor_total.cget("text")

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 3

    borde = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        for cell in row:
            cell.border = borde
            cell.alignment = Alignment(horizontal="center")

    wb.save(archivo)

    messagebox.showinfo("Éxito", "Inventario exportado a Excel correctamente.")
    os.startfile(archivo)


def imprimir_inventario():
    archivo = filedialog.asksaveasfilename(
        defaultextension=".pdf",
        filetypes=[("Archivo PDF", "*.pdf")],
        initialfile=f"Inventario_{obtener_fecha_actual()}.pdf"
    )

    if not archivo:
        return

    doc = SimpleDocTemplate(
        archivo,
        pagesize=landscape(letter),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )

    estilos = getSampleStyleSheet()
    elementos = []

    titulo = Paragraph("<b>ERP CAFÉ ALTO DE LA CRUZ</b>", estilos["Title"])
    subtitulo = Paragraph("REPORTE DE INVENTARIO", estilos["Heading2"])
    fecha = Paragraph(f"Fecha: {obtener_fecha_hora()}", estilos["Normal"])

    elementos.append(titulo)
    elementos.append(subtitulo)
    elementos.append(fecha)
    elementos.append(Spacer(1, 12))

    datos = [["ID", "Producto", "Presentación", "Cantidad", "Lote", "Costo", "Valor", "Fecha"]]

    for item in tabla.get_children():
        datos.append(list(tabla.item(item, "values")))

    tabla_pdf = Table(datos, repeatRows=1)

    tabla_pdf.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F4C81")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))

    elementos.append(tabla_pdf)
    elementos.append(Spacer(1, 15))

    resumen = Paragraph(
        f"""
        <b>Referencias:</b> {lbl_referencias.cget("text")} &nbsp;&nbsp;&nbsp;
        <b>Unidades:</b> {lbl_unidades.cget("text")} &nbsp;&nbsp;&nbsp;
        <b>Valor Inventario:</b> {lbl_valor_total.cget("text")}
        """,
        estilos["Normal"]
    )

    elementos.append(resumen)

    doc.build(elementos)

    messagebox.showinfo("Éxito", "PDF generado correctamente.")
    os.startfile(archivo)

def doble_clic_editar(event):
    editar_inventario()


tabla.bind("<Double-1>", doble_clic_editar)


for boton in frame_acciones.winfo_children():
    if boton.cget("text") == "Exportar Excel":
        boton.config(command=exportar_excel)

    if boton.cget("text") == "Imprimir":
        boton.config(command=imprimir_inventario)


# ==========================================================
# INICIO DEL MÓDULO
# ==========================================================

cargar_inventario()

ventana.mainloop()