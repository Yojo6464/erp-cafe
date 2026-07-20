"""
BME-ERP - Libro Diario Profesional
Archivo: libro_diario_profesional.py
"""

from __future__ import annotations

import os
import tkinter as tk
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from consultas_contables import (
    RUTA_DB,
    consultar_comprobante,
    consultar_libro_diario,
    listar_cuentas,
    listar_modulos,
    listar_tipos_comprobante,
    validar_estructura,
)

COLOR_FONDO = "#EEF3F8"
COLOR_TARJETA = "#FFFFFF"
COLOR_AZUL = "#0F5C8E"
COLOR_OSCURO = "#153B5B"
COLOR_VERDE = "#15803D"
COLOR_ROJO = "#B42318"
COLOR_GRIS = "#64748B"
COLOR_BORDE = "#D7E0E8"
COLOR_TEXTO = "#1F2937"

filas_actuales: list[dict] = []


def moneda(valor: float) -> str:
    return f"${float(valor or 0):,.2f}"


def validar_fecha(texto: str, nombre: str) -> str:
    texto = texto.strip()

    if not texto:
        return ""

    try:
        datetime.strptime(texto, "%Y-%m-%d")
    except ValueError as error:
        raise ValueError(
            f"{nombre} debe tener formato AAAA-MM-DD."
        ) from error

    return texto


def obtener_codigo_cuenta() -> str:
    valor = combo_cuenta.get().strip()

    if not valor or valor == "TODAS":
        return ""

    return valor.split(" | ", 1)[0].strip()


def consultar():
    global filas_actuales

    try:
        desde = validar_fecha(entry_desde.get(), "Desde")
        hasta = validar_fecha(entry_hasta.get(), "Hasta")

        if desde and hasta and desde > hasta:
            raise ValueError(
                "La fecha Desde no puede ser mayor que Hasta."
            )

        filas_actuales = consultar_libro_diario(
            desde=desde,
            hasta=hasta,
            tipo=combo_tipo.get().strip(),
            modulo=combo_modulo.get().strip(),
            cuenta_codigo=obtener_codigo_cuenta(),
            tercero=entry_tercero.get().strip(),
            buscar=entry_buscar.get().strip(),
        )

        cargar_tabla(filas_actuales)

    except Exception as error:
        messagebox.showerror(
            "Libro Diario",
            str(error)
        )


def cargar_tabla(filas: list[dict]) -> None:
    tabla.delete(*tabla.get_children())

    total_debito = 0.0
    total_credito = 0.0
    comprobantes = set()

    for indice, fila in enumerate(filas):
        total_debito += float(fila["debito"] or 0)
        total_credito += float(fila["credito"] or 0)
        comprobantes.add(int(fila["comprobante_id"]))

        tabla.insert(
            "",
            "end",
            iid=str(indice),
            values=(
                fila["fecha"],
                fila["consecutivo"],
                fila["tipo"],
                fila["documento"],
                fila["cuenta"],
                fila["nombre_cuenta"],
                fila["descripcion"],
                fila["tercero"],
                fila["centro_costo"],
                moneda(fila["debito"]),
                moneda(fila["credito"]),
                fila["modulo"],
                fila["estado"],
            )
        )

    diferencia = total_debito - total_credito

    lbl_movimientos.config(text=str(len(filas)))
    lbl_comprobantes.config(text=str(len(comprobantes)))
    lbl_debitos.config(text=moneda(total_debito))
    lbl_creditos.config(text=moneda(total_credito))
    lbl_diferencia.config(text=moneda(diferencia))

    if abs(diferencia) <= 0.01:
        lbl_estado.config(
            text="LIBRO CUADRADO",
            fg=COLOR_VERDE
        )
    else:
        lbl_estado.config(
            text="REVISAR DIFERENCIA",
            fg=COLOR_ROJO
        )


def limpiar():
    entry_desde.delete(0, "end")
    entry_hasta.delete(0, "end")
    combo_tipo.set("TODOS")
    combo_modulo.set("TODOS")
    combo_cuenta.set("TODAS")
    entry_tercero.delete(0, "end")
    entry_buscar.delete(0, "end")
    consultar()


def ver_comprobante(evento=None):
    seleccion = tabla.selection()

    if not seleccion:
        return

    indice = int(seleccion[0])
    fila = filas_actuales[indice]

    try:
        encabezado, detalle = consultar_comprobante(
            int(fila["comprobante_id"])
        )
    except Exception as error:
        messagebox.showerror(
            "Comprobante",
            str(error)
        )
        return

    top = tk.Toplevel(ventana)
    top.title(
        f"Comprobante {encabezado.get('consecutivo', '')}"
    )
    top.geometry("1250x650")
    top.configure(bg=COLOR_FONDO)

    tk.Label(
        top,
        text=(
            f"COMPROBANTE "
            f"{encabezado.get('consecutivo', '')}"
        ),
        bg=COLOR_OSCURO,
        fg="white",
        font=("Segoe UI", 18, "bold"),
        pady=14
    ).pack(fill="x")

    resumen = tk.Frame(top, bg=COLOR_TARJETA)
    resumen.pack(fill="x", padx=15, pady=15)

    datos = [
        ("Fecha", encabezado.get("fecha", "")),
        ("Tipo", encabezado.get("tipo", "")),
        ("Documento", encabezado.get("documento", "")),
        ("Tercero", encabezado.get("tercero", "")),
        ("Módulo", encabezado.get("modulo", "")),
        ("Estado", encabezado.get("estado", "")),
        ("Concepto", encabezado.get("concepto", "")),
        ("Origen", encabezado.get("origen", "")),
    ]

    for i, (titulo, valor) in enumerate(datos):
        marco = tk.Frame(resumen, bg=COLOR_TARJETA)
        marco.grid(
            row=i // 4,
            column=i % 4,
            sticky="nsew",
            padx=10,
            pady=7
        )
        tk.Label(
            marco,
            text=titulo,
            bg=COLOR_TARJETA,
            fg=COLOR_GRIS,
            font=("Segoe UI", 8, "bold")
        ).pack(anchor="w")
        tk.Label(
            marco,
            text=str(valor or ""),
            bg=COLOR_TARJETA,
            fg=COLOR_TEXTO,
            font=("Segoe UI", 10, "bold")
        ).pack(anchor="w")

    columnas = (
        "Secuencia",
        "Cuenta",
        "Nombre",
        "Descripción",
        "Tercero",
        "Centro costo",
        "Débito",
        "Crédito",
    )

    detalle_tabla = ttk.Treeview(
        top,
        columns=columnas,
        show="headings"
    )

    for columna in columnas:
        detalle_tabla.heading(columna, text=columna)

    detalle_tabla.pack(
        fill="both",
        expand=True,
        padx=15,
        pady=(0, 15)
    )

    for item in detalle:
        detalle_tabla.insert(
            "",
            "end",
            values=(
                item["secuencia"],
                item["cuenta"],
                item["nombre_cuenta"],
                item["descripcion"],
                item["tercero"],
                item["centro_costo"],
                moneda(item["debito"]),
                moneda(item["credito"]),
            )
        )


def exportar_excel():
    if not filas_actuales:
        messagebox.showwarning(
            "Exportar",
            "No hay información para exportar."
        )
        return

    carpeta_inicial = Path(
        r"C:\Users\jrive\visual\reportes"
    )
    carpeta_inicial.mkdir(parents=True, exist_ok=True)

    ruta = filedialog.asksaveasfilename(
        title="Guardar Libro Diario",
        initialdir=str(carpeta_inicial),
        initialfile=(
            "libro_diario_"
            + datetime.now().strftime("%Y%m%d_%H%M%S")
            + ".xlsx"
        ),
        defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx")],
    )

    if not ruta:
        return

    libro = Workbook()
    hoja = libro.active
    hoja.title = "Libro Diario"

    hoja.merge_cells("A1:M1")
    hoja["A1"] = "BME-ERP - LIBRO DIARIO"
    hoja["A1"].font = Font(
        bold=True,
        color="FFFFFF",
        size=16
    )
    hoja["A1"].fill = PatternFill(
        "solid",
        fgColor="153B5B"
    )
    hoja["A1"].alignment = Alignment(
        horizontal="center"
    )

    encabezados = [
        "Fecha",
        "Comprobante",
        "Tipo",
        "Documento",
        "Cuenta",
        "Nombre cuenta",
        "Descripción",
        "Tercero",
        "Centro costo",
        "Débito",
        "Crédito",
        "Módulo",
        "Estado",
    ]

    hoja.append([])
    hoja.append(encabezados)

    for celda in hoja[3]:
        celda.font = Font(
            bold=True,
            color="FFFFFF"
        )
        celda.fill = PatternFill(
            "solid",
            fgColor="0F5C8E"
        )
        celda.alignment = Alignment(
            horizontal="center"
        )

    total_debito = 0.0
    total_credito = 0.0

    for fila in filas_actuales:
        debito = float(fila["debito"] or 0)
        credito = float(fila["credito"] or 0)
        total_debito += debito
        total_credito += credito

        hoja.append([
            fila["fecha"],
            fila["consecutivo"],
            fila["tipo"],
            fila["documento"],
            fila["cuenta"],
            fila["nombre_cuenta"],
            fila["descripcion"],
            fila["tercero"],
            fila["centro_costo"],
            debito,
            credito,
            fila["modulo"],
            fila["estado"],
        ])

    fila_total = hoja.max_row + 1
    hoja.cell(fila_total, 9, "TOTALES")
    hoja.cell(fila_total, 10, total_debito)
    hoja.cell(fila_total, 11, total_credito)

    for celda in hoja[fila_total]:
        celda.font = Font(bold=True)

    for fila in range(4, hoja.max_row + 1):
        hoja.cell(fila, 10).number_format = '$#,##0.00'
        hoja.cell(fila, 11).number_format = '$#,##0.00'

    anchos = {
        1: 20,
        2: 20,
        3: 10,
        4: 16,
        5: 14,
        6: 30,
        7: 35,
        8: 24,
        9: 22,
        10: 16,
        11: 16,
        12: 15,
        13: 15,
    }

    for columna, ancho in anchos.items():
        hoja.column_dimensions[
            get_column_letter(columna)
        ].width = ancho

    hoja.freeze_panes = "A4"
    hoja.auto_filter.ref = f"A3:M{hoja.max_row}"

    libro.save(ruta)

    messagebox.showinfo(
        "Exportación",
        f"Libro Diario generado:\n\n{ruta}"
    )


validar_estructura()

ventana = tk.Tk()
ventana.title("BME-ERP - Libro Diario Profesional")
ventana.geometry("1550x900")
ventana.minsize(1200, 720)
ventana.configure(bg=COLOR_FONDO)

try:
    ventana.state("zoomed")
except tk.TclError:
    pass

estilo = ttk.Style()

try:
    estilo.theme_use("clam")
except tk.TclError:
    pass

estilo.configure(
    "Treeview",
    rowheight=27,
    font=("Segoe UI", 9)
)
estilo.configure(
    "Treeview.Heading",
    font=("Segoe UI", 9, "bold")
)

cabecera = tk.Frame(
    ventana,
    bg=COLOR_OSCURO,
    height=90
)
cabecera.pack(fill="x")
cabecera.pack_propagate(False)

tk.Label(
    cabecera,
    text="LIBRO DIARIO PROFESIONAL",
    bg=COLOR_OSCURO,
    fg="white",
    font=("Segoe UI", 21, "bold")
).pack(anchor="w", padx=28, pady=(16, 0))

tk.Label(
    cabecera,
    text=(
        "Movimientos contables, comprobantes, cuentas, "
        "terceros y centros de costo"
    ),
    bg=COLOR_OSCURO,
    fg="#BFDBFE",
    font=("Segoe UI", 9)
).pack(anchor="w", padx=29, pady=(3, 0))

filtros = tk.Frame(
    ventana,
    bg=COLOR_TARJETA,
    highlightbackground=COLOR_BORDE,
    highlightthickness=1
)
filtros.pack(
    fill="x",
    padx=16,
    pady=(14, 8)
)

campos = tk.Frame(filtros, bg=COLOR_TARJETA)
campos.pack(fill="x", padx=14, pady=12)

for i in range(9):
    campos.columnconfigure(i, weight=1)

def etiqueta(texto, columna):
    tk.Label(
        campos,
        text=texto,
        bg=COLOR_TARJETA,
        fg=COLOR_TEXTO,
        font=("Segoe UI", 8, "bold")
    ).grid(
        row=0,
        column=columna,
        sticky="w",
        padx=4
    )

etiqueta("Desde", 0)
etiqueta("Hasta", 1)
etiqueta("Tipo", 2)
etiqueta("Módulo", 3)
etiqueta("Cuenta", 4)
etiqueta("Tercero", 5)
etiqueta("Buscar", 6)

entry_desde = ttk.Entry(campos, width=14)
entry_desde.grid(row=1, column=0, sticky="ew", padx=4)

entry_hasta = ttk.Entry(campos, width=14)
entry_hasta.grid(row=1, column=1, sticky="ew", padx=4)

combo_tipo = ttk.Combobox(
    campos,
    state="readonly",
    values=["TODOS"] + listar_tipos_comprobante()
)
combo_tipo.grid(row=1, column=2, sticky="ew", padx=4)
combo_tipo.set("TODOS")

combo_modulo = ttk.Combobox(
    campos,
    state="readonly",
    values=["TODOS"] + listar_modulos()
)
combo_modulo.grid(row=1, column=3, sticky="ew", padx=4)
combo_modulo.set("TODOS")

cuentas = listar_cuentas()
combo_cuenta = ttk.Combobox(
    campos,
    state="readonly",
    values=["TODAS"] + [
        f"{c['codigo']} | {c['nombre']}"
        for c in cuentas
    ]
)
combo_cuenta.grid(row=1, column=4, sticky="ew", padx=4)
combo_cuenta.set("TODAS")

entry_tercero = ttk.Entry(campos)
entry_tercero.grid(row=1, column=5, sticky="ew", padx=4)

entry_buscar = ttk.Entry(campos)
entry_buscar.grid(row=1, column=6, sticky="ew", padx=4)

tk.Button(
    campos,
    text="Consultar",
    command=consultar,
    bg=COLOR_AZUL,
    fg="white",
    relief="flat",
    font=("Segoe UI", 9, "bold"),
    padx=16,
    pady=7
).grid(row=1, column=7, padx=4)

tk.Button(
    campos,
    text="Limpiar",
    command=limpiar,
    bg=COLOR_GRIS,
    fg="white",
    relief="flat",
    font=("Segoe UI", 9, "bold"),
    padx=16,
    pady=7
).grid(row=1, column=8, padx=4)

tk.Button(
    campos,
    text="Exportar Excel",
    command=exportar_excel,
    bg=COLOR_VERDE,
    fg="white",
    relief="flat",
    font=("Segoe UI", 9, "bold"),
    padx=16,
    pady=7
).grid(row=1, column=9, padx=4)

resumen = tk.Frame(ventana, bg=COLOR_FONDO)
resumen.pack(fill="x", padx=16, pady=(0, 8))

for i in range(6):
    resumen.columnconfigure(i, weight=1)

def tarjeta(columna, titulo):
    marco = tk.Frame(
        resumen,
        bg=COLOR_TARJETA,
        highlightbackground=COLOR_BORDE,
        highlightthickness=1
    )
    marco.grid(
        row=0,
        column=columna,
        sticky="ew",
        padx=4
    )

    tk.Label(
        marco,
        text=titulo,
        bg=COLOR_TARJETA,
        fg=COLOR_GRIS,
        font=("Segoe UI", 8, "bold")
    ).pack(anchor="w", padx=12, pady=(9, 2))

    valor = tk.Label(
        marco,
        text="0",
        bg=COLOR_TARJETA,
        fg=COLOR_TEXTO,
        font=("Segoe UI", 14, "bold")
    )
    valor.pack(anchor="w", padx=12, pady=(0, 9))
    return valor

lbl_movimientos = tarjeta(0, "MOVIMIENTOS")
lbl_comprobantes = tarjeta(1, "COMPROBANTES")
lbl_debitos = tarjeta(2, "TOTAL DÉBITOS")
lbl_creditos = tarjeta(3, "TOTAL CRÉDITOS")
lbl_diferencia = tarjeta(4, "DIFERENCIA")
lbl_estado = tarjeta(5, "ESTADO")

panel_tabla = tk.Frame(
    ventana,
    bg=COLOR_TARJETA,
    highlightbackground=COLOR_BORDE,
    highlightthickness=1
)
panel_tabla.pack(
    fill="both",
    expand=True,
    padx=16,
    pady=(0, 12)
)

tk.Label(
    panel_tabla,
    text=(
        "Detalle del Libro Diario "
        "(doble clic para abrir el comprobante)"
    ),
    bg=COLOR_TARJETA,
    fg=COLOR_TEXTO,
    font=("Segoe UI", 10, "bold")
).pack(anchor="w", padx=12, pady=(10, 5))

columnas = (
    "Fecha",
    "Comprobante",
    "Tipo",
    "Documento",
    "Cuenta",
    "Nombre cuenta",
    "Descripción",
    "Tercero",
    "Centro costo",
    "Débito",
    "Crédito",
    "Módulo",
    "Estado",
)

tabla = ttk.Treeview(
    panel_tabla,
    columns=columnas,
    show="headings"
)

anchos = {
    "Fecha": 150,
    "Comprobante": 145,
    "Tipo": 70,
    "Documento": 110,
    "Cuenta": 90,
    "Nombre cuenta": 210,
    "Descripción": 245,
    "Tercero": 170,
    "Centro costo": 160,
    "Débito": 110,
    "Crédito": 110,
    "Módulo": 100,
    "Estado": 100,
}

for columna in columnas:
    tabla.heading(columna, text=columna)
    tabla.column(
        columna,
        width=anchos[columna],
        anchor=(
            "e"
            if columna in ("Débito", "Crédito")
            else "w"
        )
    )

scroll_y = ttk.Scrollbar(
    panel_tabla,
    orient="vertical",
    command=tabla.yview
)
scroll_x = ttk.Scrollbar(
    panel_tabla,
    orient="horizontal",
    command=tabla.xview
)

tabla.configure(
    yscrollcommand=scroll_y.set,
    xscrollcommand=scroll_x.set
)

tabla.pack(
    fill="both",
    expand=True,
    side="left",
    padx=(12, 0),
    pady=(0, 12)
)
scroll_y.pack(
    fill="y",
    side="right",
    padx=(0, 12),
    pady=(0, 12)
)
scroll_x.pack(
    fill="x",
    side="bottom",
    padx=12,
    pady=(0, 12)
)

tabla.bind("<Double-1>", ver_comprobante)

tk.Label(
    ventana,
    text=f"Base de datos: {RUTA_DB}",
    bg=COLOR_FONDO,
    fg=COLOR_GRIS,
    font=("Segoe UI", 8)
).pack(pady=(0, 8))

consultar()
ventana.mainloop()
