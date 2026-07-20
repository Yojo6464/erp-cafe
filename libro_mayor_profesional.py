
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from consultas_contables import (
    RUTA_DB,
    conectar,
    consultar_libro_diario,
    consultar_comprobante,
    listar_cuentas,
)

C_FONDO = "#EEF3F8"
C_BLANCO = "#FFFFFF"
C_AZUL = "#0F5C8E"
C_OSCURO = "#153B5B"
C_VERDE = "#15803D"
C_GRIS = "#64748B"
C_TEXTO = "#1F2937"
C_BORDE = "#D7E0E8"

resumen_actual = []
detalle_actual = []


def moneda(valor):
    return f"${float(valor or 0):,.2f}"


def validar_fecha(texto, nombre):
    texto = texto.strip()
    if not texto:
        return ""
    try:
        datetime.strptime(texto, "%Y-%m-%d")
    except ValueError as error:
        raise ValueError(
            f"{nombre} debe tener formato AAAA-MM-DD."
        ) from error
    return texto


def fecha_anterior(fecha):
    if not fecha:
        return ""
    return (
        datetime.strptime(fecha, "%Y-%m-%d")
        - timedelta(days=1)
    ).strftime("%Y-%m-%d")


def codigo_cuenta():
    valor = combo_cuenta.get().strip()
    if not valor or valor == "TODAS":
        return ""
    return valor.split(" | ", 1)[0].strip()


def naturalezas():
    conexion = conectar()
    try:
        cursor = conexion.cursor()
        cursor.execute("""
            SELECT
                codigo,
                UPPER(COALESCE(naturaleza, 'DEBITO')) AS naturaleza
            FROM plan_cuentas
            WHERE UPPER(COALESCE(estado, 'ACTIVA'))='ACTIVA'
        """)
        return {
            str(f["codigo"]): str(f["naturaleza"])
            for f in cursor.fetchall()
        }
    finally:
        conexion.close()


def saldo_por_naturaleza(naturaleza, debitos, creditos):
    if naturaleza == "CREDITO":
        return creditos - debitos
    return debitos - creditos


def construir_resumen(desde, hasta, cuenta, tercero, buscar):
    mapa_naturaleza = naturalezas()

    periodo = consultar_libro_diario(
        desde=desde,
        hasta=hasta,
        cuenta_codigo=cuenta,
        tercero=tercero,
        buscar=buscar,
    )

    anteriores = []
    if desde:
        anteriores = consultar_libro_diario(
            hasta=fecha_anterior(desde),
            cuenta_codigo=cuenta,
            tercero=tercero,
            buscar=buscar,
        )

    cuentas = {}

    def asegurar(fila):
        codigo = str(fila["cuenta"])
        if codigo not in cuentas:
            cuentas[codigo] = {
                "cuenta": codigo,
                "nombre": fila["nombre_cuenta"],
                "naturaleza": mapa_naturaleza.get(
                    codigo, "DEBITO"
                ),
                "inicial_debitos": 0.0,
                "inicial_creditos": 0.0,
                "debitos": 0.0,
                "creditos": 0.0,
                "movimientos": 0,
            }
        return cuentas[codigo]

    for fila in anteriores:
        item = asegurar(fila)
        item["inicial_debitos"] += float(fila["debito"] or 0)
        item["inicial_creditos"] += float(fila["credito"] or 0)

    for fila in periodo:
        item = asegurar(fila)
        item["debitos"] += float(fila["debito"] or 0)
        item["creditos"] += float(fila["credito"] or 0)
        item["movimientos"] += 1

    resultado = []

    for codigo in sorted(cuentas):
        item = cuentas[codigo]
        inicial = saldo_por_naturaleza(
            item["naturaleza"],
            item["inicial_debitos"],
            item["inicial_creditos"],
        )
        final = inicial + saldo_por_naturaleza(
            item["naturaleza"],
            item["debitos"],
            item["creditos"],
        )
        item["saldo_inicial"] = inicial
        item["saldo_final"] = final
        resultado.append(item)

    return resultado


def consultar():
    global resumen_actual

    try:
        desde = validar_fecha(entry_desde.get(), "Desde")
        hasta = validar_fecha(entry_hasta.get(), "Hasta")

        if desde and hasta and desde > hasta:
            raise ValueError(
                "La fecha Desde no puede ser mayor que Hasta."
            )

        resumen_actual = construir_resumen(
            desde,
            hasta,
            codigo_cuenta(),
            entry_tercero.get().strip(),
            entry_buscar.get().strip(),
        )

        tabla_resumen.delete(*tabla_resumen.get_children())

        total_inicial = 0.0
        total_debitos = 0.0
        total_creditos = 0.0
        total_final = 0.0

        for fila in resumen_actual:
            total_inicial += fila["saldo_inicial"]
            total_debitos += fila["debitos"]
            total_creditos += fila["creditos"]
            total_final += fila["saldo_final"]

            tabla_resumen.insert(
                "",
                "end",
                iid=fila["cuenta"],
                values=(
                    fila["cuenta"],
                    fila["nombre"],
                    fila["naturaleza"],
                    moneda(fila["saldo_inicial"]),
                    moneda(fila["debitos"]),
                    moneda(fila["creditos"]),
                    moneda(fila["saldo_final"]),
                    fila["movimientos"],
                )
            )

        lbl_cuentas.config(text=str(len(resumen_actual)))
        lbl_inicial.config(text=moneda(total_inicial))
        lbl_debitos.config(text=moneda(total_debitos))
        lbl_creditos.config(text=moneda(total_creditos))
        lbl_final.config(text=moneda(total_final))

        if resumen_actual:
            codigo = resumen_actual[0]["cuenta"]
            tabla_resumen.selection_set(codigo)
            tabla_resumen.focus(codigo)
            cargar_detalle(codigo)
        else:
            cargar_detalle("")

    except Exception as error:
        messagebox.showerror("Libro Mayor", str(error))


def seleccionar_cuenta(evento=None):
    seleccion = tabla_resumen.selection()
    if seleccion:
        cargar_detalle(seleccion[0])


def cargar_detalle(codigo):
    global detalle_actual

    tabla_detalle.delete(*tabla_detalle.get_children())
    detalle_actual = []

    if not codigo:
        lbl_cuenta.config(text="Seleccione una cuenta")
        lbl_det_debitos.config(text="$0.00")
        lbl_det_creditos.config(text="$0.00")
        return

    try:
        desde = validar_fecha(entry_desde.get(), "Desde")
        hasta = validar_fecha(entry_hasta.get(), "Hasta")

        detalle_actual = consultar_libro_diario(
            desde=desde,
            hasta=hasta,
            cuenta_codigo=codigo,
            tercero=entry_tercero.get().strip(),
            buscar=entry_buscar.get().strip(),
        )

        resumen = next(
            (x for x in resumen_actual if x["cuenta"] == codigo),
            None
        )

        nombre = resumen["nombre"] if resumen else ""
        naturaleza = resumen["naturaleza"] if resumen else "DEBITO"
        saldo = resumen["saldo_inicial"] if resumen else 0.0

        lbl_cuenta.config(text=f"{codigo} - {nombre}")

        total_debitos = 0.0
        total_creditos = 0.0

        for i, fila in enumerate(detalle_actual):
            debito = float(fila["debito"] or 0)
            credito = float(fila["credito"] or 0)
            total_debitos += debito
            total_creditos += credito

            if naturaleza == "CREDITO":
                saldo += credito - debito
            else:
                saldo += debito - credito

            tabla_detalle.insert(
                "",
                "end",
                iid=str(i),
                values=(
                    fila["fecha"],
                    fila["consecutivo"],
                    fila["tipo"],
                    fila["documento"],
                    fila["concepto"],
                    fila["descripcion"],
                    fila["tercero"],
                    fila["centro_costo"],
                    moneda(debito),
                    moneda(credito),
                    moneda(saldo),
                    fila["modulo"],
                    fila["usuario"],
                )
            )

        lbl_det_debitos.config(text=moneda(total_debitos))
        lbl_det_creditos.config(text=moneda(total_creditos))

    except Exception as error:
        messagebox.showerror("Libro Mayor", str(error))


def ver_comprobante(evento=None):
    seleccion = tabla_detalle.selection()
    if not seleccion:
        return

    fila = detalle_actual[int(seleccion[0])]

    try:
        encabezado, detalle = consultar_comprobante(
            int(fila["comprobante_id"])
        )
    except Exception as error:
        messagebox.showerror("Comprobante", str(error))
        return

    top = tk.Toplevel(ventana)
    top.title(f"Comprobante {encabezado['consecutivo']}")
    top.geometry("1200x620")

    tk.Label(
        top,
        text=f"COMPROBANTE {encabezado['consecutivo']}",
        bg=C_OSCURO,
        fg="white",
        font=("Segoe UI", 17, "bold"),
        pady=14
    ).pack(fill="x")

    info = (
        f"Fecha: {encabezado.get('fecha', '')}   |   "
        f"Tipo: {encabezado.get('tipo', '')}   |   "
        f"Documento: {encabezado.get('documento', '')}   |   "
        f"Tercero: {encabezado.get('tercero', '')}\n"
        f"Concepto: {encabezado.get('concepto', '')}   |   "
        f"Módulo: {encabezado.get('modulo', '')}   |   "
        f"Estado: {encabezado.get('estado', '')}"
    )

    tk.Label(
        top,
        text=info,
        justify="left",
        anchor="w",
        padx=15,
        pady=12
    ).pack(fill="x")

    columnas = (
        "Sec.",
        "Cuenta",
        "Nombre",
        "Descripción",
        "Tercero",
        "Centro costo",
        "Débito",
        "Crédito",
    )

    tabla = ttk.Treeview(
        top,
        columns=columnas,
        show="headings"
    )

    for columna in columnas:
        tabla.heading(columna, text=columna)

    tabla.pack(fill="both", expand=True, padx=15, pady=15)

    for item in detalle:
        tabla.insert(
            "",
            "end",
            values=(
                item["secuencia"],
                item["cuenta"],
                item["nombre_cuenta"],
                item["descripcion"],
                item["tercero"],
                item["centro_costo"],
                moneda(item["debito"]),
                moneda(item["credito"]),
            )
        )


def limpiar():
    entry_desde.delete(0, "end")
    entry_hasta.delete(0, "end")
    combo_cuenta.set("TODAS")
    entry_tercero.delete(0, "end")
    entry_buscar.delete(0, "end")
    consultar()


def exportar_excel():
    if not resumen_actual:
        messagebox.showwarning(
            "Exportar",
            "No hay información para exportar."
        )
        return

    carpeta = Path(r"C:\Users\jrive\visual\reportes")
    carpeta.mkdir(parents=True, exist_ok=True)

    ruta = filedialog.asksaveasfilename(
        title="Guardar Libro Mayor",
        initialdir=str(carpeta),
        initialfile=(
            "libro_mayor_"
            + datetime.now().strftime("%Y%m%d_%H%M%S")
            + ".xlsx"
        ),
        defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx")],
    )

    if not ruta:
        return

    libro = Workbook()
    hoja = libro.active
    hoja.title = "Resumen Mayor"

    hoja.merge_cells("A1:H1")
    hoja["A1"] = "BME-ERP - LIBRO MAYOR PROFESIONAL"
    hoja["A1"].font = Font(
        bold=True,
        color="FFFFFF",
        size=16
    )
    hoja["A1"].fill = PatternFill(
        "solid",
        fgColor="153B5B"
    )

    hoja.append([])
    hoja.append([
        "Cuenta",
        "Nombre",
        "Naturaleza",
        "Saldo inicial",
        "Débitos",
        "Créditos",
        "Saldo final",
        "Movimientos",
    ])

    for celda in hoja[3]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="0F5C8E")

    for fila in resumen_actual:
        hoja.append([
            fila["cuenta"],
            fila["nombre"],
            fila["naturaleza"],
            fila["saldo_inicial"],
            fila["debitos"],
            fila["creditos"],
            fila["saldo_final"],
            fila["movimientos"],
        ])

    for fila in range(4, hoja.max_row + 1):
        for columna in range(4, 8):
            hoja.cell(fila, columna).number_format = '$#,##0.00'

    for columna, ancho in {
        1: 14,
        2: 34,
        3: 14,
        4: 18,
        5: 18,
        6: 18,
        7: 18,
        8: 14,
    }.items():
        hoja.column_dimensions[
            get_column_letter(columna)
        ].width = ancho

    hoja.freeze_panes = "A4"
    libro.save(ruta)

    messagebox.showinfo(
        "Exportación",
        f"Libro Mayor generado:\n\n{ruta}"
    )


ventana = tk.Tk()
ventana.title("BME-ERP - Libro Mayor Profesional")
ventana.geometry("1550x900")
ventana.minsize(1200, 720)
ventana.configure(bg=C_FONDO)

try:
    ventana.state("zoomed")
except tk.TclError:
    pass

estilo = ttk.Style()
try:
    estilo.theme_use("clam")
except tk.TclError:
    pass

estilo.configure("Treeview", rowheight=27)
estilo.configure(
    "Treeview.Heading",
    font=("Segoe UI", 9, "bold")
)

cabecera = tk.Frame(ventana, bg=C_OSCURO, height=88)
cabecera.pack(fill="x")
cabecera.pack_propagate(False)

tk.Label(
    cabecera,
    text="LIBRO MAYOR PROFESIONAL",
    bg=C_OSCURO,
    fg="white",
    font=("Segoe UI", 21, "bold")
).pack(anchor="w", padx=26, pady=(14, 0))

tk.Label(
    cabecera,
    text=(
        "Saldos iniciales, débitos, créditos, "
        "saldos finales y detalle por cuenta"
    ),
    bg=C_OSCURO,
    fg="#BFDBFE"
).pack(anchor="w", padx=27, pady=(3, 0))

filtros = tk.Frame(
    ventana,
    bg=C_BLANCO,
    highlightbackground=C_BORDE,
    highlightthickness=1
)
filtros.pack(fill="x", padx=15, pady=(12, 7))

campos = tk.Frame(filtros, bg=C_BLANCO)
campos.pack(fill="x", padx=12, pady=10)

for i, texto in enumerate(
    ["Desde", "Hasta", "Cuenta", "Tercero", "Buscar"]
):
    tk.Label(
        campos,
        text=texto,
        bg=C_BLANCO,
        fg=C_TEXTO,
        font=("Segoe UI", 8, "bold")
    ).grid(row=0, column=i, sticky="w", padx=4)

entry_desde = ttk.Entry(campos, width=15)
entry_desde.grid(row=1, column=0, padx=4)

entry_hasta = ttk.Entry(campos, width=15)
entry_hasta.grid(row=1, column=1, padx=4)

cuentas = listar_cuentas()
combo_cuenta = ttk.Combobox(
    campos,
    state="readonly",
    width=38,
    values=["TODAS"] + [
        f"{c['codigo']} | {c['nombre']}"
        for c in cuentas
    ]
)
combo_cuenta.grid(row=1, column=2, padx=4)
combo_cuenta.set("TODAS")

entry_tercero = ttk.Entry(campos, width=25)
entry_tercero.grid(row=1, column=3, padx=4)

entry_buscar = ttk.Entry(campos, width=25)
entry_buscar.grid(row=1, column=4, padx=4)

tk.Button(
    campos,
    text="Consultar",
    command=consultar,
    bg=C_AZUL,
    fg="white",
    relief="flat",
    padx=16,
    pady=6
).grid(row=1, column=5, padx=5)

tk.Button(
    campos,
    text="Limpiar",
    command=limpiar,
    bg=C_GRIS,
    fg="white",
    relief="flat",
    padx=16,
    pady=6
).grid(row=1, column=6, padx=5)

tk.Button(
    campos,
    text="Exportar Excel",
    command=exportar_excel,
    bg=C_VERDE,
    fg="white",
    relief="flat",
    padx=16,
    pady=6
).grid(row=1, column=7, padx=5)

kpis = tk.Frame(ventana, bg=C_FONDO)
kpis.pack(fill="x", padx=15, pady=(0, 7))

for i in range(5):
    kpis.columnconfigure(i, weight=1)


def tarjeta(columna, titulo):
    marco = tk.Frame(
        kpis,
        bg=C_BLANCO,
        highlightbackground=C_BORDE,
        highlightthickness=1
    )
    marco.grid(row=0, column=columna, sticky="ew", padx=4)

    tk.Label(
        marco,
        text=titulo,
        bg=C_BLANCO,
        fg=C_GRIS,
        font=("Segoe UI", 8, "bold")
    ).pack(anchor="w", padx=11, pady=(8, 2))

    valor = tk.Label(
        marco,
        text="0",
        bg=C_BLANCO,
        fg=C_TEXTO,
        font=("Segoe UI", 14, "bold")
    )
    valor.pack(anchor="w", padx=11, pady=(0, 8))
    return valor


lbl_cuentas = tarjeta(0, "CUENTAS")
lbl_inicial = tarjeta(1, "SALDO INICIAL")
lbl_debitos = tarjeta(2, "DÉBITOS")
lbl_creditos = tarjeta(3, "CRÉDITOS")
lbl_final = tarjeta(4, "SALDO FINAL")

panel_resumen = tk.LabelFrame(
    ventana,
    text="Resumen por cuenta",
    bg=C_BLANCO,
    fg=C_TEXTO,
    font=("Segoe UI", 10, "bold")
)
panel_resumen.pack(fill="x", padx=15, pady=(0, 7))

columnas_resumen = (
    "Cuenta",
    "Nombre",
    "Naturaleza",
    "Saldo inicial",
    "Débitos",
    "Créditos",
    "Saldo final",
    "Movimientos",
)

tabla_resumen = ttk.Treeview(
    panel_resumen,
    columns=columnas_resumen,
    show="headings",
    height=9
)

for columna in columnas_resumen:
    tabla_resumen.heading(columna, text=columna)

tabla_resumen.pack(
    fill="both",
    expand=True,
    padx=10,
    pady=10
)
tabla_resumen.bind("<<TreeviewSelect>>", seleccionar_cuenta)

panel_detalle = tk.LabelFrame(
    ventana,
    text="Detalle cronológico de la cuenta",
    bg=C_BLANCO,
    fg=C_TEXTO,
    font=("Segoe UI", 10, "bold")
)
panel_detalle.pack(
    fill="both",
    expand=True,
    padx=15,
    pady=(0, 10)
)

barra = tk.Frame(panel_detalle, bg=C_BLANCO)
barra.pack(fill="x", padx=10, pady=(8, 4))

lbl_cuenta = tk.Label(
    barra,
    text="Seleccione una cuenta",
    bg=C_BLANCO,
    fg=C_TEXTO,
    font=("Segoe UI", 10, "bold")
)
lbl_cuenta.pack(side="left")

tk.Label(
    barra,
    text="Débitos:",
    bg=C_BLANCO,
    fg=C_GRIS
).pack(side="left", padx=(25, 5))

lbl_det_debitos = tk.Label(
    barra,
    text="$0.00",
    bg=C_BLANCO,
    fg=C_TEXTO,
    font=("Segoe UI", 9, "bold")
)
lbl_det_debitos.pack(side="left")

tk.Label(
    barra,
    text="Créditos:",
    bg=C_BLANCO,
    fg=C_GRIS
).pack(side="left", padx=(20, 5))

lbl_det_creditos = tk.Label(
    barra,
    text="$0.00",
    bg=C_BLANCO,
    fg=C_TEXTO,
    font=("Segoe UI", 9, "bold")
)
lbl_det_creditos.pack(side="left")

columnas_detalle = (
    "Fecha",
    "Comprobante",
    "Tipo",
    "Documento",
    "Concepto",
    "Descripción",
    "Tercero",
    "Centro costo",
    "Débito",
    "Crédito",
    "Saldo",
    "Módulo",
    "Usuario",
)

tabla_detalle = ttk.Treeview(
    panel_detalle,
    columns=columnas_detalle,
    show="headings"
)

for columna in columnas_detalle:
    tabla_detalle.heading(columna, text=columna)

tabla_detalle.pack(
    fill="both",
    expand=True,
    padx=10,
    pady=(0, 10)
)
tabla_detalle.bind("<Double-1>", ver_comprobante)

tk.Label(
    ventana,
    text=f"Base de datos: {RUTA_DB}",
    bg=C_FONDO,
    fg=C_GRIS,
    font=("Segoe UI", 8)
).pack(pady=(0, 6))

consultar()
ventana.mainloop()
