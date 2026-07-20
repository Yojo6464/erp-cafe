import tkinter as tk
from tkinter import ttk, messagebox
import sqlite3
from datetime import datetime, timedelta
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ============================================================
# ERP CAFÉ ALTO DE LA CRUZ
# MÓDULO DE COMPRAS INTEGRADO v2.0
# ============================================================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
RUTA_DB = os.path.join(BASE_DIR, "erp_cafe.db")

COLOR_FONDO = "#EEF3F8"
COLOR_AZUL = "#0F5C8E"
COLOR_AZUL_OSCURO = "#153B5B"
COLOR_VERDE = "#15803D"
COLOR_NARANJA = "#C56A00"
COLOR_ROJO = "#B42318"
COLOR_TEXTO = "#1F2937"
COLOR_SUAVE = "#64748B"
COLOR_BORDE = "#D7E0E8"
COLOR_BLANCO = "#FFFFFF"

detalle_actual = []
compra_seleccionada_id = None


# ============================================================
# BASE DE DATOS
# ============================================================

def conectar():
    conexion = sqlite3.connect(RUTA_DB, timeout=20)
    conexion.execute("PRAGMA foreign_keys = ON")
    return conexion


def columnas_tabla(conexion, tabla):
    return {
        fila[1]
        for fila in conexion.execute(f"PRAGMA table_info({tabla})").fetchall()
    }


def agregar_columna_si_falta(conexion, tabla, columna, definicion):
    columnas = columnas_tabla(conexion, tabla)
    if columna not in columnas:
        conexion.execute(
            f"ALTER TABLE {tabla} ADD COLUMN {columna} {definicion}"
        )


def inicializar_bd():
    with conectar() as conexion:
        conexion.execute("""
            CREATE TABLE IF NOT EXISTS compras (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                fecha TEXT,
                proveedor TEXT,
                tipo_compra TEXT,
                descripcion TEXT,
                valor REAL,
                forma_pago TEXT,
                dias_credito INTEGER
            )
        """)

        nuevas_columnas = [
            ("factura", "TEXT DEFAULT ''"),
            ("subtotal", "REAL DEFAULT 0"),
            ("iva", "REAL DEFAULT 0"),
            ("total", "REAL DEFAULT 0"),
            ("estado", "TEXT DEFAULT 'RECIBIDA'"),
            ("observaciones", "TEXT DEFAULT ''"),
            ("vencimiento", "TEXT DEFAULT ''"),
            ("fecha_anulacion", "TEXT DEFAULT ''")
        ]

        for columna, definicion in nuevas_columnas:
            agregar_columna_si_falta(
                conexion, "compras", columna, definicion
            )

        conexion.execute("""
            CREATE TABLE IF NOT EXISTS detalle_compras (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                compra_id INTEGER NOT NULL,
                producto TEXT NOT NULL,
                presentacion TEXT NOT NULL,
                cantidad REAL NOT NULL,
                costo_unitario REAL NOT NULL,
                subtotal REAL NOT NULL,
                lote TEXT DEFAULT '',
                FOREIGN KEY(compra_id) REFERENCES compras(id)
            )
        """)

        conexion.execute("""
            CREATE TABLE IF NOT EXISTS inventario (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                producto TEXT,
                presentacion TEXT,
                cantidad REAL DEFAULT 0,
                lote TEXT DEFAULT '',
                costo_unitario REAL DEFAULT 0,
                fecha_ingreso TEXT DEFAULT '',
                numero_despacho TEXT DEFAULT '',
                stock_minimo INTEGER DEFAULT 0,
                costo REAL DEFAULT 0,
                fecha TEXT DEFAULT '',
                despacho TEXT DEFAULT ''
            )
        """)

        conexion.execute("""
            CREATE TABLE IF NOT EXISTS kardex (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                fecha TEXT,
                producto TEXT,
                presentacion TEXT,
                movimiento TEXT,
                entrada REAL DEFAULT 0,
                salida REAL DEFAULT 0,
                saldo REAL DEFAULT 0,
                costo_unitario REAL DEFAULT 0,
                lote TEXT DEFAULT '',
                origen TEXT DEFAULT '',
                observaciones TEXT DEFAULT ''
            )
        """)

        conexion.execute("""
            CREATE TABLE IF NOT EXISTS proveedores (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT NOT NULL,
                telefono TEXT,
                ciudad TEXT,
                correo TEXT
            )
        """)

        conexion.execute("""
            CREATE TABLE IF NOT EXISTS cuentas_pagar (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                proveedor TEXT,
                valor REAL,
                saldo REAL,
                fecha TEXT,
                vencimiento TEXT,
                estado TEXT
            )
        """)

        conexion.execute("""
            CREATE TABLE IF NOT EXISTS auditoria_erp (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                fecha_hora TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                usuario TEXT,
                rol TEXT,
                accion TEXT NOT NULL,
                detalle TEXT,
                modulo TEXT
            )
        """)

        conexion.execute("""
            CREATE INDEX IF NOT EXISTS idx_detalle_compra
            ON detalle_compras(compra_id)
        """)

        conexion.execute("""
            CREATE INDEX IF NOT EXISTS idx_compras_factura
            ON compras(factura)
        """)

        conexion.execute("""
            CREATE INDEX IF NOT EXISTS idx_inventario_producto_lote
            ON inventario(producto, presentacion, lote)
        """)

        conexion.commit()


# ============================================================
# UTILIDADES
# ============================================================

def fecha_actual():
    return datetime.now().strftime("%Y-%m-%d")


def fecha_hora_actual():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def a_numero(valor, nombre, permitir_cero=False):
    texto = str(valor).strip().replace(",", "")
    try:
        numero = float(texto)
    except ValueError:
        raise ValueError(f"{nombre} debe ser un número.")

    if permitir_cero:
        if numero < 0:
            raise ValueError(f"{nombre} no puede ser negativo.")
    elif numero <= 0:
        raise ValueError(f"{nombre} debe ser mayor que cero.")

    return numero


def moneda(valor):
    return f"${float(valor or 0):,.0f}"


def registrar_auditoria(
    conexion, accion, detalle="", modulo="Compras"
):
    conexion.execute("""
        INSERT INTO auditoria_erp(
            usuario, rol, accion, detalle, modulo
        )
        VALUES (?, ?, ?, ?, ?)
    """, (
        os.environ.get("ERP_USUARIO", "usuario_local"),
        os.environ.get("ERP_ROL", "OPERADOR"),
        accion,
        detalle,
        modulo
    ))


def numero_compra_siguiente():
    with conectar() as conexion:
        siguiente = conexion.execute("""
            SELECT IFNULL(MAX(id), 0) + 1 FROM compras
        """).fetchone()[0]

    return f"COMP-{int(siguiente):06d}"


def factura_duplicada(factura, proveedor):
    if not factura.strip():
        return False

    with conectar() as conexion:
        cantidad = conexion.execute("""
            SELECT COUNT(*)
            FROM compras
            WHERE UPPER(TRIM(factura)) = UPPER(TRIM(?))
              AND UPPER(TRIM(proveedor)) = UPPER(TRIM(?))
              AND UPPER(COALESCE(estado, 'RECIBIDA')) <> 'ANULADA'
        """, (factura, proveedor)).fetchone()[0]

    return cantidad > 0


# ============================================================
# CARGAS DE DATOS
# ============================================================

def cargar_proveedores():
    with conectar() as conexion:
        filas = conexion.execute("""
            SELECT nombre
            FROM proveedores
            WHERE TRIM(COALESCE(nombre, '')) <> ''
            ORDER BY nombre
        """).fetchall()

    combo_proveedor["values"] = [fila[0] for fila in filas]


def cargar_productos():
    with conectar() as conexion:
        filas = conexion.execute("""
            SELECT DISTINCT producto
            FROM inventario
            WHERE TRIM(COALESCE(producto, '')) <> ''
            ORDER BY producto
        """).fetchall()

    productos = [fila[0] for fila in filas]
    combo_producto["values"] = productos


def cargar_presentaciones(evento=None):
    producto = combo_producto.get().strip()

    if not producto:
        combo_presentacion["values"] = []
        return

    with conectar() as conexion:
        filas = conexion.execute("""
            SELECT DISTINCT presentacion
            FROM inventario
            WHERE producto = ?
              AND TRIM(COALESCE(presentacion, '')) <> ''
            ORDER BY presentacion
        """, (producto,)).fetchall()

    combo_presentacion["values"] = [fila[0] for fila in filas]



# ============================================================
# CATÁLOGO INTELIGENTE
# ============================================================

def filtrar_combobox(combo, valores_originales, texto):
    texto = texto.strip().lower()
    if not texto:
        combo["values"] = valores_originales
        return

    filtrados = [
        valor for valor in valores_originales
        if texto in str(valor).lower()
    ]
    combo["values"] = filtrados


def autocompletar_proveedor(evento=None):
    with conectar() as conexion:
        valores = [
            fila[0] for fila in conexion.execute("""
                SELECT nombre
                FROM proveedores
                WHERE TRIM(COALESCE(nombre, '')) <> ''
                ORDER BY nombre
            """).fetchall()
        ]
    filtrar_combobox(combo_proveedor, valores, combo_proveedor.get())


def autocompletar_producto(evento=None):
    with conectar() as conexion:
        valores = [
            fila[0] for fila in conexion.execute("""
                SELECT DISTINCT producto
                FROM inventario
                WHERE TRIM(COALESCE(producto, '')) <> ''
                ORDER BY producto
            """).fetchall()
        ]
    filtrar_combobox(combo_producto, valores, combo_producto.get())


def actualizar_informacion_producto(evento=None):
    producto = combo_producto.get().strip()
    presentacion = combo_presentacion.get().strip()

    if not producto:
        lbl_ultimo_costo.config(text="$0")
        lbl_costo_promedio.config(text="$0")
        lbl_ultimo_proveedor.config(text="—")
        return

    cargar_presentaciones()

    with conectar() as conexion:
        if presentacion:
            fila = conexion.execute("""
                SELECT
                    COALESCE(costo_unitario, costo, 0),
                    COALESCE(
                        SUM(cantidad * COALESCE(costo_unitario, costo, 0))
                        / NULLIF(SUM(cantidad), 0),
                        0
                    )
                FROM inventario
                WHERE producto = ?
                  AND presentacion = ?
            """, (producto, presentacion)).fetchone()
        else:
            fila = conexion.execute("""
                SELECT
                    COALESCE(MAX(COALESCE(costo_unitario, costo, 0)), 0),
                    COALESCE(
                        SUM(cantidad * COALESCE(costo_unitario, costo, 0))
                        / NULLIF(SUM(cantidad), 0),
                        0
                    )
                FROM inventario
                WHERE producto = ?
            """, (producto,)).fetchone()

        ultimo = conexion.execute("""
            SELECT
                COALESCE(dc.costo_unitario, 0),
                COALESCE(c.proveedor, '')
            FROM detalle_compras dc
            JOIN compras c ON c.id = dc.compra_id
            WHERE dc.producto = ?
              AND UPPER(COALESCE(c.estado, 'RECIBIDA')) <> 'ANULADA'
            ORDER BY c.id DESC, dc.id DESC
            LIMIT 1
        """, (producto,)).fetchone()

    costo_inventario = float(fila[0] or 0) if fila else 0
    promedio = float(fila[1] or 0) if fila else 0
    ultimo_costo = float(ultimo[0] or 0) if ultimo else costo_inventario
    ultimo_proveedor = ultimo[1] if ultimo and ultimo[1] else "—"

    lbl_ultimo_costo.config(text=moneda(ultimo_costo))
    lbl_costo_promedio.config(text=moneda(promedio))
    lbl_ultimo_proveedor.config(text=ultimo_proveedor)

    if not entry_costo.get().strip() and ultimo_costo > 0:
        entry_costo.insert(0, f"{ultimo_costo:.2f}")


def crear_proveedor_rapido():
    nombre = combo_proveedor.get().strip()
    if not nombre:
        messagebox.showwarning(
            "Proveedor",
            "Escriba el nombre del proveedor."
        )
        return

    with conectar() as conexion:
        existe = conexion.execute("""
            SELECT COUNT(*)
            FROM proveedores
            WHERE UPPER(TRIM(nombre)) = UPPER(TRIM(?))
        """, (nombre,)).fetchone()[0]

        if existe:
            messagebox.showinfo(
                "Proveedor",
                "El proveedor ya está registrado."
            )
            return

        conexion.execute("""
            INSERT INTO proveedores(nombre, telefono, ciudad, correo)
            VALUES (?, '', '', '')
        """, (nombre,))
        registrar_auditoria(
            conexion,
            "CREAR PROVEEDOR RÁPIDO",
            f"Proveedor: {nombre}"
        )
        conexion.commit()

    cargar_proveedores()
    combo_proveedor.set(nombre)
    messagebox.showinfo(
        "Proveedor",
        "Proveedor creado correctamente."
    )


def crear_producto_rapido():
    producto = combo_producto.get().strip()
    presentacion = combo_presentacion.get().strip()

    if not producto or not presentacion:
        messagebox.showwarning(
            "Producto",
            "Escriba producto y presentación."
        )
        return

    with conectar() as conexion:
        existe = conexion.execute("""
            SELECT COUNT(*)
            FROM inventario
            WHERE UPPER(TRIM(producto)) = UPPER(TRIM(?))
              AND UPPER(TRIM(presentacion)) = UPPER(TRIM(?))
              AND COALESCE(lote, '') = ''
        """, (producto, presentacion)).fetchone()[0]

        if existe:
            messagebox.showinfo(
                "Producto",
                "Ese producto y presentación ya existen."
            )
            return

        conexion.execute("""
            INSERT INTO inventario(
                producto, presentacion, cantidad, lote,
                costo_unitario, fecha_ingreso, numero_despacho,
                stock_minimo, costo, fecha, despacho
            )
            VALUES (?, ?, 0, '', 0, ?, '', 0, 0, ?, '')
        """, (producto, presentacion, fecha_actual(), fecha_actual()))

        registrar_auditoria(
            conexion,
            "CREAR PRODUCTO RÁPIDO",
            f"{producto} / {presentacion}"
        )
        conexion.commit()

    cargar_productos()
    cargar_presentaciones()
    messagebox.showinfo(
        "Producto",
        "Producto creado correctamente y listo para comprar."
    )


def exportar_historial_excel():
    filas = [
        tabla_historial.item(item, "values")
        for item in tabla_historial.get_children()
    ]

    if not filas:
        messagebox.showwarning(
            "Exportar",
            "No hay compras para exportar."
        )
        return

    carpeta = os.path.join(BASE_DIR, "reportes")
    os.makedirs(carpeta, exist_ok=True)
    ruta = os.path.join(
        carpeta,
        f"compras_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

    libro = Workbook()
    hoja = libro.active
    hoja.title = "Compras"

    encabezados = [
        "Compra", "Fecha", "Proveedor", "Factura",
        "Forma de pago", "Total", "Estado", "Vencimiento"
    ]
    hoja.append(encabezados)

    for celda in hoja[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="0F5C8E")
        celda.alignment = Alignment(horizontal="center")

    for fila in filas:
        hoja.append(list(fila))

    anchos = [16, 14, 32, 18, 18, 18, 16, 16]
    for indice, ancho in enumerate(anchos, start=1):
        hoja.column_dimensions[
            chr(64 + indice)
        ].width = ancho

    libro.save(ruta)

    with conectar() as conexion:
        registrar_auditoria(
            conexion,
            "EXPORTAR COMPRAS A EXCEL",
            ruta
        )
        conexion.commit()

    messagebox.showinfo(
        "Exportación terminada",
        f"Archivo creado correctamente:\n\n{ruta}"
    )


# ============================================================
# DETALLE TEMPORAL
# ============================================================

def recalcular_totales():
    subtotal = sum(item["subtotal"] for item in detalle_actual)

    try:
        porcentaje_iva = a_numero(
            entry_iva_pct.get() or 0,
            "IVA",
            permitir_cero=True
        )
    except ValueError:
        porcentaje_iva = 0

    valor_iva = subtotal * porcentaje_iva / 100
    total = subtotal + valor_iva

    lbl_subtotal.config(text=moneda(subtotal))
    lbl_iva.config(text=moneda(valor_iva))
    lbl_total.config(text=moneda(total))


def mostrar_detalle_temporal():
    tabla_detalle.delete(*tabla_detalle.get_children())

    for indice, item in enumerate(detalle_actual, start=1):
        tabla_detalle.insert(
            "",
            "end",
            iid=str(indice - 1),
            values=(
                indice,
                item["producto"],
                item["presentacion"],
                f'{item["cantidad"]:,.2f}',
                item["lote"],
                f'{item["costo_unitario"]:,.2f}',
                f'{item["subtotal"]:,.2f}'
            )
        )

    recalcular_totales()


def limpiar_linea():
    combo_producto.set("")
    combo_presentacion.set("")
    entry_cantidad.delete(0, "end")
    entry_costo.delete(0, "end")
    entry_lote.delete(0, "end")
    combo_producto.focus_set()


def agregar_linea():
    producto = combo_producto.get().strip()
    presentacion = combo_presentacion.get().strip()
    lote = entry_lote.get().strip()

    if not producto:
        messagebox.showerror("Detalle", "Ingrese o seleccione el producto.")
        return

    if not presentacion:
        messagebox.showerror(
            "Detalle",
            "Ingrese o seleccione la presentación."
        )
        return

    try:
        cantidad = a_numero(entry_cantidad.get(), "Cantidad")
        costo = a_numero(entry_costo.get(), "Costo unitario")
    except ValueError as error:
        messagebox.showerror("Detalle", str(error))
        return

    subtotal = cantidad * costo

    detalle_actual.append({
        "producto": producto,
        "presentacion": presentacion,
        "cantidad": cantidad,
        "costo_unitario": costo,
        "subtotal": subtotal,
        "lote": lote
    })

    mostrar_detalle_temporal()
    limpiar_linea()


def eliminar_linea():
    seleccion = tabla_detalle.selection()

    if not seleccion:
        messagebox.showwarning(
            "Detalle",
            "Seleccione una línea para eliminar."
        )
        return

    indice = int(seleccion[0])

    if 0 <= indice < len(detalle_actual):
        detalle_actual.pop(indice)

    mostrar_detalle_temporal()


def limpiar_compra():
    global detalle_actual, compra_seleccionada_id

    detalle_actual = []
    compra_seleccionada_id = None

    entry_fecha.delete(0, "end")
    entry_fecha.insert(0, fecha_actual())

    combo_proveedor.set("")
    entry_factura.delete(0, "end")
    combo_forma_pago.set("CRÉDITO")
    entry_dias.delete(0, "end")
    entry_dias.insert(0, "30")
    entry_iva_pct.delete(0, "end")
    entry_iva_pct.insert(0, "0")
    entry_observaciones.delete("1.0", "end")

    lbl_numero.config(text=numero_compra_siguiente())
    lbl_estado_edicion.config(text="NUEVA COMPRA", fg=COLOR_VERDE)

    limpiar_linea()
    mostrar_detalle_temporal()


# ============================================================
# INTEGRACIÓN: INVENTARIO Y KARDEX
# ============================================================

def registrar_kardex(
    conexion,
    producto,
    presentacion,
    lote,
    movimiento,
    entrada,
    salida,
    costo,
    origen,
    observaciones
):
    saldo = conexion.execute("""
        SELECT IFNULL(SUM(entrada - salida), 0)
        FROM kardex
        WHERE producto = ?
          AND presentacion = ?
          AND COALESCE(lote, '') = ?
    """, (producto, presentacion, lote)).fetchone()[0]

    saldo_nuevo = float(saldo or 0) + entrada - salida

    conexion.execute("""
        INSERT INTO kardex(
            fecha, producto, presentacion, movimiento,
            entrada, salida, saldo, costo_unitario,
            lote, origen, observaciones
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        fecha_hora_actual(),
        producto,
        presentacion,
        movimiento,
        entrada,
        salida,
        saldo_nuevo,
        costo,
        lote,
        origen,
        observaciones
    ))


def ingresar_inventario(
    conexion,
    producto,
    presentacion,
    cantidad,
    costo_nuevo,
    lote,
    fecha,
    documento
):
    fila = conexion.execute("""
        SELECT id, cantidad, COALESCE(costo_unitario, costo, 0)
        FROM inventario
        WHERE producto = ?
          AND presentacion = ?
          AND COALESCE(lote, '') = ?
        LIMIT 1
    """, (producto, presentacion, lote)).fetchone()

    if fila:
        inventario_id, cantidad_actual, costo_actual = fila
        cantidad_actual = float(cantidad_actual or 0)
        costo_actual = float(costo_actual or 0)

        cantidad_total = cantidad_actual + cantidad

        if cantidad_total > 0:
            costo_promedio = (
                cantidad_actual * costo_actual
                + cantidad * costo_nuevo
            ) / cantidad_total
        else:
            costo_promedio = costo_nuevo

        conexion.execute("""
            UPDATE inventario
            SET cantidad = ?,
                costo_unitario = ?,
                costo = ?,
                fecha_ingreso = ?,
                fecha = ?,
                numero_despacho = ?,
                despacho = ?
            WHERE id = ?
        """, (
            cantidad_total,
            costo_promedio,
            costo_promedio,
            fecha,
            fecha,
            documento,
            documento,
            inventario_id
        ))
    else:
        conexion.execute("""
            INSERT INTO inventario(
                producto, presentacion, cantidad, lote,
                costo_unitario, fecha_ingreso, numero_despacho,
                stock_minimo, costo, fecha, despacho
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, 0, ?, ?, ?)
        """, (
            producto,
            presentacion,
            cantidad,
            lote,
            costo_nuevo,
            fecha,
            documento,
            costo_nuevo,
            fecha,
            documento
        ))

    registrar_kardex(
        conexion,
        producto,
        presentacion,
        lote,
        "ENTRADA POR COMPRA",
        cantidad,
        0,
        costo_nuevo,
        documento,
        "Entrada automática desde módulo de Compras"
    )


def revertir_inventario(
    conexion,
    producto,
    presentacion,
    cantidad,
    costo,
    lote,
    documento
):
    fila = conexion.execute("""
        SELECT id, cantidad
        FROM inventario
        WHERE producto = ?
          AND presentacion = ?
          AND COALESCE(lote, '') = ?
        LIMIT 1
    """, (producto, presentacion, lote)).fetchone()

    if not fila:
        raise ValueError(
            f"No existe inventario para revertir: "
            f"{producto} / {presentacion} / lote {lote or 'sin lote'}."
        )

    inventario_id, existencia = fila
    existencia = float(existencia or 0)

    if existencia < cantidad:
        raise ValueError(
            f"No se puede anular. La existencia de {producto} "
            f"({existencia:,.2f}) es menor que la cantidad comprada "
            f"({cantidad:,.2f})."
        )

    conexion.execute("""
        UPDATE inventario
        SET cantidad = cantidad - ?
        WHERE id = ?
    """, (cantidad, inventario_id))

    registrar_kardex(
        conexion,
        producto,
        presentacion,
        lote,
        "ANULACIÓN DE COMPRA",
        0,
        cantidad,
        costo,
        documento,
        "Salida automática por anulación de compra"
    )


# ============================================================
# GUARDAR COMPRA
# ============================================================

def guardar_compra():
    fecha = entry_fecha.get().strip()
    proveedor = combo_proveedor.get().strip()
    factura = entry_factura.get().strip()
    forma_pago = combo_forma_pago.get().strip().upper()
    observaciones = entry_observaciones.get("1.0", "end").strip()

    if not fecha:
        messagebox.showerror("Compra", "Ingrese la fecha.")
        return

    try:
        datetime.strptime(fecha, "%Y-%m-%d")
    except ValueError:
        messagebox.showerror(
            "Compra",
            "La fecha debe tener formato AAAA-MM-DD."
        )
        return

    if not proveedor:
        messagebox.showerror("Compra", "Seleccione un proveedor.")
        return

    if not factura:
        messagebox.showerror(
            "Compra",
            "Ingrese el número de factura o documento del proveedor."
        )
        return

    if factura_duplicada(factura, proveedor):
        messagebox.showerror(
            "Compra duplicada",
            "Ya existe una compra activa con esa factura y proveedor."
        )
        return

    if forma_pago not in ("CONTADO", "CRÉDITO"):
        messagebox.showerror(
            "Compra",
            "Seleccione CONTADO o CRÉDITO."
        )
        return

    if not detalle_actual:
        messagebox.showerror(
            "Compra",
            "Agregue al menos un producto al detalle."
        )
        return

    try:
        dias_credito = int(entry_dias.get() or 0)
        porcentaje_iva = a_numero(
            entry_iva_pct.get() or 0,
            "IVA",
            permitir_cero=True
        )
    except ValueError as error:
        messagebox.showerror("Compra", str(error))
        return

    if forma_pago == "CRÉDITO" and dias_credito <= 0:
        messagebox.showerror(
            "Compra",
            "Para compra a crédito, los días deben ser mayores que cero."
        )
        return

    if forma_pago == "CONTADO":
        dias_credito = 0

    subtotal = sum(item["subtotal"] for item in detalle_actual)
    valor_iva = subtotal * porcentaje_iva / 100
    total = subtotal + valor_iva
    vencimiento = (
        datetime.strptime(fecha, "%Y-%m-%d")
        + timedelta(days=dias_credito)
    ).strftime("%Y-%m-%d")

    confirmacion = messagebox.askyesno(
        "Confirmar compra",
        f"Proveedor: {proveedor}\n"
        f"Factura: {factura}\n"
        f"Forma de pago: {forma_pago}\n"
        f"Total: {moneda(total)}\n\n"
        "¿Desea registrar la compra e ingresar el inventario?"
    )

    if not confirmacion:
        return

    conexion = conectar()

    try:
        conexion.execute("BEGIN IMMEDIATE")

        cursor = conexion.execute("""
            INSERT INTO compras(
                fecha, proveedor, tipo_compra, descripcion,
                valor, forma_pago, dias_credito,
                factura, subtotal, iva, total,
                estado, observaciones, vencimiento
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            fecha,
            proveedor,
            "MERCANCÍA",
            f"Factura {factura}",
            total,
            forma_pago,
            dias_credito,
            factura,
            subtotal,
            valor_iva,
            total,
            "RECIBIDA",
            observaciones,
            vencimiento
        ))

        compra_id = cursor.lastrowid
        documento = f"COMP-{compra_id:06d} / FACT-{factura}"

        for item in detalle_actual:
            conexion.execute("""
                INSERT INTO detalle_compras(
                    compra_id, producto, presentacion,
                    cantidad, costo_unitario, subtotal, lote
                )
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                compra_id,
                item["producto"],
                item["presentacion"],
                item["cantidad"],
                item["costo_unitario"],
                item["subtotal"],
                item["lote"]
            ))

            ingresar_inventario(
                conexion,
                item["producto"],
                item["presentacion"],
                item["cantidad"],
                item["costo_unitario"],
                item["lote"],
                fecha,
                documento
            )

        if forma_pago == "CRÉDITO":
            conexion.execute("""
                INSERT INTO cuentas_pagar(
                    proveedor, valor, saldo, fecha,
                    vencimiento, estado
                )
                VALUES (?, ?, ?, ?, ?, 'PENDIENTE')
            """, (
                proveedor,
                total,
                total,
                fecha,
                vencimiento
            ))

        registrar_auditoria(
            conexion,
            "REGISTRAR COMPRA",
            (
                f"Compra {documento}; proveedor {proveedor}; "
                f"total {total:.2f}; pago {forma_pago}"
            )
        )

        conexion.commit()

        messagebox.showinfo(
            "Compra registrada",
            f"Compra COMP-{compra_id:06d} registrada correctamente.\n\n"
            "El inventario y el kardex fueron actualizados."
            + (
                "\nSe creó la cuenta por pagar."
                if forma_pago == "CRÉDITO"
                else ""
            )
        )

        limpiar_compra()
        cargar_historial()
        cargar_productos()

    except Exception as error:
        conexion.rollback()
        messagebox.showerror(
            "Error al guardar",
            "No se registró ningún cambio porque la operación fue "
            f"revertida completamente.\n\nDetalle:\n{error}"
        )
    finally:
        conexion.close()


# ============================================================
# HISTORIAL Y CONSULTA
# ============================================================

def cargar_historial():
    criterio = entry_buscar.get().strip()
    estado = combo_filtro_estado.get().strip()

    sql = """
        SELECT
            id,
            fecha,
            proveedor,
            COALESCE(factura, ''),
            COALESCE(forma_pago, ''),
            COALESCE(total, valor, 0),
            COALESCE(estado, 'RECIBIDA'),
            COALESCE(vencimiento, '')
        FROM compras
        WHERE 1 = 1
    """
    parametros = []

    if criterio:
        sql += """
            AND (
                proveedor LIKE ?
                OR COALESCE(factura, '') LIKE ?
                OR CAST(id AS TEXT) LIKE ?
            )
        """
        patron = f"%{criterio}%"
        parametros.extend([patron, patron, patron])

    if estado and estado != "TODOS":
        sql += " AND UPPER(COALESCE(estado, 'RECIBIDA')) = ?"
        parametros.append(estado)

    sql += " ORDER BY id DESC"

    with conectar() as conexion:
        filas = conexion.execute(sql, parametros).fetchall()

    tabla_historial.delete(*tabla_historial.get_children())

    for fila in filas:
        tag = "anulada" if str(fila[6]).upper() == "ANULADA" else "normal"
        tabla_historial.insert(
            "",
            "end",
            iid=str(fila[0]),
            values=(
                f"COMP-{fila[0]:06d}",
                fila[1],
                fila[2],
                fila[3],
                fila[4],
                moneda(fila[5]),
                fila[6],
                fila[7]
            ),
            tags=(tag,)
        )


def seleccionar_compra(evento=None):
    global compra_seleccionada_id

    seleccion = tabla_historial.selection()
    if not seleccion:
        return

    compra_seleccionada_id = int(seleccion[0])
    cargar_detalle_consulta(compra_seleccionada_id)


def cargar_detalle_consulta(compra_id):
    with conectar() as conexion:
        compra = conexion.execute("""
            SELECT
                fecha, proveedor, COALESCE(factura, ''),
                COALESCE(forma_pago, ''),
                COALESCE(total, valor, 0),
                COALESCE(estado, 'RECIBIDA'),
                COALESCE(observaciones, '')
            FROM compras
            WHERE id = ?
        """, (compra_id,)).fetchone()

        detalles = conexion.execute("""
            SELECT
                producto, presentacion, cantidad,
                costo_unitario, subtotal, COALESCE(lote, '')
            FROM detalle_compras
            WHERE compra_id = ?
            ORDER BY id
        """, (compra_id,)).fetchall()

    tabla_consulta.delete(*tabla_consulta.get_children())

    for indice, fila in enumerate(detalles, start=1):
        tabla_consulta.insert(
            "",
            "end",
            values=(
                indice,
                fila[0],
                fila[1],
                f"{fila[2]:,.2f}",
                fila[5],
                moneda(fila[3]),
                moneda(fila[4])
            )
        )

    if compra:
        lbl_resumen_consulta.config(
            text=(
                f"COMP-{compra_id:06d}  |  {compra[0]}  |  "
                f"{compra[1]}  |  Factura {compra[2]}  |  "
                f"{compra[3]}  |  {moneda(compra[4])}  |  "
                f"Estado: {compra[5]}"
            )
        )


# ============================================================
# ANULACIÓN TRANSACCIONAL
# ============================================================

def anular_compra():
    seleccion = tabla_historial.selection()

    if not seleccion:
        messagebox.showwarning(
            "Anular compra",
            "Seleccione una compra del historial."
        )
        return

    compra_id = int(seleccion[0])

    with conectar() as conexion:
        compra = conexion.execute("""
            SELECT
                proveedor,
                COALESCE(factura, ''),
                COALESCE(forma_pago, ''),
                COALESCE(total, valor, 0),
                COALESCE(estado, 'RECIBIDA'),
                fecha,
                COALESCE(vencimiento, '')
            FROM compras
            WHERE id = ?
        """, (compra_id,)).fetchone()

    if not compra:
        messagebox.showerror("Anular compra", "No se encontró la compra.")
        return

    if str(compra[4]).upper() == "ANULADA":
        messagebox.showinfo("Anular compra", "La compra ya está anulada.")
        return

    confirmar = messagebox.askyesno(
        "Confirmar anulación",
        f"Compra: COMP-{compra_id:06d}\n"
        f"Proveedor: {compra[0]}\n"
        f"Factura: {compra[1]}\n"
        f"Total: {moneda(compra[3])}\n\n"
        "La operación retirará del inventario todas las cantidades "
        "ingresadas y cancelará la cuenta por pagar pendiente.\n\n"
        "¿Desea continuar?"
    )

    if not confirmar:
        return

    conexion = conectar()

    try:
        conexion.execute("BEGIN IMMEDIATE")

        detalles = conexion.execute("""
            SELECT
                producto, presentacion, cantidad,
                costo_unitario, COALESCE(lote, '')
            FROM detalle_compras
            WHERE compra_id = ?
        """, (compra_id,)).fetchall()

        if not detalles:
            raise ValueError(
                "La compra no tiene detalle integrado y no puede anularse "
                "automáticamente."
            )

        documento = f"COMP-{compra_id:06d} / FACT-{compra[1]}"

        for producto, presentacion, cantidad, costo, lote in detalles:
            revertir_inventario(
                conexion,
                producto,
                presentacion,
                float(cantidad),
                float(costo),
                lote,
                documento
            )

        conexion.execute("""
            UPDATE compras
            SET estado = 'ANULADA',
                fecha_anulacion = ?
            WHERE id = ?
        """, (fecha_hora_actual(), compra_id))

        if str(compra[2]).upper() == "CRÉDITO":
            cuenta = conexion.execute("""
                SELECT id, saldo
                FROM cuentas_pagar
                WHERE proveedor = ?
                  AND ABS(valor - ?) < 0.01
                  AND fecha = ?
                  AND vencimiento = ?
                  AND UPPER(COALESCE(estado, '')) = 'PENDIENTE'
                ORDER BY id DESC
                LIMIT 1
            """, (
                compra[0],
                float(compra[3]),
                compra[5],
                compra[6]
            )).fetchone()

            if cuenta:
                if float(cuenta[1] or 0) < float(compra[3]) - 0.01:
                    raise ValueError(
                        "La cuenta por pagar asociada ya tiene pagos. "
                        "Primero revierta los pagos antes de anular la compra."
                    )

                conexion.execute("""
                    UPDATE cuentas_pagar
                    SET saldo = 0,
                        estado = 'ANULADA'
                    WHERE id = ?
                """, (cuenta[0],))

        registrar_auditoria(
            conexion,
            "ANULAR COMPRA",
            f"Compra COMP-{compra_id:06d}; factura {compra[1]}"
        )

        conexion.commit()

        messagebox.showinfo(
            "Compra anulada",
            "La compra fue anulada y el inventario fue revertido."
        )

        cargar_historial()
        tabla_consulta.delete(*tabla_consulta.get_children())
        lbl_resumen_consulta.config(text="Seleccione una compra")

    except Exception as error:
        conexion.rollback()
        messagebox.showerror(
            "No fue posible anular",
            "La operación fue revertida y no se hicieron cambios.\n\n"
            f"Detalle:\n{error}"
        )
    finally:
        conexion.close()


# ============================================================
# INDICADORES
# ============================================================

def actualizar_indicadores():
    with conectar() as conexion:
        hoy = conexion.execute("""
            SELECT IFNULL(SUM(COALESCE(total, valor, 0)), 0)
            FROM compras
            WHERE date(fecha) = date('now', 'localtime')
              AND UPPER(COALESCE(estado, 'RECIBIDA')) <> 'ANULADA'
        """).fetchone()[0]

        mes = conexion.execute("""
            SELECT IFNULL(SUM(COALESCE(total, valor, 0)), 0)
            FROM compras
            WHERE strftime('%Y-%m', fecha) =
                  strftime('%Y-%m', 'now', 'localtime')
              AND UPPER(COALESCE(estado, 'RECIBIDA')) <> 'ANULADA'
        """).fetchone()[0]

        pendientes = conexion.execute("""
            SELECT IFNULL(SUM(saldo), 0)
            FROM cuentas_pagar
            WHERE UPPER(COALESCE(estado, '')) = 'PENDIENTE'
        """).fetchone()[0]

        cantidad = conexion.execute("""
            SELECT COUNT(*)
            FROM compras
            WHERE UPPER(COALESCE(estado, 'RECIBIDA')) <> 'ANULADA'
        """).fetchone()[0]

    lbl_kpi_hoy.config(text=moneda(hoy))
    lbl_kpi_mes.config(text=moneda(mes))
    lbl_kpi_cxp.config(text=moneda(pendientes))
    lbl_kpi_cantidad.config(text=f"{cantidad:,}")


def refrescar_todo():
    cargar_proveedores()
    cargar_productos()
    cargar_historial()
    actualizar_indicadores()


# ============================================================
# INTERFAZ
# ============================================================

inicializar_bd()

ventana = tk.Tk()
ventana.title("ERP Café Alto de la Cruz - Compras Integradas")
ventana.geometry("1450x850")
ventana.minsize(1180, 700)
ventana.configure(bg=COLOR_FONDO)

try:
    ventana.state("zoomed")
except tk.TclError:
    pass

estilo = ttk.Style()
try:
    estilo.theme_use("clam")
except tk.TclError:
    pass

estilo.configure(
    "Treeview",
    rowheight=28,
    font=("Segoe UI", 9)
)
estilo.configure(
    "Treeview.Heading",
    font=("Segoe UI", 9, "bold")
)
estilo.map(
    "Treeview",
    background=[("selected", COLOR_AZUL)],
    foreground=[("selected", "white")]
)

# Encabezado
header = tk.Frame(ventana, bg=COLOR_AZUL_OSCURO, height=82)
header.pack(fill="x")
header.pack_propagate(False)

tk.Label(
    header,
    text="COMPRAS INTEGRADAS",
    font=("Segoe UI", 22, "bold"),
    bg=COLOR_AZUL_OSCURO,
    fg="white"
).pack(side="left", padx=24, pady=20)

lbl_estado_edicion = tk.Label(
    header,
    text="NUEVA COMPRA",
    font=("Segoe UI", 10, "bold"),
    bg=COLOR_AZUL_OSCURO,
    fg="#86EFAC"
)
lbl_estado_edicion.pack(side="right", padx=24)

# KPIs
kpis = tk.Frame(ventana, bg=COLOR_FONDO)
kpis.pack(fill="x", padx=18, pady=(14, 5))

for columna in range(4):
    kpis.grid_columnconfigure(columna, weight=1)

def tarjeta_kpi(columna, titulo, color):
    marco = tk.Frame(
        kpis,
        bg=COLOR_BLANCO,
        highlightbackground=COLOR_BORDE,
        highlightthickness=1
    )
    marco.grid(row=0, column=columna, sticky="ew", padx=6)
    tk.Frame(marco, bg=color, width=5).pack(side="left", fill="y")
    interno = tk.Frame(marco, bg=COLOR_BLANCO)
    interno.pack(fill="both", expand=True, padx=14, pady=10)
    tk.Label(
        interno,
        text=titulo,
        bg=COLOR_BLANCO,
        fg=COLOR_SUAVE,
        font=("Segoe UI", 8, "bold")
    ).pack(anchor="w")
    etiqueta = tk.Label(
        interno,
        text="$0",
        bg=COLOR_BLANCO,
        fg=COLOR_TEXTO,
        font=("Segoe UI", 16, "bold")
    )
    etiqueta.pack(anchor="w", pady=(3, 0))
    return etiqueta

lbl_kpi_hoy = tarjeta_kpi(0, "COMPRAS DE HOY", COLOR_AZUL)
lbl_kpi_mes = tarjeta_kpi(1, "COMPRAS DEL MES", COLOR_VERDE)
lbl_kpi_cxp = tarjeta_kpi(2, "CUENTAS POR PAGAR", COLOR_NARANJA)
lbl_kpi_cantidad = tarjeta_kpi(3, "COMPRAS ACTIVAS", "#7C3AED")

notebook = ttk.Notebook(ventana)
notebook.pack(fill="both", expand=True, padx=18, pady=10)

pestana_registro = tk.Frame(notebook, bg=COLOR_FONDO)
pestana_historial = tk.Frame(notebook, bg=COLOR_FONDO)

notebook.add(pestana_registro, text="  Registrar compra  ")
notebook.add(pestana_historial, text="  Historial y anulación  ")

# ------------------------------------------------------------
# PESTAÑA REGISTRO
# ------------------------------------------------------------

encabezado = tk.LabelFrame(
    pestana_registro,
    text="ENCABEZADO DE LA COMPRA",
    bg=COLOR_BLANCO,
    fg=COLOR_TEXTO,
    font=("Segoe UI", 10, "bold"),
    padx=12,
    pady=10
)
encabezado.pack(fill="x", padx=10, pady=10)

for columna in range(7):
    encabezado.grid_columnconfigure(columna, weight=1)

tk.Label(
    encabezado,
    text="Número interno",
    bg=COLOR_BLANCO,
    fg=COLOR_SUAVE
).grid(row=0, column=0, sticky="w")
lbl_numero = tk.Label(
    encabezado,
    text=numero_compra_siguiente(),
    bg=COLOR_BLANCO,
    fg=COLOR_AZUL,
    font=("Segoe UI", 10, "bold")
)
lbl_numero.grid(row=1, column=0, sticky="w", pady=(3, 8))

tk.Label(
    encabezado, text="Fecha", bg=COLOR_BLANCO, fg=COLOR_SUAVE
).grid(row=0, column=1, sticky="w")
entry_fecha = ttk.Entry(encabezado, width=14)
entry_fecha.grid(row=1, column=1, sticky="ew", padx=(0, 8), pady=(3, 8))
entry_fecha.insert(0, fecha_actual())

tk.Label(
    encabezado, text="Proveedor", bg=COLOR_BLANCO, fg=COLOR_SUAVE
).grid(row=0, column=2, sticky="w")
combo_proveedor = ttk.Combobox(encabezado, state="normal", width=26)
combo_proveedor.grid(
    row=1, column=2, columnspan=2, sticky="ew", padx=(0, 8), pady=(3, 8)
)
combo_proveedor.bind("<KeyRelease>", autocompletar_proveedor)

tk.Button(
    encabezado,
    text="+ Proveedor",
    command=crear_proveedor_rapido,
    bg="#475569",
    fg="white",
    relief="flat",
    cursor="hand2"
).grid(row=1, column=4, sticky="e", padx=(0, 8), pady=(3, 8))

tk.Label(
    encabezado,
    text="Factura proveedor",
    bg=COLOR_BLANCO,
    fg=COLOR_SUAVE
).grid(row=2, column=4, sticky="w")
entry_factura = ttk.Entry(encabezado, width=18)
entry_factura.grid(row=3, column=4, sticky="ew", padx=(0, 8), pady=(3, 8))

tk.Label(
    encabezado,
    text="Forma de pago",
    bg=COLOR_BLANCO,
    fg=COLOR_SUAVE
).grid(row=0, column=5, sticky="w")
combo_forma_pago = ttk.Combobox(
    encabezado,
    values=["CONTADO", "CRÉDITO"],
    state="readonly",
    width=14
)
combo_forma_pago.grid(
    row=1, column=5, sticky="ew", padx=(0, 8), pady=(3, 8)
)
combo_forma_pago.set("CRÉDITO")

tk.Label(
    encabezado,
    text="Días crédito",
    bg=COLOR_BLANCO,
    fg=COLOR_SUAVE
).grid(row=0, column=6, sticky="w")
entry_dias = ttk.Entry(encabezado, width=10)
entry_dias.grid(row=1, column=6, sticky="ew", pady=(3, 8))
entry_dias.insert(0, "30")

tk.Label(
    encabezado,
    text="IVA %",
    bg=COLOR_BLANCO,
    fg=COLOR_SUAVE
).grid(row=2, column=0, sticky="w")
entry_iva_pct = ttk.Entry(encabezado, width=10)
entry_iva_pct.grid(row=3, column=0, sticky="ew", padx=(0, 8))
entry_iva_pct.insert(0, "0")
entry_iva_pct.bind("<KeyRelease>", lambda evento: recalcular_totales())

tk.Label(
    encabezado,
    text="Observaciones",
    bg=COLOR_BLANCO,
    fg=COLOR_SUAVE
).grid(row=2, column=1, columnspan=6, sticky="w")
entry_observaciones = tk.Text(
    encabezado,
    height=2,
    relief="solid",
    bd=1,
    font=("Segoe UI", 9)
)
entry_observaciones.grid(
    row=3,
    column=1,
    columnspan=6,
    sticky="ew"
)

detalle_frame = tk.LabelFrame(
    pestana_registro,
    text="DETALLE DE PRODUCTOS",
    bg=COLOR_BLANCO,
    fg=COLOR_TEXTO,
    font=("Segoe UI", 10, "bold"),
    padx=12,
    pady=10
)
detalle_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))

linea = tk.Frame(detalle_frame, bg=COLOR_BLANCO)
linea.pack(fill="x", pady=(0, 8))

campos = [
    ("Producto", 0),
    ("Presentación", 1),
    ("Cantidad", 2),
    ("Lote", 3),
    ("Costo unitario", 4)
]
for texto, columna in campos:
    tk.Label(
        linea, text=texto, bg=COLOR_BLANCO, fg=COLOR_SUAVE
    ).grid(row=0, column=columna, sticky="w", padx=4)

combo_producto = ttk.Combobox(linea, state="normal", width=28)
combo_producto.grid(row=1, column=0, padx=4, sticky="ew")
combo_producto.bind("<KeyRelease>", autocompletar_producto)
combo_producto.bind("<<ComboboxSelected>>", actualizar_informacion_producto)

combo_presentacion = ttk.Combobox(linea, state="normal", width=20)
combo_presentacion.grid(row=1, column=1, padx=4, sticky="ew")
combo_presentacion.bind("<<ComboboxSelected>>", actualizar_informacion_producto)

entry_cantidad = ttk.Entry(linea, width=12)
entry_cantidad.grid(row=1, column=2, padx=4, sticky="ew")

entry_lote = ttk.Entry(linea, width=16)
entry_lote.grid(row=1, column=3, padx=4, sticky="ew")

entry_costo = ttk.Entry(linea, width=16)
entry_costo.grid(row=1, column=4, padx=4, sticky="ew")

tk.Button(
    linea,
    text="Agregar producto",
    command=agregar_linea,
    bg=COLOR_AZUL,
    fg="white",
    activebackground="#0B4B75",
    activeforeground="white",
    relief="flat",
    font=("Segoe UI", 9, "bold"),
    cursor="hand2",
    padx=12,
    pady=6
).grid(row=1, column=5, padx=8)

tk.Button(
    linea,
    text="+ Crear producto",
    command=crear_producto_rapido,
    bg="#475569",
    fg="white",
    activebackground="#334155",
    activeforeground="white",
    relief="flat",
    font=("Segoe UI", 9, "bold"),
    cursor="hand2",
    padx=10,
    pady=6
).grid(row=1, column=6, padx=4)

for columna in range(5):
    linea.grid_columnconfigure(columna, weight=1)

info_producto = tk.Frame(detalle_frame, bg="#F8FAFC")
info_producto.pack(fill="x", pady=(0, 8))

tk.Label(
    info_producto,
    text="Último costo:",
    bg="#F8FAFC",
    fg=COLOR_SUAVE
).pack(side="left", padx=(10, 4))
lbl_ultimo_costo = tk.Label(
    info_producto,
    text="$0",
    bg="#F8FAFC",
    fg=COLOR_TEXTO,
    font=("Segoe UI", 9, "bold")
)
lbl_ultimo_costo.pack(side="left", padx=(0, 18))

tk.Label(
    info_producto,
    text="Costo promedio:",
    bg="#F8FAFC",
    fg=COLOR_SUAVE
).pack(side="left", padx=(0, 4))
lbl_costo_promedio = tk.Label(
    info_producto,
    text="$0",
    bg="#F8FAFC",
    fg=COLOR_TEXTO,
    font=("Segoe UI", 9, "bold")
)
lbl_costo_promedio.pack(side="left", padx=(0, 18))

tk.Label(
    info_producto,
    text="Último proveedor:",
    bg="#F8FAFC",
    fg=COLOR_SUAVE
).pack(side="left", padx=(0, 4))
lbl_ultimo_proveedor = tk.Label(
    info_producto,
    text="—",
    bg="#F8FAFC",
    fg=COLOR_AZUL,
    font=("Segoe UI", 9, "bold")
)
lbl_ultimo_proveedor.pack(side="left")

columnas_detalle = (
    "N", "Producto", "Presentacion",
    "Cantidad", "Lote", "Costo", "Subtotal"
)
tabla_detalle = ttk.Treeview(
    detalle_frame,
    columns=columnas_detalle,
    show="headings",
    height=10
)

anchos = {
    "N": 45,
    "Producto": 240,
    "Presentacion": 150,
    "Cantidad": 100,
    "Lote": 130,
    "Costo": 120,
    "Subtotal": 130
}

for columna in columnas_detalle:
    tabla_detalle.heading(columna, text=columna)
    tabla_detalle.column(
        columna,
        width=anchos[columna],
        anchor="center" if columna != "Producto" else "w"
    )

tabla_detalle.pack(fill="both", expand=True)

acciones_detalle = tk.Frame(detalle_frame, bg=COLOR_BLANCO)
acciones_detalle.pack(fill="x", pady=(8, 0))

tk.Button(
    acciones_detalle,
    text="Eliminar línea seleccionada",
    command=eliminar_linea,
    bg=COLOR_ROJO,
    fg="white",
    relief="flat",
    font=("Segoe UI", 9, "bold"),
    cursor="hand2",
    padx=12,
    pady=6
).pack(side="left")

totales = tk.Frame(acciones_detalle, bg=COLOR_BLANCO)
totales.pack(side="right")

tk.Label(
    totales, text="Subtotal:", bg=COLOR_BLANCO, fg=COLOR_SUAVE
).grid(row=0, column=0, padx=6)
lbl_subtotal = tk.Label(
    totales, text="$0", bg=COLOR_BLANCO, fg=COLOR_TEXTO,
    font=("Segoe UI", 11, "bold")
)
lbl_subtotal.grid(row=0, column=1, padx=6)

tk.Label(
    totales, text="IVA:", bg=COLOR_BLANCO, fg=COLOR_SUAVE
).grid(row=0, column=2, padx=6)
lbl_iva = tk.Label(
    totales, text="$0", bg=COLOR_BLANCO, fg=COLOR_TEXTO,
    font=("Segoe UI", 11, "bold")
)
lbl_iva.grid(row=0, column=3, padx=6)

tk.Label(
    totales, text="TOTAL:", bg=COLOR_BLANCO, fg=COLOR_AZUL,
    font=("Segoe UI", 10, "bold")
).grid(row=0, column=4, padx=6)
lbl_total = tk.Label(
    totales, text="$0", bg=COLOR_BLANCO, fg=COLOR_AZUL,
    font=("Segoe UI", 16, "bold")
)
lbl_total.grid(row=0, column=5, padx=6)

botonera = tk.Frame(pestana_registro, bg=COLOR_FONDO)
botonera.pack(fill="x", padx=10, pady=(0, 12))

tk.Button(
    botonera,
    text="GUARDAR E INTEGRAR COMPRA",
    command=guardar_compra,
    bg=COLOR_VERDE,
    fg="white",
    activebackground="#166534",
    activeforeground="white",
    relief="flat",
    font=("Segoe UI", 10, "bold"),
    cursor="hand2",
    padx=20,
    pady=10
).pack(side="left")

tk.Button(
    botonera,
    text="Nueva / Limpiar",
    command=limpiar_compra,
    bg=COLOR_AZUL_OSCURO,
    fg="white",
    relief="flat",
    font=("Segoe UI", 10, "bold"),
    cursor="hand2",
    padx=20,
    pady=10
).pack(side="left", padx=8)

tk.Button(
    botonera,
    text="Actualizar listas",
    command=refrescar_todo,
    bg="#475569",
    fg="white",
    relief="flat",
    font=("Segoe UI", 10, "bold"),
    cursor="hand2",
    padx=20,
    pady=10
).pack(side="left")

# ------------------------------------------------------------
# PESTAÑA HISTORIAL
# ------------------------------------------------------------

filtros = tk.Frame(pestana_historial, bg=COLOR_BLANCO)
filtros.pack(fill="x", padx=10, pady=10)

tk.Label(
    filtros,
    text="Buscar proveedor, factura o ID:",
    bg=COLOR_BLANCO,
    fg=COLOR_TEXTO
).pack(side="left", padx=(12, 5))

entry_buscar = ttk.Entry(filtros, width=30)
entry_buscar.pack(side="left", padx=5)
entry_buscar.bind("<Return>", lambda evento: cargar_historial())

combo_filtro_estado = ttk.Combobox(
    filtros,
    values=["TODOS", "RECIBIDA", "ANULADA"],
    state="readonly",
    width=14
)
combo_filtro_estado.pack(side="left", padx=5)
combo_filtro_estado.set("TODOS")

tk.Button(
    filtros,
    text="Buscar",
    command=cargar_historial,
    bg=COLOR_AZUL,
    fg="white",
    relief="flat",
    cursor="hand2",
    padx=14,
    pady=6
).pack(side="left", padx=5)

tk.Button(
    filtros,
    text="Exportar Excel",
    command=exportar_historial_excel,
    bg=COLOR_VERDE,
    fg="white",
    relief="flat",
    cursor="hand2",
    font=("Segoe UI", 9, "bold"),
    padx=14,
    pady=6
).pack(side="right", padx=5)

tk.Button(
    filtros,
    text="Anular compra seleccionada",
    command=anular_compra,
    bg=COLOR_ROJO,
    fg="white",
    relief="flat",
    cursor="hand2",
    font=("Segoe UI", 9, "bold"),
    padx=14,
    pady=6
).pack(side="right", padx=12)

columnas_historial = (
    "Compra", "Fecha", "Proveedor", "Factura",
    "Pago", "Total", "Estado", "Vencimiento"
)
tabla_historial = ttk.Treeview(
    pestana_historial,
    columns=columnas_historial,
    show="headings",
    height=11
)

for columna in columnas_historial:
    tabla_historial.heading(columna, text=columna)

tabla_historial.column("Compra", width=110, anchor="center")
tabla_historial.column("Fecha", width=100, anchor="center")
tabla_historial.column("Proveedor", width=230, anchor="w")
tabla_historial.column("Factura", width=140, anchor="center")
tabla_historial.column("Pago", width=100, anchor="center")
tabla_historial.column("Total", width=130, anchor="e")
tabla_historial.column("Estado", width=100, anchor="center")
tabla_historial.column("Vencimiento", width=110, anchor="center")
tabla_historial.pack(fill="both", expand=True, padx=10)

tabla_historial.tag_configure("anulada", foreground=COLOR_ROJO)
tabla_historial.bind("<<TreeviewSelect>>", seleccionar_compra)

consulta_frame = tk.LabelFrame(
    pestana_historial,
    text="DETALLE DE LA COMPRA SELECCIONADA",
    bg=COLOR_BLANCO,
    fg=COLOR_TEXTO,
    font=("Segoe UI", 10, "bold")
)
consulta_frame.pack(fill="both", expand=True, padx=10, pady=10)

lbl_resumen_consulta = tk.Label(
    consulta_frame,
    text="Seleccione una compra",
    bg=COLOR_BLANCO,
    fg=COLOR_AZUL,
    font=("Segoe UI", 10, "bold")
)
lbl_resumen_consulta.pack(anchor="w", padx=10, pady=8)

tabla_consulta = ttk.Treeview(
    consulta_frame,
    columns=columnas_detalle,
    show="headings",
    height=7
)

for columna in columnas_detalle:
    tabla_consulta.heading(columna, text=columna)
    tabla_consulta.column(
        columna,
        width=anchos[columna],
        anchor="center" if columna != "Producto" else "w"
    )

tabla_consulta.pack(fill="both", expand=True, padx=10, pady=(0, 10))

# Barra inferior
barra_estado = tk.Frame(ventana, bg=COLOR_BLANCO, height=28)
barra_estado.pack(fill="x")

tk.Label(
    barra_estado,
    text=f"Base de datos: {RUTA_DB}",
    bg=COLOR_BLANCO,
    fg=COLOR_SUAVE,
    font=("Segoe UI", 8)
).pack(side="left", padx=12)

tk.Label(
    barra_estado,
    text="Compras Integradas v2.0",
    bg=COLOR_BLANCO,
    fg=COLOR_SUAVE,
    font=("Segoe UI", 8)
).pack(side="right", padx=12)

# Inicio
cargar_proveedores()
cargar_productos()
cargar_historial()
actualizar_indicadores()
limpiar_compra()

ventana.mainloop()
