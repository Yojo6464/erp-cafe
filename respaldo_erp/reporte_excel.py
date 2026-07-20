import tkinter as tk
from tkinter import ttk, messagebox
import sqlite3
import os

from openpyxl import Workbook

# =====================================
# CONEXION
# =====================================

RUTA_DB = r"C:\Users\jrive\visual\erp_cafe.db"

conexion = sqlite3.connect(RUTA_DB)
cursor = conexion.cursor()

# =====================================
# CARPETA REPORTES
# =====================================

CARPETA_REPORTES = r"C:\Users\jrive\visual\reportes"

if not os.path.exists(CARPETA_REPORTES):
    os.makedirs(CARPETA_REPORTES)

# =====================================
# REPORTE VENTAS
# =====================================

def reporte_ventas():

    try:

        archivo = os.path.join(
            CARPETA_REPORTES,
            "REPORTE_VENTAS.xlsx"
        )

        wb = Workbook()
        ws = wb.active

        ws.title = "Ventas"

        ws.append([
            "ID",
            "Fecha",
            "Cliente",
            "Producto",
            "Presentación",
            "Cantidad",
            "Precio Unitario",
            "Total"
        ])

        cursor.execute("""
        SELECT
            id,
            fecha,
            cliente,
            producto,
            presentacion,
            cantidad,
            precio_unitario,
            total
        FROM ventas
        ORDER BY id
        """)

        for fila in cursor.fetchall():
            ws.append(fila)

        wb.save(archivo)

        messagebox.showinfo(
            "Éxito",
            f"Reporte generado:\n{archivo}"
        )

    except Exception as e:

        messagebox.showerror(
            "Error",
            str(e)
        )

# =====================================
# REPORTE INVENTARIO
# =====================================

def reporte_inventario():

    try:

        archivo = os.path.join(
            CARPETA_REPORTES,
            "REPORTE_INVENTARIO.xlsx"
        )

        wb = Workbook()
        ws = wb.active

        ws.title = "Inventario"

        cursor.execute("""
        SELECT *
        FROM inventario
        """)

        columnas = [
            item[1]
            for item in cursor.execute(
                "PRAGMA table_info(inventario)"
            )
        ]

        ws.append(columnas)

        cursor.execute("""
        SELECT *
        FROM inventario
        """)

        for fila in cursor.fetchall():
            ws.append(fila)

        wb.save(archivo)

        messagebox.showinfo(
            "Éxito",
            f"Reporte generado:\n{archivo}"
        )

    except Exception as e:

        messagebox.showerror(
            "Error",
            str(e)
        )

# =====================================
# REPORTE CLIENTES
# =====================================

def reporte_clientes():

    try:

        archivo = os.path.join(
            CARPETA_REPORTES,
            "REPORTE_CLIENTES.xlsx"
        )

        wb = Workbook()
        ws = wb.active

        ws.title = "Clientes"

        ws.append([
            "ID",
            "Nombre",
            "Fecha Registro",
            "Teléfono",
            "Ciudad",
            "Correo"
        ])

        cursor.execute("""
        SELECT
            id,
            nombre,
            fecha_registro,
            telefono,
            ciudad,
            correo
        FROM clientes
        ORDER BY nombre
        """)

        for fila in cursor.fetchall():
            ws.append(fila)

        wb.save(archivo)

        messagebox.showinfo(
            "Éxito",
            f"Reporte generado:\n{archivo}"
        )

    except Exception as e:

        messagebox.showerror(
            "Error",
            str(e)
        )

# =====================================
# REPORTE CUENTAS POR COBRAR
# =====================================

def reporte_cxc():

    try:

        archivo = os.path.join(
            CARPETA_REPORTES,
            "REPORTE_CXC.xlsx"
        )

        wb = Workbook()
        ws = wb.active

        ws.title = "CXC"

        cursor.execute("""
        SELECT *
        FROM cuentas_cobrar
        """)

        columnas = [
            item[1]
            for item in cursor.execute(
                "PRAGMA table_info(cuentas_cobrar)"
            )
        ]

        ws.append(columnas)

        cursor.execute("""
        SELECT *
        FROM cuentas_cobrar
        """)

        for fila in cursor.fetchall():
            ws.append(fila)

        wb.save(archivo)

        messagebox.showinfo(
            "Éxito",
            f"Reporte generado:\n{archivo}"
        )

    except Exception as e:

        messagebox.showerror(
            "Error",
            str(e)
        )

# =====================================
# REPORTE RENTABILIDAD
# =====================================

def reporte_rentabilidad():

    try:

        archivo = os.path.join(
            CARPETA_REPORTES,
            "REPORTE_RENTABILIDAD.xlsx"
        )

        wb = Workbook()
        ws = wb.active

        ws.title = "Rentabilidad"

        ws.append([
            "Ventas Totales",
            "Costos Totales",
            "Utilidad Total"
        ])

        cursor.execute("""
        SELECT
            IFNULL(SUM(total),0),
            IFNULL(SUM(costo_unitario*cantidad),0),
            IFNULL(SUM(utilidad_total),0)
        FROM ventas
        """)

        ws.append(
            cursor.fetchone()
        )

        wb.save(archivo)

        messagebox.showinfo(
            "Éxito",
            f"Reporte generado:\n{archivo}"
        )

    except Exception as e:

        messagebox.showerror(
            "Error",
            str(e)
        )

# =====================================
# VENTANA
# =====================================

ventana = tk.Tk()

ventana.title(
    "ERP Café Alto de la Cruz - Reportes Excel"
)

ventana.geometry("600x500")

titulo = tk.Label(
    ventana,
    text="REPORTES EXCEL",
    font=("Arial", 18, "bold")
)

titulo.pack(pady=20)

tk.Button(
    ventana,
    text="Reporte de Ventas",
    width=35,
    height=2,
    command=reporte_ventas
).pack(pady=5)

tk.Button(
    ventana,
    text="Reporte de Inventario",
    width=35,
    height=2,
    command=reporte_inventario
).pack(pady=5)

tk.Button(
    ventana,
    text="Reporte de Clientes",
    width=35,
    height=2,
    command=reporte_clientes
).pack(pady=5)

tk.Button(
    ventana,
    text="Reporte Cuentas por Cobrar",
    width=35,
    height=2,
    command=reporte_cxc
).pack(pady=5)

tk.Button(
    ventana,
    text="Reporte de Rentabilidad",
    width=35,
    height=2,
    command=reporte_rentabilidad
).pack(pady=5)

ventana.mainloop()

conexion.close()