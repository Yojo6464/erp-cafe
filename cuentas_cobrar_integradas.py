import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
import sqlite3
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ============================================================
# BME-ERP - CUENTAS POR COBRAR INTEGRADAS v1.0
# ============================================================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
RUTA_DB = os.path.join(BASE_DIR, "erp_cafe.db")

C_FONDO = "#EEF3F8"
C_AZUL = "#0F5C8E"
C_OSCURO = "#153B5B"
C_VERDE = "#15803D"
C_NARANJA = "#C56A00"
C_ROJO = "#B42318"
C_MORADO = "#7C3AED"
C_TEXTO = "#1F2937"
C_SUAVE = "#64748B"
C_BLANCO = "#FFFFFF"
C_BORDE = "#D7E0E8"


def conectar():
    con = sqlite3.connect(RUTA_DB, timeout=20)
    con.execute("PRAGMA foreign_keys = ON")
    return con


def columnas_tabla(con, tabla):
    return {r[1] for r in con.execute(f"PRAGMA table_info({tabla})").fetchall()}


def agregar_columna(con, tabla, columna, definicion):
    if columna not in columnas_tabla(con, tabla):
        con.execute(f"ALTER TABLE {tabla} ADD COLUMN {columna} {definicion}")


def inicializar_bd():
    with conectar() as con:
        con.executescript("""
            CREATE TABLE IF NOT EXISTS cuentas_cobrar(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                fecha TEXT,
                cliente TEXT,
                concepto TEXT,
                valor REAL DEFAULT 0,
                saldo REAL DEFAULT 0,
                vencimiento TEXT,
                estado TEXT DEFAULT 'PENDIENTE'
            );

            CREATE TABLE IF NOT EXISTS recaudos_cartera(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                cuenta_id INTEGER NOT NULL,
                fecha TEXT NOT NULL,
                banco_id INTEGER NOT NULL,
                valor REAL NOT NULL,
                referencia TEXT DEFAULT '',
                observaciones TEXT DEFAULT '',
                usuario TEXT DEFAULT '',
                estado TEXT DEFAULT 'ACTIVO',
                creado_en TEXT DEFAULT CURRENT_TIMESTAMP,
                reversado_en TEXT DEFAULT '',
                motivo_reversion TEXT DEFAULT '',
                FOREIGN KEY(cuenta_id) REFERENCES cuentas_cobrar(id)
            );

            CREATE TABLE IF NOT EXISTS auditoria_erp(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                fecha_hora TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                usuario TEXT,
                rol TEXT,
                accion TEXT NOT NULL,
                detalle TEXT,
                modulo TEXT
            );

            CREATE INDEX IF NOT EXISTS idx_cxc_cliente
            ON cuentas_cobrar(cliente);

            CREATE INDEX IF NOT EXISTS idx_cxc_vencimiento
            ON cuentas_cobrar(vencimiento);

            CREATE INDEX IF NOT EXISTS idx_recaudos_cuenta
            ON recaudos_cartera(cuenta_id);
        """)

        for col, definicion in [
            ("venta_id", "INTEGER"),
            ("factura", "TEXT DEFAULT ''"),
            ("documento_cliente", "TEXT DEFAULT ''"),
            ("dias_mora", "INTEGER DEFAULT 0"),
            ("fecha_ultimo_pago", "TEXT DEFAULT ''"),
            ("observaciones", "TEXT DEFAULT ''")
        ]:
            agregar_columna(con, "cuentas_cobrar", col, definicion)

        con.commit()


def ahora():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def hoy():
    return datetime.now().strftime("%Y-%m-%d")


def moneda(v):
    return f"${float(v or 0):,.0f}"


def numero(v):
    return f"{float(v or 0):,.2f}"


def a_numero(valor, nombre):
    try:
        n = float(str(valor).replace(",", "").strip())
    except ValueError:
        raise ValueError(f"{nombre} debe ser numérico.")
    if n <= 0:
        raise ValueError(f"{nombre} debe ser mayor que cero.")
    return n


def auditoria(con, accion, detalle):
    con.execute("""
        INSERT INTO auditoria_erp(usuario, rol, accion, detalle, modulo)
        VALUES (?, ?, ?, ?, 'Cuentas por Cobrar')
    """, (
        os.environ.get("ERP_USUARIO", "usuario_local"),
        os.environ.get("ERP_ROL", "OPERADOR"),
        accion,
        detalle
    ))


def extraer_id_combo(valor):
    try:
        return int(str(valor).split("|")[0].strip())
    except Exception:
        return None


def sincronizar_desde_ventas():
    with conectar() as con:
        tablas = {
            r[0] for r in con.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
            ).fetchall()
        }

        if "ventas_integradas" not in tablas:
            return

        ventas = con.execute("""
            SELECT
                id, numero, fecha, cliente,
                documento_cliente, total, vencimiento, estado
            FROM ventas_integradas
            WHERE forma_pago='CRÉDITO'
              AND estado='ACTIVA'
        """).fetchall()

        for venta in ventas:
            existe = con.execute("""
                SELECT id
                FROM cuentas_cobrar
                WHERE venta_id=?
                   OR concepto=?
            """, (venta[0], f"Venta {venta[1]}")).fetchone()

            if not existe:
                con.execute("""
                    INSERT INTO cuentas_cobrar(
                        fecha, cliente, concepto, valor, saldo,
                        vencimiento, estado, venta_id, factura,
                        documento_cliente, observaciones
                    )
                    VALUES (?, ?, ?, ?, ?, ?, 'PENDIENTE', ?, ?, ?, ?)
                """, (
                    venta[2], venta[3], f"Venta {venta[1]}",
                    venta[5], venta[5], venta[6], venta[0],
                    venta[1], venta[4], "Generada automáticamente desde Ventas"
                ))
        con.commit()


def actualizar_estados():
    with conectar() as con:
        filas = con.execute("""
            SELECT id, valor, saldo, vencimiento, estado
            FROM cuentas_cobrar
        """).fetchall()

        hoy_dt = datetime.strptime(hoy(), "%Y-%m-%d")

        for fila in filas:
            cuenta_id, valor, saldo, vencimiento, estado = fila
            saldo = float(saldo or 0)
            valor = float(valor or 0)

            if str(estado).upper() == "ANULADA":
                continue

            dias_mora = 0
            if vencimiento:
                try:
                    fecha_v = datetime.strptime(vencimiento[:10], "%Y-%m-%d")
                    dias_mora = max((hoy_dt - fecha_v).days, 0)
                except ValueError:
                    dias_mora = 0

            if saldo <= 0.0001:
                nuevo = "PAGADA"
            elif saldo < valor:
                nuevo = "VENCIDA" if dias_mora > 0 else "PARCIAL"
            else:
                nuevo = "VENCIDA" if dias_mora > 0 else "PENDIENTE"

            con.execute("""
                UPDATE cuentas_cobrar
                SET estado=?, dias_mora=?
                WHERE id=?
            """, (nuevo, dias_mora, cuenta_id))

        con.commit()


def cargar_bancos():
    with conectar() as con:
        filas = con.execute("""
            SELECT id, banco, numero_cuenta, saldo
            FROM bancos
            WHERE UPPER(estado)='ACTIVA'
            ORDER BY banco, numero_cuenta
        """).fetchall()

    valores = [
        f"{f[0]} | {f[1]} | {f[2]} | {moneda(f[3])}"
        for f in filas
    ]
    combo_banco["values"] = valores


def cargar_cartera():
    actualizar_estados()

    criterio = entry_buscar.get().strip()
    estado = combo_estado.get().strip()

    sql = """
        SELECT
            id, COALESCE(factura,''), fecha, cliente, concepto,
            valor, saldo, vencimiento, dias_mora, estado
        FROM cuentas_cobrar
        WHERE 1=1
    """
    params = []

    if criterio:
        patron = f"%{criterio}%"
        sql += """
            AND (
                cliente LIKE ?
                OR concepto LIKE ?
                OR COALESCE(factura,'') LIKE ?
            )
        """
        params.extend([patron, patron, patron])

    if estado and estado != "TODOS":
        sql += " AND estado=?"
        params.append(estado)

    sql += " ORDER BY CASE WHEN saldo>0 THEN 0 ELSE 1 END, vencimiento, id DESC"

    with conectar() as con:
        filas = con.execute(sql, params).fetchall()

    tabla_cartera.delete(*tabla_cartera.get_children())

    for f in filas:
        if f[9] == "PAGADA":
            tag = "pagada"
        elif f[9] == "VENCIDA":
            tag = "vencida"
        elif f[9] == "PARCIAL":
            tag = "parcial"
        elif f[9] == "ANULADA":
            tag = "anulada"
        else:
            tag = "pendiente"

        tabla_cartera.insert(
            "", "end", iid=str(f[0]),
            values=(
                f[1], f[2], f[3], f[4],
                moneda(f[5]), moneda(f[6]),
                f[7], f[8], f[9]
            ),
            tags=(tag,)
        )

    actualizar_kpis()


def datos_cuenta_seleccionada():
    sel = tabla_cartera.selection()
    if not sel:
        return None

    cuenta_id = int(sel[0])
    with conectar() as con:
        return con.execute("""
            SELECT
                id, COALESCE(factura,''), fecha, cliente, concepto,
                valor, saldo, vencimiento, estado
            FROM cuentas_cobrar
            WHERE id=?
        """, (cuenta_id,)).fetchone()


def registrar_recaudo():
    cuenta = datos_cuenta_seleccionada()

    if not cuenta:
        messagebox.showwarning("Recaudo", "Seleccione una cuenta por cobrar.")
        return

    if cuenta[8] in ("PAGADA", "ANULADA"):
        messagebox.showerror(
            "Recaudo",
            f"No se puede recaudar una cuenta en estado {cuenta[8]}."
        )
        return

    banco_id = extraer_id_combo(combo_banco.get())
    if not banco_id:
        messagebox.showerror("Recaudo", "Seleccione la cuenta bancaria.")
        return

    try:
        valor = a_numero(entry_valor_recaudo.get(), "Valor del recaudo")
    except ValueError as e:
        messagebox.showerror("Recaudo", str(e))
        return

    saldo_actual = float(cuenta[6] or 0)

    if valor > saldo_actual + 0.0001:
        messagebox.showerror(
            "Recaudo",
            f"El recaudo no puede superar el saldo de {moneda(saldo_actual)}."
        )
        return

    referencia = entry_referencia.get().strip()
    observaciones = txt_obs_recaudo.get("1.0", "end").strip()

    if not messagebox.askyesno(
        "Confirmar recaudo",
        (
            f"Cliente: {cuenta[3]}\n"
            f"Cuenta: {cuenta[1] or cuenta[4]}\n"
            f"Saldo actual: {moneda(saldo_actual)}\n"
            f"Recaudo: {moneda(valor)}\n\n"
            "¿Desea registrar el recaudo?"
        )
    ):
        return

    con = conectar()

    try:
        con.execute("BEGIN IMMEDIATE")

        banco = con.execute("""
            SELECT saldo
            FROM bancos
            WHERE id=? AND UPPER(estado)='ACTIVA'
        """, (banco_id,)).fetchone()

        if not banco:
            raise ValueError("La cuenta bancaria no está disponible.")

        saldo_banco_anterior = float(banco[0] or 0)
        saldo_banco_nuevo = saldo_banco_anterior + valor

        con.execute("""
            INSERT INTO recaudos_cartera(
                cuenta_id, fecha, banco_id, valor,
                referencia, observaciones, usuario, estado
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, 'ACTIVO')
        """, (
            cuenta[0], hoy(), banco_id, valor,
            referencia, observaciones,
            os.environ.get("ERP_USUARIO", "usuario_local")
        ))

        nuevo_saldo = saldo_actual - valor

        if nuevo_saldo <= 0.0001:
            nuevo_estado = "PAGADA"
            nuevo_saldo = 0
        else:
            nuevo_estado = "PARCIAL"

        con.execute("""
            UPDATE cuentas_cobrar
            SET saldo=?,
                estado=?,
                fecha_ultimo_pago=?
            WHERE id=?
        """, (nuevo_saldo, nuevo_estado, hoy(), cuenta[0]))

        con.execute("""
            UPDATE bancos
            SET saldo=?
            WHERE id=?
        """, (saldo_banco_nuevo, banco_id))

        columnas_mov = columnas_tabla(con, "movimientos_bancos")

        if "saldo_anterior" in columnas_mov and "saldo_nuevo" in columnas_mov:
            con.execute("""
                INSERT INTO movimientos_bancos(
                    fecha, banco_id, tipo, concepto, valor,
                    saldo_anterior, saldo_nuevo, autorizado_por
                )
                VALUES (?, ?, 'CONSIGNACION', ?, ?, ?, ?, ?)
            """, (
                ahora(), banco_id,
                f"Recaudo cartera - {cuenta[3]} - {cuenta[1] or cuenta[4]}",
                valor, saldo_banco_anterior, saldo_banco_nuevo,
                os.environ.get("ERP_USUARIO", "usuario_local")
            ))
        else:
            con.execute("""
                INSERT INTO movimientos_bancos(
                    fecha, banco_id, tipo, concepto, valor,
                    saldo_resultante, referencia
                )
                VALUES (?, ?, 'INGRESO CARTERA', ?, ?, ?, ?)
            """, (
                ahora(), banco_id,
                f"Recaudo cartera - {cuenta[3]}",
                valor, saldo_banco_nuevo,
                referencia
            ))

        auditoria(
            con,
            "REGISTRAR RECAUDO",
            f"Cuenta {cuenta[0]}; cliente {cuenta[3]}; valor {valor:.2f}"
        )

        con.commit()

        messagebox.showinfo(
            "Recaudo",
            (
                "Recaudo registrado correctamente.\n\n"
                f"Nuevo saldo de cartera: {moneda(nuevo_saldo)}\n"
                f"Nuevo saldo bancario: {moneda(saldo_banco_nuevo)}"
            )
        )

        entry_valor_recaudo.delete(0, "end")
        entry_referencia.delete(0, "end")
        txt_obs_recaudo.delete("1.0", "end")

        cargar_cartera()
        cargar_recaudos()
        cargar_bancos()

    except Exception as e:
        con.rollback()
        messagebox.showerror(
            "Recaudo",
            f"La operación fue revertida completamente.\n\n{e}"
        )
    finally:
        con.close()


def cargar_recaudos():
    with conectar() as con:
        filas = con.execute("""
            SELECT
                r.id, r.fecha, c.cliente,
                COALESCE(c.factura, c.concepto),
                b.banco, b.numero_cuenta,
                r.valor, r.referencia, r.estado
            FROM recaudos_cartera r
            JOIN cuentas_cobrar c ON c.id=r.cuenta_id
            JOIN bancos b ON b.id=r.banco_id
            ORDER BY r.id DESC
        """).fetchall()

    tabla_recaudos.delete(*tabla_recaudos.get_children())

    for f in filas:
        tabla_recaudos.insert(
            "", "end", iid=str(f[0]),
            values=(
                f[1], f[2], f[3],
                f"{f[4]} / {f[5]}",
                moneda(f[6]), f[7], f[8]
            ),
            tags=("reversado" if f[8] == "REVERSADO" else "activo",)
        )


def reversar_recaudo():
    sel = tabla_recaudos.selection()

    if not sel:
        messagebox.showwarning("Reversión", "Seleccione un recaudo.")
        return

    recaudo_id = int(sel[0])

    with conectar() as con:
        r = con.execute("""
            SELECT
                r.cuenta_id, r.banco_id, r.valor, r.estado,
                c.cliente, COALESCE(c.factura, c.concepto)
            FROM recaudos_cartera r
            JOIN cuentas_cobrar c ON c.id=r.cuenta_id
            WHERE r.id=?
        """, (recaudo_id,)).fetchone()

    if not r:
        return

    if r[3] == "REVERSADO":
        messagebox.showinfo("Reversión", "El recaudo ya está reversado.")
        return

    motivo = simpledialog.askstring(
        "Motivo de reversión",
        "Indique el motivo:",
        parent=ventana
    )

    if not motivo:
        return

    if not messagebox.askyesno(
        "Confirmar reversión",
        (
            f"Cliente: {r[4]}\n"
            f"Documento: {r[5]}\n"
            f"Valor: {moneda(r[2])}\n\n"
            "¿Desea reversar este recaudo?"
        )
    ):
        return

    con = conectar()

    try:
        con.execute("BEGIN IMMEDIATE")

        banco = con.execute(
            "SELECT saldo FROM bancos WHERE id=?",
            (r[1],)
        ).fetchone()

        saldo_banco = float(banco[0] or 0)

        if saldo_banco < float(r[2] or 0):
            raise ValueError(
                "El banco no tiene saldo suficiente para reversar el recaudo."
            )

        cuenta = con.execute("""
            SELECT valor, saldo
            FROM cuentas_cobrar
            WHERE id=?
        """, (r[0],)).fetchone()

        nuevo_saldo_cartera = min(
            float(cuenta[0] or 0),
            float(cuenta[1] or 0) + float(r[2] or 0)
        )
        nuevo_estado = (
            "PENDIENTE"
            if abs(nuevo_saldo_cartera - float(cuenta[0] or 0)) < 0.0001
            else "PARCIAL"
        )
        nuevo_saldo_banco = saldo_banco - float(r[2] or 0)

        con.execute("""
            UPDATE recaudos_cartera
            SET estado='REVERSADO',
                reversado_en=?,
                motivo_reversion=?
            WHERE id=?
        """, (ahora(), motivo, recaudo_id))

        con.execute("""
            UPDATE cuentas_cobrar
            SET saldo=?, estado=?
            WHERE id=?
        """, (nuevo_saldo_cartera, nuevo_estado, r[0]))

        con.execute("""
            UPDATE bancos
            SET saldo=?
            WHERE id=?
        """, (nuevo_saldo_banco, r[1]))

        columnas_mov = columnas_tabla(con, "movimientos_bancos")

        if "saldo_anterior" in columnas_mov and "saldo_nuevo" in columnas_mov:
            con.execute("""
                INSERT INTO movimientos_bancos(
                    fecha, banco_id, tipo, concepto, valor,
                    saldo_anterior, saldo_nuevo, autorizado_por
                )
                VALUES (?, ?, 'RETIRO', ?, ?, ?, ?, ?)
            """, (
                ahora(), r[1],
                f"Reversión recaudo - {r[4]} - {r[5]}",
                r[2], saldo_banco, nuevo_saldo_banco,
                os.environ.get("ERP_USUARIO", "usuario_local")
            ))

        auditoria(
            con,
            "REVERSAR RECAUDO",
            f"Recaudo {recaudo_id}; valor {r[2]}; motivo {motivo}"
        )

        con.commit()

        messagebox.showinfo(
            "Reversión",
            "Recaudo reversado correctamente."
        )

        cargar_cartera()
        cargar_recaudos()
        cargar_bancos()

    except Exception as e:
        con.rollback()
        messagebox.showerror(
            "Reversión",
            f"La operación fue revertida.\n\n{e}"
        )
    finally:
        con.close()


def ver_estado_cuenta():
    cuenta = datos_cuenta_seleccionada()

    if not cuenta:
        messagebox.showwarning("Estado de cuenta", "Seleccione una cuenta.")
        return

    with conectar() as con:
        recaudos = con.execute("""
            SELECT
                r.fecha, b.banco, r.valor, r.referencia, r.estado
            FROM recaudos_cartera r
            JOIN bancos b ON b.id=r.banco_id
            WHERE r.cuenta_id=?
            ORDER BY r.id
        """, (cuenta[0],)).fetchall()

    top = tk.Toplevel(ventana)
    top.title(f"Estado de cuenta - {cuenta[3]}")
    top.geometry("950x580")
    top.configure(bg=C_FONDO)

    tk.Label(
        top,
        text=f"ESTADO DE CUENTA - {cuenta[3]}",
        bg=C_OSCURO,
        fg="white",
        font=("Segoe UI", 16, "bold"),
        pady=14
    ).pack(fill="x")

    resumen = tk.Frame(top, bg=C_BLANCO)
    resumen.pack(fill="x", padx=15, pady=15)

    datos = [
        ("Documento", cuenta[1] or cuenta[4]),
        ("Fecha", cuenta[2]),
        ("Valor inicial", moneda(cuenta[5])),
        ("Saldo actual", moneda(cuenta[6])),
        ("Vencimiento", cuenta[7]),
        ("Estado", cuenta[8])
    ]

    for i, (titulo, valor) in enumerate(datos):
        f = tk.Frame(resumen, bg=C_BLANCO)
        f.grid(row=i // 3, column=i % 3, sticky="nsew", padx=15, pady=10)
        tk.Label(
            f, text=titulo, bg=C_BLANCO, fg=C_SUAVE,
            font=("Segoe UI", 8, "bold")
        ).pack(anchor="w")
        tk.Label(
            f, text=valor, bg=C_BLANCO, fg=C_TEXTO,
            font=("Segoe UI", 11, "bold")
        ).pack(anchor="w")

    for c in range(3):
        resumen.grid_columnconfigure(c, weight=1)

    cols = ("Fecha", "Banco", "Valor", "Referencia", "Estado")
    tv = ttk.Treeview(top, columns=cols, show="headings")

    for col in cols:
        tv.heading(col, text=col)

    tv.pack(fill="both", expand=True, padx=15, pady=(0, 15))

    for f in recaudos:
        tv.insert(
            "", "end",
            values=(f[0], f[1], moneda(f[2]), f[3], f[4])
        )


def actualizar_kpis():
    actualizar_estados()

    with conectar() as con:
        total = con.execute("""
            SELECT IFNULL(SUM(saldo),0)
            FROM cuentas_cobrar
            WHERE estado NOT IN ('PAGADA','ANULADA')
        """).fetchone()[0]

        vencida = con.execute("""
            SELECT IFNULL(SUM(saldo),0)
            FROM cuentas_cobrar
            WHERE estado='VENCIDA'
        """).fetchone()[0]

        recaudado_mes = con.execute("""
            SELECT IFNULL(SUM(valor),0)
            FROM recaudos_cartera
            WHERE estado='ACTIVO'
              AND substr(fecha,1,7)=substr(?,1,7)
        """, (hoy(),)).fetchone()[0]

        clientes = con.execute("""
            SELECT COUNT(DISTINCT cliente)
            FROM cuentas_cobrar
            WHERE saldo>0
              AND estado NOT IN ('PAGADA','ANULADA')
        """).fetchone()[0]

    lbl_kpi_total.config(text=moneda(total))
    lbl_kpi_vencida.config(text=moneda(vencida))
    lbl_kpi_recaudado.config(text=moneda(recaudado_mes))
    lbl_kpi_clientes.config(text=str(clientes))


def exportar_excel():
    carpeta = os.path.join(BASE_DIR, "reportes")
    os.makedirs(carpeta, exist_ok=True)

    ruta = os.path.join(
        carpeta,
        "cuentas_cobrar_"
        + datetime.now().strftime("%Y%m%d_%H%M%S")
        + ".xlsx"
    )

    with conectar() as con:
        cartera = con.execute("""
            SELECT
                COALESCE(factura,''), fecha, cliente, concepto,
                valor, saldo, vencimiento, dias_mora, estado
            FROM cuentas_cobrar
            ORDER BY vencimiento, cliente
        """).fetchall()

        recaudos = con.execute("""
            SELECT
                r.fecha, c.cliente, COALESCE(c.factura,c.concepto),
                b.banco, b.numero_cuenta, r.valor,
                r.referencia, r.estado
            FROM recaudos_cartera r
            JOIN cuentas_cobrar c ON c.id=r.cuenta_id
            JOIN bancos b ON b.id=r.banco_id
            ORDER BY r.id DESC
        """).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Cartera"

    ws.append([
        "Factura", "Fecha", "Cliente", "Concepto",
        "Valor", "Saldo", "Vencimiento", "Días mora", "Estado"
    ])

    for celda in ws[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="153B5B")
        celda.alignment = Alignment(horizontal="center")

    for fila in cartera:
        ws.append(list(fila))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    ws2 = wb.create_sheet("Recaudos")
    ws2.append([
        "Fecha", "Cliente", "Documento", "Banco",
        "Cuenta", "Valor", "Referencia", "Estado"
    ])

    for celda in ws2[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="0F5C8E")

    for fila in recaudos:
        ws2.append(list(fila))

    wb.save(ruta)

    with conectar() as con:
        auditoria(con, "EXPORTAR CARTERA", ruta)
        con.commit()

    messagebox.showinfo(
        "Exportación",
        f"Reporte generado correctamente:\n\n{ruta}"
    )


def refrescar_todo():
    sincronizar_desde_ventas()
    cargar_bancos()
    cargar_cartera()
    cargar_recaudos()
    actualizar_kpis()


# ============================================================
# INTERFAZ
# ============================================================

inicializar_bd()

ventana = tk.Tk()
ventana.title("BME-ERP - Cuentas por Cobrar Integradas v1.0")
ventana.geometry("1500x900")
ventana.minsize(1180, 720)
ventana.configure(bg=C_FONDO)

try:
    ventana.state("zoomed")
except tk.TclError:
    pass

estilo = ttk.Style()
try:
    estilo.theme_use("clam")
except tk.TclError:
    pass

estilo.configure("Treeview", rowheight=28, font=("Segoe UI", 9))
estilo.configure("Treeview.Heading", font=("Segoe UI", 9, "bold"))

header = tk.Frame(ventana, bg=C_OSCURO, height=82)
header.pack(fill="x")
header.pack_propagate(False)

tk.Label(
    header,
    text="CUENTAS POR COBRAR INTEGRADAS",
    font=("Segoe UI", 22, "bold"),
    bg=C_OSCURO,
    fg="white"
).pack(side="left", padx=24, pady=20)

tk.Label(
    header,
    text="Cartera, recaudos, bancos y vencimientos",
    font=("Segoe UI", 10),
    bg=C_OSCURO,
    fg="#BFDBFE"
).pack(side="right", padx=24)

panel_kpi = tk.Frame(ventana, bg=C_FONDO)
panel_kpi.pack(fill="x", padx=18, pady=(14, 5))

for c in range(4):
    panel_kpi.grid_columnconfigure(c, weight=1)


def crear_kpi(columna, titulo, color):
    marco = tk.Frame(
        panel_kpi,
        bg=C_BLANCO,
        highlightbackground=C_BORDE,
        highlightthickness=1
    )
    marco.grid(row=0, column=columna, sticky="ew", padx=6)

    tk.Frame(marco, bg=color, width=5).pack(side="left", fill="y")

    interior = tk.Frame(marco, bg=C_BLANCO)
    interior.pack(fill="both", expand=True, padx=14, pady=10)

    tk.Label(
        interior,
        text=titulo,
        bg=C_BLANCO,
        fg=C_SUAVE,
        font=("Segoe UI", 8, "bold")
    ).pack(anchor="w")

    valor = tk.Label(
        interior,
        text="$0",
        bg=C_BLANCO,
        fg=C_TEXTO,
        font=("Segoe UI", 16, "bold")
    )
    valor.pack(anchor="w", pady=(3, 0))
    return valor


lbl_kpi_total = crear_kpi(0, "CARTERA TOTAL", C_AZUL)
lbl_kpi_vencida = crear_kpi(1, "CARTERA VENCIDA", C_ROJO)
lbl_kpi_recaudado = crear_kpi(2, "RECAUDADO DEL MES", C_VERDE)
lbl_kpi_clientes = crear_kpi(3, "CLIENTES CON SALDO", C_MORADO)

notebook = ttk.Notebook(ventana)
notebook.pack(fill="both", expand=True, padx=18, pady=10)

tab_cartera = tk.Frame(notebook, bg=C_FONDO)
tab_recaudos = tk.Frame(notebook, bg=C_FONDO)
tab_historial = tk.Frame(notebook, bg=C_FONDO)

notebook.add(tab_cartera, text="  Cartera  ")
notebook.add(tab_recaudos, text="  Registrar recaudo  ")
notebook.add(tab_historial, text="  Historial de recaudos  ")

# CARTERA
barra = tk.Frame(tab_cartera, bg=C_BLANCO)
barra.pack(fill="x", padx=10, pady=10)

tk.Label(
    barra, text="Buscar:", bg=C_BLANCO, fg=C_TEXTO
).pack(side="left", padx=(12, 5), pady=10)

entry_buscar = ttk.Entry(barra, width=30)
entry_buscar.pack(side="left", padx=5)

combo_estado = ttk.Combobox(
    barra,
    values=["TODOS", "PENDIENTE", "PARCIAL", "VENCIDA", "PAGADA", "ANULADA"],
    state="readonly",
    width=14
)
combo_estado.pack(side="left", padx=5)
combo_estado.set("TODOS")

tk.Button(
    barra,
    text="Actualizar",
    command=cargar_cartera,
    bg=C_AZUL,
    fg="white",
    relief="flat",
    padx=14,
    pady=6
).pack(side="left", padx=5)

tk.Button(
    barra,
    text="Exportar Excel",
    command=exportar_excel,
    bg=C_VERDE,
    fg="white",
    relief="flat",
    padx=14,
    pady=6
).pack(side="right", padx=8)

tk.Button(
    barra,
    text="Estado de cuenta",
    command=ver_estado_cuenta,
    bg=C_OSCURO,
    fg="white",
    relief="flat",
    padx=14,
    pady=6
).pack(side="right", padx=5)

cols_cartera = (
    "Factura", "Fecha", "Cliente", "Concepto",
    "Valor", "Saldo", "Vencimiento", "DiasMora", "Estado"
)

tabla_cartera = ttk.Treeview(
    tab_cartera,
    columns=cols_cartera,
    show="headings"
)

for col in cols_cartera:
    tabla_cartera.heading(col, text=col)

tabla_cartera.pack(fill="both", expand=True, padx=10, pady=(0, 10))

tabla_cartera.tag_configure("pendiente", foreground=C_NARANJA)
tabla_cartera.tag_configure("parcial", foreground=C_AZUL)
tabla_cartera.tag_configure("vencida", foreground=C_ROJO)
tabla_cartera.tag_configure("pagada", foreground=C_VERDE)
tabla_cartera.tag_configure("anulada", foreground=C_SUAVE)

# REGISTRAR RECAUDO
form = tk.LabelFrame(
    tab_recaudos,
    text="REGISTRO DE RECAUDO",
    bg=C_BLANCO,
    fg=C_TEXTO,
    font=("Segoe UI", 10, "bold"),
    padx=12,
    pady=10
)
form.pack(fill="x", padx=10, pady=10)

for c in range(4):
    form.grid_columnconfigure(c, weight=1)

for texto, col in [
    ("Cuenta bancaria", 0),
    ("Valor", 1),
    ("Referencia", 2),
    ("Observaciones", 3)
]:
    tk.Label(
        form, text=texto, bg=C_BLANCO, fg=C_SUAVE
    ).grid(row=0, column=col, sticky="w")

combo_banco = ttk.Combobox(form, state="readonly")
combo_banco.grid(row=1, column=0, sticky="ew", padx=(0, 8))

entry_valor_recaudo = ttk.Entry(form)
entry_valor_recaudo.grid(row=1, column=1, sticky="ew", padx=(0, 8))

entry_referencia = ttk.Entry(form)
entry_referencia.grid(row=1, column=2, sticky="ew", padx=(0, 8))

txt_obs_recaudo = tk.Text(form, height=2, relief="solid", bd=1)
txt_obs_recaudo.grid(row=1, column=3, sticky="ew")

tk.Label(
    tab_recaudos,
    text="Seleccione primero la cuenta en la pestaña Cartera y luego registre el recaudo.",
    bg=C_FONDO,
    fg=C_SUAVE,
    font=("Segoe UI", 9, "italic")
).pack(anchor="w", padx=16, pady=(0, 8))

tk.Button(
    tab_recaudos,
    text="REGISTRAR RECAUDO E INGRESAR AL BANCO",
    command=registrar_recaudo,
    bg=C_VERDE,
    fg="white",
    relief="flat",
    font=("Segoe UI", 10, "bold"),
    padx=20,
    pady=10
).pack(anchor="w", padx=10)

# HISTORIAL
barra_h = tk.Frame(tab_historial, bg=C_BLANCO)
barra_h.pack(fill="x", padx=10, pady=10)

tk.Button(
    barra_h,
    text="Actualizar",
    command=cargar_recaudos,
    bg=C_AZUL,
    fg="white",
    relief="flat",
    padx=14,
    pady=6
).pack(side="left", padx=8, pady=8)

tk.Button(
    barra_h,
    text="Reversar recaudo",
    command=reversar_recaudo,
    bg=C_ROJO,
    fg="white",
    relief="flat",
    padx=14,
    pady=6
).pack(side="right", padx=8, pady=8)

cols_recaudos = (
    "Fecha", "Cliente", "Documento",
    "Banco", "Valor", "Referencia", "Estado"
)

tabla_recaudos = ttk.Treeview(
    tab_historial,
    columns=cols_recaudos,
    show="headings"
)

for col in cols_recaudos:
    tabla_recaudos.heading(col, text=col)

tabla_recaudos.pack(fill="both", expand=True, padx=10, pady=(0, 10))
tabla_recaudos.tag_configure("activo", foreground=C_VERDE)
tabla_recaudos.tag_configure("reversado", foreground=C_ROJO)

barra_estado = tk.Frame(ventana, bg=C_BLANCO, height=28)
barra_estado.pack(fill="x")

tk.Label(
    barra_estado,
    text=f"Base de datos: {RUTA_DB}",
    bg=C_BLANCO,
    fg=C_SUAVE,
    font=("Segoe UI", 8)
).pack(side="left", padx=12)

tk.Label(
    barra_estado,
    text="BME-ERP Cuentas por Cobrar Integradas v1.0",
    bg=C_BLANCO,
    fg=C_SUAVE,
    font=("Segoe UI", 8)
).pack(side="right", padx=12)

refrescar_todo()
ventana.mainloop()
