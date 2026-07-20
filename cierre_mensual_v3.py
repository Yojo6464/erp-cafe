import tkinter as tk
import sqlite3

from tkinter import messagebox
from openpyxl import Workbook
from datetime import datetime

# =====================================
# BASE DE DATOS
# =====================================

RUTA_DB = r"C:\Users\jrive\visual\erp_cafe.db"

# =====================================
# FUNCION VALOR
# =====================================

def valor(sql):

    try:

        con = sqlite3.connect(RUTA_DB)
        cur = con.cursor()

        cur.execute(sql)

        dato = cur.fetchone()

        con.close()

        if dato is None:
            return 0

        if dato[0] is None:
            return 0

        return dato[0]

    except:

        return 0

# =====================================
# ACTUALIZAR
# =====================================

def actualizar():

    ventas = valor(
        "SELECT IFNULL(SUM(total),0) FROM ventas_cafe"
    )

    compras = valor(
        "SELECT IFNULL(SUM(total),0) FROM compras"
    )

    produccion = valor(
        "SELECT IFNULL(SUM(cafe_tostado),0) FROM produccion_cafe"
    )

    cartera = valor(
        "SELECT IFNULL(SUM(valor),0) FROM cuentas_cobrar_v1 WHERE estado='Pendiente'"
    )

    cxp = valor(
        "SELECT IFNULL(SUM(saldo),0) FROM cuentas_pagar WHERE estado='Pendiente'"
    )

    bancos = valor(
        "SELECT IFNULL(SUM(saldo),0) FROM bancos"
    )

    nomina = valor(
        "SELECT IFNULL(SUM(neto_pagar),0) FROM nomina"
    )
    gastos = valor(
    "SELECT IFNULL(SUM(valor),0) FROM gastos_operativos"
    )

    prestaciones = valor(
        "SELECT IFNULL(SUM(total_prestaciones),0) FROM prestaciones"
    )
   

    utilidad = (
    ventas
    - compras
    - nomina
    - gastos
)

    lbl_ventas.config(text=f"${ventas:,.0f}")
    lbl_compras.config(text=f"${compras:,.0f}")
    lbl_produccion.config(text=f"{produccion:,.0f} Kg")
    lbl_cartera.config(text=f"${cartera:,.0f}")
    lbl_cxp.config(text=f"${cxp:,.0f}")
    lbl_bancos.config(text=f"${bancos:,.0f}")
    lbl_gastos.config(
    text=f"${gastos:,.0f}"
)
    lbl_nomina.config(text=f"${nomina:,.0f}")
    lbl_prestaciones.config(text=f"${prestaciones:,.0f}")
   
    if utilidad >= 0:

     lbl_utilidad.config(
        text=f"${utilidad:,.0f}",
        fg="green"
    )

    else:

     lbl_utilidad.config(
        text=f"${utilidad:,.0f}",
        fg="red"
    )
 # =====================================
# EXPORTAR EXCEL
# =====================================

def exportar_excel():

    try:

        ventas = valor(
            "SELECT IFNULL(SUM(total),0) FROM ventas_cafe"
        )

        compras = valor(
            "SELECT IFNULL(SUM(total),0) FROM compras"
        )

        produccion = valor(
            "SELECT IFNULL(SUM(cafe_tostado),0) FROM produccion_cafe"
        )

        cartera = valor(
            "SELECT IFNULL(SUM(valor),0) FROM cuentas_cobrar_v1 WHERE estado='Pendiente'"
        )

        cxp = valor(
            "SELECT IFNULL(SUM(saldo),0) FROM cuentas_pagar WHERE estado='Pendiente'"
        )

        bancos = valor(
            "SELECT IFNULL(SUM(saldo),0) FROM bancos"
        )

        nomina = valor(
            "SELECT IFNULL(SUM(neto_pagar),0) FROM nomina"
        )

        prestaciones = valor(
            "SELECT IFNULL(SUM(total_prestaciones),0) FROM prestaciones"
        )

        utilidad = (
            ventas
            - compras
            - nomina
        )

        wb = Workbook()

        ws = wb.active

        ws.title = "Resumen"

        ws["A1"] = "CIERRE MENSUAL ERP CAFE"
        ws["A3"] = "Ventas"
        ws["B3"] = ventas

        ws["A4"] = "Compras"
        ws["B4"] = compras

        ws["A5"] = "Produccion Kg"
        ws["B5"] = produccion

        ws["A6"] = "Cartera"
        ws["B6"] = cartera

        ws["A7"] = "Cuentas por Pagar"
        ws["B7"] = cxp

        ws["A8"] = "Bancos"
        ws["B8"] = bancos

        ws["A9"] = "Nomina"
        ws["B9"] = nomina

        ws["A10"] = "Prestaciones"
        ws["B10"] = prestaciones

        ws["A11"] = "Utilidad Estimada"
        ws["B11"] = utilidad

        fecha = datetime.now().strftime("%Y_%m")

        archivo = (
            rf"C:\Users\jrive\visual\CIERRE_MENSUAL_{fecha}.xlsx"
        )

        wb.save(archivo)

        messagebox.showinfo(
            "Excel",
            f"Archivo generado:\n\n{archivo}"
        )

    except Exception as e:

        messagebox.showerror(
            "Error",
            str(e)
        )   




# =====================================
# VENTANA
# =====================================

ventana = tk.Tk()

ventana.title(
    "ERP Café Alto de la Cruz - Cierre Mensual V1"
)

ventana.geometry("900x650")

titulo = tk.Label(
    ventana,
    text="CIERRE MENSUAL",
    font=("Arial",22,"bold")
)

titulo.pack(pady=20)

frame = tk.Frame(ventana)
frame.pack(pady=20)

# =====================================
# INDICADORES
# =====================================

def fila(texto, fila_num):

    tk.Label(
        frame,
        text=texto,
        font=("Arial",12,"bold")
    ).grid(row=fila_num,column=0,padx=20,pady=8,sticky="w")

labels = {}

fila("Ventas",0)
lbl_ventas = tk.Label(
    frame,
    text="$0",
    fg="green"
)
lbl_ventas.grid(row=0,column=1)

fila("Compras",1)
lbl_compras = tk.Label(frame,text="$0")
lbl_compras.grid(row=1,column=1)

fila("Producción",2)
lbl_produccion = tk.Label(frame,text="0")
lbl_produccion.grid(row=2,column=1)

fila("Cartera",3)
lbl_cartera = tk.Label(frame,text="$0")
lbl_cartera.grid(row=3,column=1)

fila("Cuentas por Pagar",4)
lbl_cxp = tk.Label(frame,text="$0")
lbl_cxp.grid(row=4,column=1)

fila("Bancos",5)
lbl_bancos = tk.Label(
    frame,
    text="$0",
    fg="green"
)
lbl_bancos.grid(row=5,column=1)

fila("Nomina",6)
lbl_nomina = tk.Label(frame,text="$0")
lbl_nomina.grid(row=6,column=1)

fila("Gastos Operativos",7)
lbl_gastos = tk.Label(frame,text="$0")
lbl_gastos.grid(row=7,column=1)

fila("Prestaciones",8)
lbl_prestaciones = tk.Label(frame,text="$0")
lbl_prestaciones.grid(row=8,column=1)

fila("Utilidad Estimada",9)
lbl_utilidad = tk.Label(
    frame,
    text="$0",
    font=("Arial",12,"bold")
)
lbl_utilidad.grid(row=9,column=1)


btn = tk.Button(
    ventana,
    text="Actualizar Cierre",
    width=25,
    height=2,
    command=actualizar
)

btn.pack(pady=20)
btn_excel = tk.Button(
    ventana,
    text="Exportar a Excel",
    width=25,
    height=2,
    command=exportar_excel
)

btn_excel.pack(pady=10)

actualizar()

ventana.mainloop()